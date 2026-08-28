"""Importa vínculos, capacidades e snapshots manuais do QL a partir do painel.

Uso:
    venv/Scripts/python.exe scripts/import_ql_manual_snapshot.py caminho.xlsx

O processo é idempotente: os mesmos dias e vínculos podem ser reaplicados.
"""

from collections import defaultdict
from datetime import datetime
from pathlib import Path
import os
import sys
from unicodedata import normalize

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import create_engine, text


BACKEND_DIR = Path(__file__).resolve().parents[1]
load_dotenv(BACKEND_DIR / ".env")


def _normalized(value):
    raw = "".join(
        character
        for character in normalize("NFD", str(value or "").strip().upper())
        if not 0x300 <= ord(character) <= 0x36F
    )
    return " ".join(raw.split())


def _integer(value):
    if value in (None, ""):
        return None
    return int(round(float(value)))


def _company_id(value):
    name = _normalized(value)
    if name.startswith("COSTA OESTE"):
        return 1
    if "GRABIN" in name:
        return 4
    if "FACILITIES" in name:
        return 3
    if "MAG SUL" in name:
        return 8
    raise ValueError(f"Empresa da planilha não reconhecida: {value!r}")


def _records(workbook):
    sheet = workbook["ATIVOS"]
    for row in range(3, sheet.max_row + 1):
        company_name = sheet.cell(row, 1).value
        department = _integer(sheet.cell(row, 2).value)
        if not company_name or department is None or _normalized(company_name).startswith("TOTAL"):
            continue
        daily = {}
        for column in range(9, 32, 2):
            raw_date = sheet.cell(1, column).value
            day = raw_date.date() if isinstance(raw_date, datetime) else datetime(2026, 8, ((column - 9) // 2) + 1).date()
            daily[day] = _integer(sheet.cell(row, column).value) or 0
        yield {
            "empresa_id": _company_id(company_name),
            "departamento": department,
            "filial": _normalized(sheet.cell(row, 4).value),
            "capacidade": _integer(sheet.cell(row, 8).value) or 0,
            "diario": daily,
        }


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Informe o caminho da planilha XLSX.")
    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Planilha não encontrada: {workbook_path}")

    grouped = defaultdict(lambda: {"capacidade": 0, "diario": defaultdict(int)})
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    for record in _records(workbook):
        key = (record["empresa_id"], record["filial"], record["departamento"])
        grouped[key]["capacidade"] += record["capacidade"]
        for day, headcount in record["diario"].items():
            grouped[key]["diario"][day] += headcount

    engine = create_engine(os.environ["DB_URI"])
    linked_centers = snapshots = capacities = 0
    with engine.begin() as connection:
        connection.execute(text("""
            CREATE TABLE IF NOT EXISTS ql_capacidades_empresa (
                empresa_id INTEGER NOT NULL REFERENCES empresas(id) ON DELETE CASCADE,
                departamento INTEGER NOT NULL,
                capacidade_esperada INTEGER NOT NULL,
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                PRIMARY KEY (empresa_id, departamento)
            )
        """))
        connection.execute(text("""
            CREATE TABLE IF NOT EXISTS ql_historico_empresa_diario (
                id SERIAL PRIMARY KEY,
                data_referencia DATE NOT NULL,
                filial_id INTEGER NOT NULL REFERENCES filiais(id) ON DELETE CASCADE,
                empresa_id INTEGER NOT NULL REFERENCES empresas(id) ON DELETE CASCADE,
                departamento INTEGER NOT NULL,
                colaboradores_ativos INTEGER NOT NULL DEFAULT 0,
                capacidade_esperada INTEGER,
                centros_quantidade INTEGER NOT NULL DEFAULT 0,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                CONSTRAINT uq_ql_historico_empresa_diario_escopo
                    UNIQUE (data_referencia, filial_id, empresa_id, departamento)
            )
        """))
        connection.execute(text("CREATE INDEX IF NOT EXISTS ix_ql_hist_empresa_data ON ql_historico_empresa_diario (data_referencia)"))
        connection.execute(text("CREATE INDEX IF NOT EXISTS ix_ql_hist_empresa_filial ON ql_historico_empresa_diario (filial_id)"))

        branches = {
            _normalized(row.nome): row.id
            for row in connection.execute(text("SELECT id, nome FROM filiais"))
        }
        missing_branches = sorted({key[1] for key in grouped if key[1] not in branches})
        if missing_branches:
            raise RuntimeError(f"Filiais não cadastradas no TMHub: {', '.join(missing_branches)}")

        # A planilha enviada é a referência operacional destas filiais.
        connection.execute(
            text("UPDATE filiais SET ativa = TRUE WHERE id = ANY(:ids)"),
            {"ids": list({branches[key[1]] for key in grouped})},
        )

        by_company_department = defaultdict(int)
        for (company_id, branch_name, department), values in grouped.items():
            branch_id = branches[branch_name]
            center_ids = [
                row.id for row in connection.execute(
                    text("""
                        SELECT id FROM centro_de_custo
                        WHERE empresa_id = :empresa_id
                        AND (
                            departamento = :departamento
                            OR (
                                departamento IS NULL
                                AND FLOOR(centro_id / 1000.0)::INTEGER = :departamento
                            )
                        )
                    """),
                    {"empresa_id": company_id, "departamento": department},
                )
            ]
            if not center_ids:
                raise RuntimeError(
                    f"Nenhum centro encontrado para empresa {company_id}, DPTO {department}."
                )

            connection.execute(
                text("""
                    UPDATE centro_de_custo
                    SET departamento = :departamento
                    WHERE id = ANY(:ids) AND departamento IS NULL
                """),
                {"departamento": department, "ids": center_ids},
            )

            connection.execute(
                text("DELETE FROM filial_centros_custo WHERE centro_custo_id = ANY(:ids)"),
                {"ids": center_ids},
            )
            connection.execute(
                text("""
                    INSERT INTO filial_centros_custo (filial_id, centro_custo_id)
                    SELECT :filial_id, unnest(:ids)
                    ON CONFLICT DO NOTHING
                """),
                {"filial_id": branch_id, "ids": center_ids},
            )
            linked_centers += len(center_ids)
            by_company_department[(company_id, department)] += values["capacidade"]

            for day, headcount in values["diario"].items():
                connection.execute(text("""
                    INSERT INTO ql_historico_empresa_diario (
                        data_referencia, filial_id, empresa_id, departamento,
                        colaboradores_ativos, capacidade_esperada, centros_quantidade
                    ) VALUES (
                        :data, :filial_id, :empresa_id, :departamento,
                        :ativos, :capacidade, :centros
                    )
                    ON CONFLICT (data_referencia, filial_id, empresa_id, departamento)
                    DO UPDATE SET
                        colaboradores_ativos = EXCLUDED.colaboradores_ativos,
                        capacidade_esperada = EXCLUDED.capacidade_esperada,
                        centros_quantidade = EXCLUDED.centros_quantidade,
                        updated_at = NOW()
                """), {
                    "data": day,
                    "filial_id": branch_id,
                    "empresa_id": company_id,
                    "departamento": department,
                    "ativos": headcount,
                    "capacidade": values["capacidade"],
                    "centros": len(center_ids),
                })
                snapshots += 1

        for (company_id, department), capacity in by_company_department.items():
            connection.execute(text("""
                INSERT INTO ql_capacidades_empresa (empresa_id, departamento, capacidade_esperada)
                VALUES (:empresa_id, :departamento, :capacidade)
                ON CONFLICT (empresa_id, departamento) DO UPDATE SET
                    capacidade_esperada = EXCLUDED.capacidade_esperada,
                    updated_at = NOW()
            """), {"empresa_id": company_id, "departamento": department, "capacidade": capacity})
            capacities += 1

    print(
        f"Importação concluída: {len(grouped)} grupos, {linked_centers} centros vinculados, "
        f"{capacities} capacidades e {snapshots} snapshots de 01 a 12/08/2026."
    )


if __name__ == "__main__":
    main()

"""Importa metas de QL por departamento a partir da aba ATIVOS de uma planilha.

Uso:
    python scripts/importar_metas_ql.py "C:\\caminho\\Painel.xlsx"

O script soma os "POSTOS AUTORIZADOS O.S + RT" por DPTO, cria somente as
filiais explicitamente informadas na planilha e não cria vínculos de
departamento/centro de custo com filiais.
"""

import argparse
import os
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

# Permite executar o arquivo diretamente a partir da pasta scripts/.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import app
from models.centros_de_custo import DepartmentConfiguration
from models.filiais import Branch
from services.dashboard_ql import QLDashboardService
from utils.db import db
from utils.socket import socketio


BRANCH_NAMES = {
    "BELO HORIZONTE": "Belo Horizonte",
    "CASCAVEL": "Cascavel",
    "CUIABA": "Cuiabá",
    "CURITIBA": "Curitiba",
    "LONDRINA": "Londrina",
    "OESTE": "Oeste",
    "SC": "SC",
    "TOLEDO": "Toledo",
}


def normalize(value):
    text = str(value or "").strip()
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    ).upper()


def as_integer(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def source_data(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["ATIVOS"]
    except KeyError as error:
        raise ValueError("A planilha precisa conter a aba 'ATIVOS'.") from error

    header_row = None
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        normalized = [normalize(value) for value in row]
        if "DPTO" in normalized and "FILIAL" in normalized:
            header_row = row_index
            headers = normalized
            break
    if not header_row:
        raise ValueError("Não foi possível localizar o cabeçalho da aba ATIVOS.")

    try:
        department_index = headers.index("DPTO")
        branch_index = headers.index("FILIAL")
        target_index = next(
            index
            for index, value in enumerate(headers)
            if "POSTOS AUTORIZADOS" in value and "O.S" in value
        )
    except (StopIteration, ValueError) as error:
        raise ValueError(
            "Não encontrei as colunas DPTO, Filial e Postos Autorizados O.S + RT."
        ) from error

    targets = defaultdict(int)
    branches = set()
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        department = as_integer(row[department_index] if department_index < len(row) else None)
        target = as_integer(row[target_index] if target_index < len(row) else None)
        branch = normalize(row[branch_index] if branch_index < len(row) else None)
        if department is None or target is None:
            continue
        targets[department] += target
        if branch:
            branches.add(branch)
    return dict(targets), branches


def import_data(path):
    targets, source_branches = source_data(path)
    if not targets:
        raise ValueError("Nenhuma meta válida foi encontrada na aba ATIVOS.")

    existing_branches = {
        normalize(branch.nome): branch
        for branch in Branch.query.all()
    }
    created_branches = []
    for source_branch in sorted(source_branches):
        if source_branch in existing_branches:
            continue
        branch = Branch(nome=BRANCH_NAMES.get(source_branch, source_branch.title()), ativa=True)
        db.session.add(branch)
        created_branches.append(branch.nome)

    updated_departments = []
    for department, target in sorted(targets.items()):
        configuration = db.session.get(DepartmentConfiguration, department)
        if not configuration:
            configuration = DepartmentConfiguration(departamento=department, ativo=True)
            db.session.add(configuration)
        configuration.capacidade_pessoas = target
        updated_departments.append(department)

    db.session.commit()
    QLDashboardService.capture_daily()
    socketio.emit("ql_update", {"action": "planning_imported"})
    return {
        "departamentos_atualizados": len(updated_departments),
        "filiais_criadas": created_branches,
        "filiais_ja_existentes": sorted(
            BRANCH_NAMES.get(branch, branch.title())
            for branch in source_branches
            if branch in existing_branches
        ),
        "vinculos_departamento_filial": 0,
    }


def main():
    parser = argparse.ArgumentParser(description="Importa metas de QL por departamento.")
    parser.add_argument("arquivo", type=Path, help="Caminho da planilha .xlsx")
    args = parser.parse_args()
    if not args.arquivo.is_file():
        raise FileNotFoundError(f"Planilha não encontrada: {args.arquivo}")

    with app.app_context():
        try:
            summary = import_data(args.arquivo)
        except Exception:
            db.session.rollback()
            raise

    print("Metas de QL importadas com sucesso.")
    print(f"Departamentos atualizados: {summary['departamentos_atualizados']}")
    print(f"Filiais criadas: {', '.join(summary['filiais_criadas']) or 'nenhuma'}")
    print(f"Filiais já existentes: {', '.join(summary['filiais_ja_existentes']) or 'nenhuma'}")
    print("Vínculos de departamento com filial: não alterados")
    return 0


if __name__ == "__main__":
    status = main()
    sys.stdout.flush()
    os._exit(status)

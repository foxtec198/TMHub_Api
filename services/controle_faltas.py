from datetime import date, datetime as dt, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from unicodedata import normalize
from zoneinfo import ZoneInfo

from dateutil import relativedelta
from flask import jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import String, case, cast, or_, extract
from sqlalchemy.orm import aliased

from models.centros_de_custo import CostCenters
from models.colaboradores import Employees
from models.controle_faltas import AbsenceControl
from models.rp_historico import History
from models.rp_requisicao import Requisicao
from models.rp_timeline import Timeline
from models.supervisores import Supervisors
from models.usuarios import Users
from utils.db import db
from utils.filial_scope import apply_cost_center_scope, can_access_cost_center, can_access_supervisor, is_admin
from utils.permissions import has_permission
from utils.safe_route import safe_route
from utils.socket import socketio

SAO_PAULO = ZoneInfo("America/Sao_Paulo")
NON_ABSENCE_REASON_TERMS = ("REMANEJAMENTO", "FERIAS", "POSTO VAGO", "AFASTAMENTO")
DECLARATION_PARTIAL_HOURS = Decimal("4")


class AbsenceControlService:
    @staticmethod
    def _excel_datetime(value):
        """Converte datetimes com fuso para o formato aceito pelo Excel."""
        if isinstance(value, dt) and value.tzinfo is not None:
            return value.astimezone(SAO_PAULO).replace(tzinfo=None)
        return value

    @staticmethod
    def _normalized_reason(reason):
        raw = str(reason or "").strip()
        return "".join(
            character
            for character in normalize("NFD", raw)
            if ord(character) < 0x300 or ord(character) > 0x36F
        ).upper()

    @classmethod
    def _requires_document_deadline(cls, reason):
        normalized = cls._normalized_reason(reason)
        return "ATESTADO" in normalized or "DECLARACAO" in normalized

    @classmethod
    def _is_declaration(cls, reason):
        return "DECLARACAO" in cls._normalized_reason(reason)
    
    @staticmethod
    def _is_historical(value):
        if not value:
            return False
        local_value = value.astimezone(SAO_PAULO) if value.tzinfo else value.replace(tzinfo=SAO_PAULO)
        return local_value.date() < dt.now(SAO_PAULO).date()

    @staticmethod
    def _mark_historical_as_treated(absence):
        now = dt.now(SAO_PAULO)
        absence.status = "tratada"
        absence.prazo_atestado = None
        absence.tratado_por_usuario_id = None
        absence.tratado_em = now
        absence.automatizado_em = now

    @staticmethod
    def _can_manage(token_data):
        if is_admin(token_data):
            return True
        user = db.session.get(Users, (token_data or {}).get("id"))
        return bool(user and user.gerencia_faltas)

    @classmethod
    def _initial_classification(cls, reason):
        normalized = cls._normalized_reason(reason)
        if normalized == "INJUSTIFICADA":
            return "injustificada"
        if "ATESTADO" in normalized or "AFASTAMENTO" in normalized or "DECLARACAO" in normalized:
            return "justificada"
        return "em_analise"

    @classmethod
    def _is_absence_reason(cls, reason):
        normalized = cls._normalized_reason(reason)
        return not any(term in normalized for term in NON_ABSENCE_REASON_TERMS)

    @staticmethod
    def _deadline(req):
        if not AbsenceControlService._requires_document_deadline(req.motivo):
            return None
        opened = req.opened_at or dt.now(SAO_PAULO)
        if opened.tzinfo is None:
            opened = opened.replace(tzinfo=SAO_PAULO)
        scheduled = req.created_at
        if scheduled and scheduled.tzinfo is None:
            scheduled = scheduled.replace(tzinfo=SAO_PAULO)
        reference = max(opened, scheduled) if scheduled else opened
        return reference + timedelta(hours=48)

    @classmethod
    def ensure_for_request(cls, req):
        # Férias, remanejamento e posto vago não representam uma falta. A
        # requisição operacional continua existindo, sem criar tratativa.
        if not cls._is_absence_reason(req.motivo):
            absence = AbsenceControl.query.filter_by(requisicao_id=req.id).first()
            if absence:
                db.session.delete(absence)
            return None

        with db.session.no_autoflush:
            absence = AbsenceControl.query.filter_by(requisicao_id=req.id).first()
            employee = db.session.get(Employees, req.ausente_id)
        is_new = absence is None
        if not absence:
            absence = AbsenceControl(requisicao_id=req.id)
            db.session.add(absence)
        absence.colaborador_id = employee.id if employee else None
        absence.colaborador_nome = employee.nome if employee else "Colaborador não encontrado"
        absence.colaborador_matricula = employee.matricula if employee else None
        absence.centro_custo_id = req.cc
        absence.supervisor_id = req.supervisor_id
        absence.motivo = req.motivo
        if cls._is_declaration(req.motivo):
            absence.tipo_ausencia = "parcial"
            if absence.quantidade_horas is None:
                absence.quantidade_horas = DECLARATION_PARTIAL_HOURS
        else:
            absence.tipo_ausencia = "integral"
            absence.quantidade_horas = None
        absence.data_falta = req.created_at
        if is_new and req.obs:
            absence.observacao = req.obs
        if absence.status != "tratada":
            absence.classificacao = cls._initial_classification(req.motivo)
            if is_new and cls._is_historical(req.created_at):
                cls._mark_historical_as_treated(absence)
            else:
                absence.prazo_atestado = cls._deadline(req)
        if absence.status == "tratada":
            from services.glosas import DisallowanceService

            DisallowanceService.ensure_for_absence(absence)
        return absence

    @staticmethod
    def _expire_certificates():
        now = dt.now(SAO_PAULO)
        expired = AbsenceControl.query.filter(
            AbsenceControl.status == "pendente",
            or_(
                db.func.upper(AbsenceControl.motivo).like("%ATESTADO%"),
                db.func.upper(AbsenceControl.motivo).like("%DECLARA%"),
            ),
            AbsenceControl.prazo_atestado.isnot(None),
            AbsenceControl.prazo_atestado <= now,
            AbsenceControl.classificacao != "injustificada",
        ).all()
        for absence in expired:
            absence.classificacao = "injustificada"
            absence.automatizado_em = now
        if expired:
            db.session.commit()

    # @safe_route
    def total(self):
        month = request.args.get("month")

        if month == 'last': month = dt.now().month - 1
        if not month: month = dt.now().month
        year = dt.now().year

        query = AbsenceControl.query.filter(
            extract("month", AbsenceControl.created_at) == month,
            extract("year", AbsenceControl.created_at) == year,
        ).all()
        
        return jsonify(len([e for e in query]))
    
    @safe_route
    def read(self, token_data):
        if not has_permission(token_data, "controle_faltas", "view"):
            return jsonify("Você não possui acesso ao Controle de Faltas."), 403
        self._expire_certificates()
        Tratador = aliased(Users)
        query = (
            db.session.query(
                AbsenceControl.id,
                AbsenceControl.requisicao_id,
                AbsenceControl.data_falta,
                AbsenceControl.motivo,
                AbsenceControl.tipo_ausencia,
                AbsenceControl.quantidade_horas,
                AbsenceControl.prazo_atestado,
                AbsenceControl.classificacao,
                AbsenceControl.status,
                AbsenceControl.observacao,
                AbsenceControl.tratado_em,
                AbsenceControl.automatizado_em,
                db.func.coalesce(Employees.nome, AbsenceControl.colaborador_nome).label("colaborador"),
                db.func.coalesce(
                    cast(Employees.matricula, String),
                    AbsenceControl.colaborador_matricula,
                ).label("matricula"),
                CostCenters.local.label("contrato"),
                CostCenters.departamento,
                Supervisors.nome.label("supervisor"),
                Tratador.nome.label("tratado_por"),
                Requisicao.status.label("status_requisicao"),
            )
            .select_from(AbsenceControl)
            .outerjoin(Employees, Employees.id == AbsenceControl.colaborador_id)
            .join(CostCenters, CostCenters.id == AbsenceControl.centro_custo_id)
            .join(Supervisors, Supervisors.id == AbsenceControl.supervisor_id)
            .join(Requisicao, Requisicao.id == AbsenceControl.requisicao_id)
            .outerjoin(Tratador, Tratador.id == AbsenceControl.tratado_por_usuario_id)
            .filter(~db.func.upper(AbsenceControl.motivo).in_(NON_ABSENCE_REASON_TERMS))
            .order_by(
                case((AbsenceControl.status == "pendente", 0), else_=1),
                AbsenceControl.prazo_atestado.asc().nullslast(),
                AbsenceControl.data_falta.desc(),
            )
        )
        rows = apply_cost_center_scope(query, AbsenceControl.centro_custo_id, token_data).all()

        filter_options = {
            "departamentos": sorted(
                {str(row.departamento) for row in rows if row.departamento is not None},
                key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value),
            ),
            "supervisores": sorted({row.supervisor for row in rows if row.supervisor}),
            "motivos": sorted({row.motivo for row in rows if row.motivo}),
            "contratos": sorted({row.contrato for row in rows if row.contrato}),
            "colaboradores": sorted({row.colaborador for row in rows if row.colaborador}),
        }
        
        def selected_values(name):
            return {
                value.strip()
                for raw in request.args.getlist(name)
                for value in str(raw).split(",")
                if value.strip() and value.strip() != "__all__"
            }

        department = selected_values("departamento")
        supervisor = selected_values("supervisor")
        reason = selected_values("motivo")
        contract = selected_values("contrato")
        collaborator = selected_values("colaborador")
        status = selected_values("status")
        classification = selected_values("classificacao")
        rows = [
            row for row in rows
            if (not department or str(row.departamento) in department)
            and (not supervisor or row.supervisor in supervisor)
            and (not reason or row.motivo in reason)
            and (not contract or row.contrato in contract)
            and (not collaborator or row.colaborador in collaborator)
            and (not status or row.status in status)
            and (not classification or row.classificacao in classification)
        ]
        return jsonify([row._asdict() for row in rows]), 200

    @safe_route
    def create_manual(self, token_data):
        if not has_permission(token_data, "controle_faltas", "edit"):
            return jsonify("Você não possui permissão para lançar faltas manualmente."), 403

        body = request.get_json(silent=True) or {}
        try:
            employee_id = int(body.get("colaborador_id"))
            supervisor_id = int(body.get("supervisor_id"))
        except (TypeError, ValueError):
            return jsonify("Selecione o colaborador e o supervisor."), 400

        employee = db.session.get(Employees, employee_id)
        if not employee:
            return jsonify("Colaborador não encontrado."), 404
        if not employee.centro_id:
            return jsonify("O colaborador selecionado não possui contrato/local vinculado."), 400
        if not can_access_cost_center(token_data, employee.centro_id):
            return jsonify("Você não possui acesso à filial deste colaborador."), 403
        if not db.session.get(Supervisors, supervisor_id):
            return jsonify("Supervisor não encontrado."), 404
        if not can_access_supervisor(token_data, supervisor_id):
            return jsonify("Você não possui acesso à filial deste supervisor."), 403

        reason = str(body.get("motivo") or "").strip().upper()
        if not reason:
            return jsonify("Informe o motivo da falta."), 400
        if not self._is_absence_reason(reason):
            return jsonify("Remanejamento não deve ser lançado como falta."), 400

        absence_type = str(body.get("tipo_ausencia") or "").strip().lower()
        if absence_type not in {"integral", "parcial"}:
            return jsonify("Informe se a falta foi integral ou parcial."), 400

        try:
            absence_date = dt.fromisoformat(str(body.get("data_falta") or "").replace("Z", "+00:00"))
        except ValueError:
            return jsonify("Informe uma data válida para a falta."), 400
        if absence_date.tzinfo:
            absence_date = absence_date.astimezone(SAO_PAULO).replace(tzinfo=None)

        absence_hours = None
        if absence_type == "parcial":
            try:
                absence_hours = Decimal(str(body.get("quantidade_horas")).replace(",", ".")).quantize(Decimal("0.01"))
            except (InvalidOperation, TypeError, ValueError):
                return jsonify("Informe a quantidade de horas da falta parcial."), 400
            if absence_hours <= 0 or absence_hours >= 24:
                return jsonify("As horas da falta parcial devem ser maiores que zero e menores que 24."), 400

        observation = str(body.get("observacao") or "").strip()
        type_label = "PARCIAL" if absence_type == "parcial" else "INTEGRAL"
        coverage_note = f"FALTA {type_label}"
        if absence_hours is not None:
            formatted_hours = format(absence_hours, "f").rstrip("0").rstrip(".")
            coverage_note += f" DE {formatted_hours}H"
        request_observation = f"{coverage_note} · LANÇADA PELO CONTROLE DE FALTAS · SEM COBERTURA"
        if observation:
            request_observation += f" · {observation.upper()}"

        try:
            requisition = Requisicao(
                reserva_id=0,
                ausente_id=employee.id,
                cc=employee.centro_id,
                supervisor_id=supervisor_id,
                warning=False,
                motivo=reason,
                obs=request_observation,
                created_at=absence_date,
                opened_at=dt.now(SAO_PAULO),
                status="pending",
            )
            db.session.add(requisition)
            db.session.flush()

            absence = self.ensure_for_request(requisition)
            absence.tipo_ausencia = absence_type
            absence.quantidade_horas = absence_hours
            absence.observacao = observation or request_observation

            db.session.add(Timeline(
                requisicao_id=requisition.id,
                reserva_id=0,
                ausente_id=employee.id,
                cc=employee.centro_id,
                supervisor_id=supervisor_id,
                criado_por_usuario_id=token_data.get("id"),
                status="pending",
                tipo="Requisição criada através do Controle de Faltas",
                motivo=reason,
                obs=request_observation,
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise

        socketio.emit("absence_control_update", {"id": absence.id, "action": "created_manually"})
        socketio.emit("new_request")
        socketio.emit("new_history")
        socketio.emit("kds_update", {
            "action": "created_from_absence_control",
            "request_id": requisition.id,
            "status": requisition.status,
            "emitted_at": dt.now(SAO_PAULO).isoformat(),
        })
        return jsonify({
            "message": "Falta lançada e requisição sem cobertura criada.",
            "falta_id": absence.id,
            "requisicao_id": requisition.id,
        }), 201

    @safe_route
    def update(self, token_data):
        if not has_permission(token_data, "controle_faltas", "edit"):
            return jsonify("Você não possui permissão para alterar o Controle de Faltas."), 403
        body = request.get_json(silent=True) or {}
        absence = db.session.get(AbsenceControl, body.get("id"))
        if not absence:
            return jsonify("Registro de falta não encontrado."), 404
        if not can_access_cost_center(token_data, absence.centro_custo_id):
            return jsonify("Você não possui acesso à filial deste registro."), 403

        if "motivo" in body:
            reason = str(body.get("motivo") or "").strip().upper()
            if not reason:
                return jsonify("Informe o motivo."), 400
            if not self._is_absence_reason(reason):
                return jsonify("Remanejamento não é uma falta. Altere a requisição diretamente."), 400
            previous_reason = absence.motivo
            absence.motivo = reason
            req = db.session.get(Requisicao, absence.requisicao_id)
            if req:
                req.motivo = reason
                History.query.filter_by(requisicao_id=req.id).update(
                    {History.motivo: reason},
                    synchronize_session=False,
                )
                if previous_reason != reason:
                    db.session.add(Timeline(
                        requisicao_id=req.id,
                        reserva_id=req.reserva_id,
                        ausente_id=req.ausente_id,
                        cc=req.cc,
                        supervisor_id=req.supervisor_id,
                        alterado_por_usuario_id=token_data.get("id"),
                        status=req.status,
                        tipo="Motivo alterado no Controle de Faltas",
                        motivo=reason,
                        obs=f"Motivo alterado de {previous_reason or 'NÃO INFORMADO'} para {reason}.",
                    ))
            if absence.status != "tratada":
                absence.classificacao = self._initial_classification(reason)
                if req:
                    absence.prazo_atestado = self._deadline(req)
        if "data_falta" in body:
            try:
                value = dt.fromisoformat(str(body.get("data_falta")).replace("Z", "+00:00"))
            except ValueError:
                return jsonify("Data da falta inválida."), 400
            absence.data_falta = value
            req = db.session.get(Requisicao, absence.requisicao_id)
            if req:
                req.created_at = value
                if absence.status != "tratada" and self._is_historical(value):
                    absence.classificacao = self._initial_classification(absence.motivo)
                    self._mark_historical_as_treated(absence)
                elif absence.status != "tratada":
                    absence.prazo_atestado = self._deadline(req)
        if "observacao" in body:
            absence.observacao = str(body.get("observacao") or "").strip() or None

        glosa_created = False
        if body.get("status") == "tratada":
            classification = str(body.get("classificacao") or "").lower()
            if classification not in {"justificada", "injustificada"}:
                return jsonify("Informe se a falta foi justificada ou injustificada."), 400
            absence.status = "tratada"
            absence.classificacao = classification
            absence.tratado_por_usuario_id = token_data.get("id")
            absence.tratado_em = dt.now(SAO_PAULO)
        elif body.get("status") == "pendente":
            absence.status = "pendente"
            absence.tratado_por_usuario_id = None
            absence.tratado_em = None

        if absence.status == "tratada":
            from services.glosas import DisallowanceService

            _, glosa_created = DisallowanceService.ensure_for_absence(
                absence,
                user_id=token_data.get("id"),
            )
        db.session.commit()
        socketio.emit("absence_control_update", {"id": absence.id})
        if glosa_created:
            socketio.emit(
                "disallowance_update",
                {"falta_id": absence.id, "action": "created_preventively"},
            )
        socketio.emit("new_request")
        socketio.emit("new_history")
        socketio.emit("kds_update", {
            "action": "absence_updated",
            "request_id": absence.requisicao_id,
            "emitted_at": dt.now(SAO_PAULO).isoformat(),
        })
        return jsonify("Registro de falta atualizado."), 200

    @safe_route
    def export(self, token_data):
        # A exportação usa a mesma permissão de visualização da tela principal.
        if not has_permission(token_data, "controle_faltas", "view"):
            return jsonify("Você não possui acesso ao Controle de Faltas."), 403

        # Atualiza classificações vencidas antes de consultar os registros.
        self._expire_certificates()
        try:
            start = dt.fromisoformat(request.args["inicio"]) if request.args.get("inicio") else None
            end = dt.fromisoformat(request.args["fim"]) if request.args.get("fim") else None
        except ValueError:
            return jsonify("Período inválido."), 400

        if start and start.tzinfo is None:
            start = start.replace(tzinfo=SAO_PAULO)
        if end:
            if end.tzinfo is None:
                end = end.replace(tzinfo=SAO_PAULO)
            end = end.replace(hour=23, minute=59, second=59, microsecond=999999)

        # Mantém o usuário que tratou a falta separado dos demais vínculos.
        Tratador = aliased(Users)
        query = (
            db.session.query(
                AbsenceControl.data_falta,
                AbsenceControl.motivo,
                AbsenceControl.tipo_ausencia,
                AbsenceControl.quantidade_horas,
                AbsenceControl.prazo_atestado,
                AbsenceControl.classificacao,
                AbsenceControl.status,
                AbsenceControl.observacao,
                AbsenceControl.tratado_em,
                db.func.coalesce(Employees.nome, AbsenceControl.colaborador_nome).label("colaborador"),
                db.func.coalesce(
                    cast(Employees.matricula, String),
                    AbsenceControl.colaborador_matricula,
                ).label("matricula"),
                CostCenters.local.label("contrato"),
                CostCenters.departamento,
                Supervisors.nome.label("supervisor"),
                Tratador.nome.label("tratado_por"),
            )
            .select_from(AbsenceControl)
            .outerjoin(Employees, Employees.id == AbsenceControl.colaborador_id)
            .join(CostCenters, CostCenters.id == AbsenceControl.centro_custo_id)
            .join(Supervisors, Supervisors.id == AbsenceControl.supervisor_id)
            .outerjoin(Tratador, Tratador.id == AbsenceControl.tratado_por_usuario_id)
            .filter(~db.func.upper(AbsenceControl.motivo).in_(NON_ABSENCE_REASON_TERMS))
            .order_by(
                case((AbsenceControl.status == "pendente", 0), else_=1),
                AbsenceControl.prazo_atestado.asc().nullslast(),
                AbsenceControl.data_falta.desc(),
            )
        )
        if start:
            query = query.filter(AbsenceControl.data_falta >= start)
        if end:
            query = query.filter(AbsenceControl.data_falta <= end)
        rows = apply_cost_center_scope(query, AbsenceControl.centro_custo_id, token_data).all()

        # Aceita filtros repetidos ou valores separados por vírgula.
        def selected_values(name):
            return {
                value.strip()
                for raw in request.args.getlist(name)
                for value in str(raw).split(",")
                if value.strip() and value.strip() != "__all__"
            }

        departments = selected_values("departamento")
        supervisors = selected_values("supervisor")
        reasons = selected_values("motivo")
        contracts = selected_values("contrato")
        collaborators = selected_values("colaborador")
        statuses = selected_values("status")
        classifications = selected_values("classificacao")
        rows = [
            row for row in rows
            if (not departments or str(row.departamento) in departments)
            and (not supervisors or row.supervisor in supervisors)
            and (not reasons or row.motivo in reasons)
            and (not contracts or row.contrato in contracts)
            and (not collaborators or row.colaborador in collaborators)
            and (not statuses or row.status in statuses)
            and (not classifications or row.classificacao in classifications)
        ]

        # Os cartões do topo refletem apenas as linhas filtradas para exportação.
        summary = {
            "total": len(rows),
            "pendentes": sum(row.status == "pendente" for row in rows),
            "tratadas": sum(row.status == "tratada" for row in rows),
            "justificadas": sum(row.classificacao == "justificada" for row in rows),
            "injustificadas": sum(row.classificacao == "injustificada" for row in rows),
            "horas": sum(
                float(row.quantidade_horas or (8 if row.tipo_ausencia == "integral" else 0))
                for row in rows
            ),
        }

        # Cria uma planilha independente, sem grade visual, como a de Glosas.
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Controle de Faltas"
        sheet.sheet_view.showGridLines = False
        green, dark, light, red, amber, white = (
            "20A65A", "173925", "EAF6EF", "D64545", "D99000", "FFFFFF"
        )
        thin = Side(style="thin", color="DDE7E1")

        sheet.merge_cells("A1:M2")
        title = sheet["A1"]
        title.value = "CONTROLE DE FALTAS"
        title.font = Font(size=20, bold=True, color=white)
        title.fill = PatternFill("solid", fgColor=dark)
        title.alignment = Alignment(vertical="center", horizontal="left")

        # Exibe os indicadores principais antes da tabela detalhada.
        cards = [
            ("REGISTROS", summary["total"], dark),
            ("PENDENTES", summary["pendentes"], amber),
            ("TRATADAS", summary["tratadas"], green),
            ("JUSTIFICADAS", summary["justificadas"], "2E8B57"),
            ("INJUSTIFICADAS", summary["injustificadas"], red),
            ("HORAS AFASTADAS", summary["horas"], dark),
        ]
        for (label, value, color), start_column in zip(cards, (1, 3, 5, 7, 9, 12)):
            end_column = start_column + 1
            sheet.merge_cells(start_row=4, start_column=start_column, end_row=4, end_column=end_column)
            sheet.merge_cells(start_row=5, start_column=start_column, end_row=6, end_column=end_column)
            for row in range(4, 7):
                for column in range(start_column, end_column + 1):
                    cell = sheet.cell(row, column)
                    cell.fill = PatternFill("solid", fgColor=light)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            label_cell = sheet.cell(4, start_column, label)
            value_cell = sheet.cell(5, start_column, value)
            label_cell.font = Font(size=9, bold=True, color=color)
            value_cell.font = Font(size=14, bold=True, color=dark)
            label_cell.alignment = value_cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # As colunas priorizam identificação, classificação e andamento da tratativa.
        headers = [
            "Data da falta", "Colaborador", "Matrícula", "Departamento", "Contrato",
            "Supervisor", "Motivo", "Tipo de ausência", "Horas", "Classificação",
            "Status", "Prazo do documento", "Observação",
        ]
        header_row = 8
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(header_row, column, label)
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Traduz valores persistidos para os rótulos exibidos ao usuário.
        classification_labels = {
            "em_analise": "Em análise",
            "justificada": "Justificada",
            "injustificada": "Injustificada",
        }
        status_labels = {"pendente": "Pendente", "tratada": "Tratada"}
        type_labels = {"integral": "Integral", "parcial": "Parcial"}
        for row_index, row in enumerate(rows, header_row + 1):
            hours = float(row.quantidade_horas or (8 if row.tipo_ausencia == "integral" else 0))
            values = [
                self._excel_datetime(row.data_falta),
                row.colaborador,
                row.matricula,
                row.departamento,
                row.contrato,
                row.supervisor,
                row.motivo,
                type_labels.get(row.tipo_ausencia, row.tipo_ausencia),
                hours,
                classification_labels.get(row.classificacao, row.classificacao),
                status_labels.get(row.status, row.status),
                self._excel_datetime(row.prazo_atestado),
                row.observacao,
            ]
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_index, column, value)
                cell.fill = PatternFill("solid", fgColor="FFFFFF" if row_index % 2 else "F5F9F7")
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=column == 13)
            for column in (1, 12):
                sheet.cell(row_index, column).number_format = "dd/mm/yyyy hh:mm"
            sheet.cell(row_index, 9).number_format = "0.00"

        # Mantém o cabeçalho visível e permite filtrar qualquer coluna da tabela.
        sheet.freeze_panes = "A9"
        sheet.auto_filter.ref = f"A8:M{max(header_row, header_row + len(rows))}"
        widths = [18, 34, 14, 15, 40, 28, 24, 18, 12, 18, 14, 21, 48]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[5].height = 24
        sheet.row_dimensions[8].height = 32

        # Mantém o arquivo em memória para enviá-lo diretamente na resposta HTTP.
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"controle_faltas_{date.today().isoformat()}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @safe_route
    def dashboard(self, token_data):
        if not has_permission(token_data, "dashboard_faltas", "view"):
            return jsonify("Você não possui acesso ao Dashboard de Faltas."), 403

        try:
            start = (
                dt.fromisoformat(request.args.get("inicio"))
                if request.args.get("inicio")
                else dt.now(SAO_PAULO).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            )
            end = (
                dt.fromisoformat(request.args.get("fim"))
                if request.args.get("fim")
                else dt.now(SAO_PAULO)
            )
        except ValueError:
            return jsonify("Período inválido."), 400
        if start.tzinfo is None:
            start = start.replace(tzinfo=SAO_PAULO)
        if end.tzinfo is None:
            end = end.replace(tzinfo=SAO_PAULO)
        end = end.replace(hour=23, minute=59, second=59, microsecond=999999)

        query = (
            db.session.query(
                AbsenceControl.id,
                AbsenceControl.data_falta,
                AbsenceControl.motivo,
                AbsenceControl.classificacao,
                AbsenceControl.status,
                AbsenceControl.created_at,
                AbsenceControl.tratado_em,
                db.func.coalesce(Employees.nome, AbsenceControl.colaborador_nome).label("colaborador"),
                CostCenters.local.label("contrato"),
                CostCenters.departamento,
                Supervisors.nome.label("supervisor"),
            )
            .select_from(AbsenceControl)
            .outerjoin(Employees, Employees.id == AbsenceControl.colaborador_id)
            .join(CostCenters, CostCenters.id == AbsenceControl.centro_custo_id)
            .join(Supervisors, Supervisors.id == AbsenceControl.supervisor_id)
            .filter(~db.func.upper(AbsenceControl.motivo).in_(NON_ABSENCE_REASON_TERMS))
            .filter(AbsenceControl.data_falta.between(start, end))
        )
        rows = apply_cost_center_scope(query, AbsenceControl.centro_custo_id, token_data).all()

        filter_options = {
            "departamentos": sorted(
                {str(row.departamento) for row in rows if row.departamento is not None},
                key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value),
            ),
            "supervisores": sorted({row.supervisor for row in rows if row.supervisor}),
            "motivos": sorted({row.motivo for row in rows if row.motivo}),
            "contratos": sorted({row.contrato for row in rows if row.contrato}),
            "colaboradores": sorted({row.colaborador for row in rows if row.colaborador}),
        }

        # Aplica filtros adicionais (os filter_options são calculados antes para manter os dropdowns completos)
        def selected_values(name):
            return {
                value.strip()
                for raw in request.args.getlist(name)
                for value in str(raw).split(",")
                if value.strip() and value.strip() != "__all__"
            }

        department = selected_values("departamento")
        supervisor = selected_values("supervisor")
        reason = selected_values("motivo")
        contract = selected_values("contrato")
        collaborator = selected_values("colaborador")
        status = selected_values("status")
        classification = selected_values("classificacao")
        rows = [
            row for row in rows
            if (not department or str(row.departamento) in department)
            and (not supervisor or row.supervisor in supervisor)
            and (not reason or row.motivo in reason)
            and (not contract or row.contrato in contract)
            and (not collaborator or row.colaborador in collaborator)
            and (not status or row.status in status)
            and (not classification or row.classificacao in classification)
        ]

        # Opções procedurais: uma escolha reduz imediatamente as alternativas
        # disponíveis no próximo filtro, evitando combinações sem registros.
        filter_options = {
            "departamentos": sorted(
                {str(row.departamento) for row in rows if row.departamento is not None},
                key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value),
            ),
            "supervisores": sorted({row.supervisor for row in rows if row.supervisor}),
            "motivos": sorted({row.motivo for row in rows if row.motivo}),
            "contratos": sorted({row.contrato for row in rows if row.contrato}),
            "colaboradores": sorted({row.colaborador for row in rows if row.colaborador}),
        }

        def rank(field):
            counts = {}
            for row in rows:
                value = getattr(row, field) or "Não informado"
                counts[str(value)] = counts.get(str(value), 0) + 1
            ordered = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
            return [{"label": key, "total": total} for key, total in ordered]

        treatment_hours = []
        for row in rows:
            if row.tratado_em and row.created_at:
                treated = row.tratado_em if row.tratado_em.tzinfo else row.tratado_em.replace(tzinfo=SAO_PAULO)
                created = row.created_at if row.created_at.tzinfo else row.created_at.replace(tzinfo=SAO_PAULO)
                treatment_hours.append(max(0, (treated - created).total_seconds() / 3600))

        indicators = {
            "total": len(rows),
            "pendentes": sum(row.status == "pendente" for row in rows),
            "tratadas": sum(row.status == "tratada" for row in rows),
            "justificadas": sum(row.classificacao == "justificada" for row in rows),
            "injustificadas": sum(row.classificacao == "injustificada" for row in rows),
            "em_analise": sum(row.classificacao == "em_analise" for row in rows),
            "tempo_medio_tratativa_horas": round(sum(treatment_hours) / len(treatment_hours), 1) if treatment_hours else None,
        }
        recent = sorted(rows, key=lambda row: row.data_falta, reverse=True)[:20]
        return jsonify({
            "periodo": {"inicio": start.date().isoformat(), "fim": end.date().isoformat()},
            "indicadores": indicators,
            "motivos": rank("motivo"),
            "contratos": rank("contrato")[:15],
            "departamentos": rank("departamento"),
            "supervisores": rank("supervisor"),
            "filtros": {
                **filter_options,
                "colaboradores": [{"label": name, "value": name} for name in filter_options["colaboradores"]],
            },
            "recentes": [{
                "id": row.id,
                "data_falta": row.data_falta.isoformat() if row.data_falta else None,
                "colaborador": row.colaborador,
                "contrato": row.contrato,
                "motivo": row.motivo,
                "classificacao": row.classificacao,
                "status": row.status,
            } for row in recent],
        }), 200

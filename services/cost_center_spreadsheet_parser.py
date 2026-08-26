"""Parser do relatório corporativo de centros de custo em XLS/XLSX."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable


def _clean(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _code(value: Any) -> int | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        number = int(float(value))
        return number if number > 0 else None
    except (TypeError, ValueError):
        return None


def _rows(path: Path) -> Iterable[tuple[Any, ...]]:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                yield from sheet.iter_rows(values_only=True)
        finally:
            workbook.close()
        return
    if path.suffix.lower() != ".xls":
        raise ValueError("Formato não suportado. Envie .xls ou .xlsx.")
    try:
        from python_calamine import load_workbook
        workbook = load_workbook(path)
        for sheet_name in workbook.sheet_names:
            yield from (tuple(row) for row in workbook.get_sheet_by_name(sheet_name).to_python())
        return
    except (ImportError, OSError, ValueError, RuntimeError) as error:
        raise ValueError("Não foi possível abrir o XLS de centros de custo.") from error


def parse_cost_center_spreadsheet(path: str | Path) -> dict:
    source = Path(path)
    centers: dict[int, dict] = {}
    invalid: list[str] = []
    duplicates = 0
    for row_number, row in enumerate(_rows(source), start=1):
        code = _code(row[0] if row else None)
        name = _clean(row[4] if len(row) > 4 else None)
        if code is None:
            continue
        if not name:
            invalid.append(f"Linha {row_number}: centro {code} sem nome.")
            continue
        if code in centers:
            duplicates += 1
        centers[code] = {"centro_id": code, "nome": name.upper(), "local": name.upper()}
    return {"centers": list(centers.values()), "invalid": invalid, "duplicates": duplicates}

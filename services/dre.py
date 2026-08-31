"""Importação de fontes e cálculo do demonstrativo DRE de Londrina."""
# Biblioteca padrão.
from collections import defaultdict
from datetime import date, datetime as dt
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from pathlib import Path
import re
from unicodedata import normalize

# Dependências externas.
from flask import jsonify, request
from openpyxl import load_workbook
from sqlalchemy import func, or_

# Módulos internos da aplicação.
from models.centros_de_custo import CostCenters
from models.colaboradores import Employees
from models.dre import DreEntry, DreImport
from models.empresas import Company
from models.filiais import (
    Branch,
    filial_centros_custo,
    filial_departamentos,
    filial_usuarios,
)
from utils.db import db
from utils.filial_scope import allowed_cost_center_ids, is_admin, requested_branch_ids
from utils.permissions import has_permission
from utils.safe_route import safe_route


MAX_IMPORT_SIZE = 100 * 1024 * 1024
MAX_IMPORT_ROWS = 100_000
MONEY = Decimal("0.01")
DEFAULT_TAX_RATE = Decimal("0.09")
VA_DISCOUNT_RATE = Decimal("0.20")
ACTIVE_EMPLOYEE_STATUSES = {1, 9, 18}

IMPORT_TYPES = {"contratos", "historico_dre", "folha", "glosas", "custos"}
HISTORIC_DRE_COMPETENCES = {
    date(2026, 5, 1),
    date(2026, 6, 1),
    date(2026, 7, 1),
}
MANUAL_CATEGORIES = {
    "rob",
    "receita_prevista",
    "impostos",
    "encargos",
    "glosa_analitica",
    "salarios",
    "horas_extras",
    "valor_va",
    "valor_vt",
    "materiais",
    "uniformes",
    "epis",
    "combustivel",
    "locacao",
    "manutencao",
    "maquinario",
    "ajuste_custos",
}
COST_SHEETS = {
    "MATERIAIS": "materiais",
    "UNIFORMES": "uniformes",
    "EPIS": "epis",
    "EPI'S": "epis",
    "COMBUSTIVEL": "combustivel",
    "IMOBILIZADO": "locacao",
    "MANUTENCAO": "manutencao",
    "MAQUINAS - EQUIPAMENTOS": "maquinario",
}


def _normalise(value):
    """Padroniza texto para comparar abas e cabeçalhos sem depender de acentos."""
    raw = normalize("NFKD", str(value or ""))
    raw = raw.replace("º", "").replace("°", "").replace("�", "")
    raw = "".join(character for character in raw if not character.isspace() or character == " ")
    raw = "".join(character for character in raw if ord(character) < 0x300 or ord(character) > 0x36F)
    return re.sub(r"\s+", " ", raw).strip().upper()


def _as_decimal(value, label="Valor"):
    """Converte células financeiras para Decimal com duas casas."""
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(MONEY)
    try:
        raw = str(value).strip().replace("R$", "").replace(" ", "")
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        return Decimal(raw).quantize(MONEY)
    except (InvalidOperation, ValueError):
        raise ValueError(f"{label} inválido.") from None


def _as_int(value):
    """Converte códigos exportados pelo Excel, inclusive valores terminados em .0."""
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _as_date(value):
    """Aceita datas nativas do Excel e os formatos recebidos nas planilhas."""
    if isinstance(value, dt):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.strptime(str(value).strip(), pattern).date()
        except ValueError:
            continue
    return None


def _va_company_cost(gross_value):
    """Calcula o custo empresarial do VA após o desconto de 20%."""
    return (gross_value * (Decimal("1.00") - VA_DISCOUNT_RATE)).quantize(MONEY)


def _first_day(value):
    return date(value.year, value.month, 1)


def _month_from_text(value, year):
    """Converte competências como JULHO ou 2026 - 07 em uma data mensal."""
    raw = _normalise(value)
    if not raw:
        return None
    match = re.search(r"(20\d{2})\D+(0?[1-9]|1[0-2])", raw)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)

    months = {
        "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "ABRIL": 4,
        "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8,
        "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
    }
    month = next((number for name, number in months.items() if name in raw), None)
    return date(year, month, 1) if month else None


def _department_from_value(value):
    """Extrai o código do departamento e reconhece o administrativo de Londrina."""
    raw = _normalise(value)
    if "ADMINISTRAT" in raw and "LONDRINA" in raw:
        return 28
    match = re.search(r"(?:DEPARTAMENTO|DPTO|CT)\s*[-.:]*\s*(\d+)", raw)
    if not match:
        match = re.match(r"\s*(\d+)\b", raw)
    return int(match.group(1)) if match else None


def _cell(row, index):
    return row[index] if index is not None and index < len(row) else None


def _header_index(row):
    """Monta o índice por cabeçalho e elimina células vazias repetidas."""
    result = {}
    for index, value in enumerate(row):
        label = _normalise(value)
        if label and label not in result:
            result[label] = index
    return result


def _header_value(headers, *labels):
    """Localiza variações conhecidas de cabeçalhos exportados pelo Excel."""
    return next((headers[label] for label in labels if label in headers), None)


def _uploads_hash(uploads):
    """Gera uma assinatura única para um conjunto de planilhas de folha."""
    digest = sha256()
    for upload in sorted(uploads, key=lambda item: Path(item.filename or "").name.lower()):
        content = upload.read()
        upload.stream.seek(0)
        digest.update(Path(upload.filename or "").name.encode("utf-8"))
        digest.update(content)
    return digest.hexdigest()


class DreService:
    """Centraliza as fontes e o cálculo da DRE por filial e empresa."""

    @classmethod
    def _department_map(cls, branch, company):
        """Retorna um centro representativo para cada departamento do contexto."""
        rows = (
            db.session.query(CostCenters.departamento, func.min(CostCenters.id))
            .outerjoin(
                filial_centros_custo,
                filial_centros_custo.c.centro_custo_id == CostCenters.id,
            )
            .outerjoin(
                filial_departamentos,
                filial_departamentos.c.departamento == CostCenters.departamento,
            )
            .filter(
                CostCenters.departamento.is_not(None),
                CostCenters.empresa_id == company.id,
                or_(
                    filial_centros_custo.c.filial_id == branch.id,
                    filial_departamentos.c.filial_id == branch.id,
                ),
            )
            .group_by(CostCenters.departamento)
            .all()
        )
        return {
            department: center_id
            for department, center_id in rows
        }

    @classmethod
    def _allowed_departments(cls, token_data, branch, company):
        """Respeita simultaneamente o escopo da filial, empresa e usuário atual."""
        department_map = cls._department_map(branch, company)
        allowed_centers = allowed_cost_center_ids(token_data)
        if allowed_centers is None:
            return department_map

        allowed = {
            department: center_id
            for department, center_id in department_map.items()
            if center_id in allowed_centers
        }
        return allowed

    @classmethod
    def _historic_department_map(cls, token_data, branch, company, competence):
        """Recupera departamentos de uma competência histórica já importada na DRE."""
        query = (
            db.session.query(DreEntry.departamento, func.min(DreEntry.centro_custo_id))
            .join(DreImport, DreImport.id == DreEntry.importacao_id)
            .filter(
                DreImport.ativo.is_(True),
                DreImport.filial_id == branch.id,
                DreImport.empresa_id == company.id,
                DreEntry.empresa_id == company.id,
                DreEntry.competencia == competence,
                DreEntry.departamento.is_not(None),
                DreEntry.centro_custo_id.is_not(None),
            )
        )
        allowed_centers = allowed_cost_center_ids(token_data)
        if allowed_centers is not None:
            query = query.filter(DreEntry.centro_custo_id.in_(allowed_centers))

        return {
            department: center_id
            for department, center_id in query.group_by(DreEntry.departamento).all()
        }

    @staticmethod
    def _read_workbook(upload):
        """Lê somente XLSX nesta primeira versão para manter o importador previsível."""
        suffix = Path(upload.filename or "").suffix.lower()
        if suffix != ".xlsx":
            raise ValueError("Envie a planilha no formato .xlsx.")
        content = upload.read()
        upload.stream.seek(0)
        return load_workbook(BytesIO(content), read_only=True, data_only=True)

    @staticmethod
    def _source_key(source_type, sheet_name, line_number, category):
        raw = f"{source_type}|{sheet_name}|{line_number}|{category}"
        return sha256(raw.encode("utf-8")).hexdigest()

    @staticmethod
    def _entry(**values):
        """Uniformiza o contrato interno antes de gravar ou exibir a prévia."""
        values["valor"] = _as_decimal(values.get("valor"))
        values["descricao"] = str(values.get("descricao") or "").strip()[:500] or None
        values["fornecedor"] = str(values.get("fornecedor") or "").strip()[:255] or None
        values["documento"] = str(values.get("documento") or "").strip()[:100] or None
        values["ordem_compra"] = str(values.get("ordem_compra") or "").strip()[:100] or None
        values["contrato_codigo"] = str(values.get("contrato_codigo") or "").strip()[:40] or None
        return values

    @classmethod
    def _parse_contracts(cls, workbook, requested_competence, departments, branch):
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        header_row = next(
            (
                index
                for index, row in enumerate(rows)
                if _header_value(_header_index(row), "N CT/DPTO", "NO CT/DPTO") is not None
            ),
            None,
        )
        if header_row is None:
            raise ValueError("Não localizamos o cabeçalho 'Nº CT/DPTO' na planilha de contratos.")

        headers = _header_index(rows[header_row])
        contract_column = _header_value(headers, "N CT/DPTO", "NO CT/DPTO")
        required = ("FILIAL (PADRONIZADA)", "COMPETENCIA FAT.", "VALOR CONTRATO (R$)", "REALIZADO (R$)")
        if any(label not in headers for label in required):
            raise ValueError("A planilha de contratos não possui todas as colunas necessárias.")

        current_year = requested_competence.year
        result = []
        for line_number, row in enumerate(rows[header_row + 1:], start=header_row + 2):
            source_branch = _normalise(_cell(row, headers["FILIAL (PADRONIZADA)"]))
            if source_branch != _normalise(branch.nome):
                continue
            competence = _month_from_text(_cell(row, headers["COMPETENCIA FAT."]), current_year)
            if competence != requested_competence:
                continue
            department = _department_from_value(_cell(row, contract_column))
            if department not in departments:
                continue

            common = {
                "competencia": competence,
                "departamento": department,
                "centro_custo_id": departments[department],
                "contrato_codigo": _cell(row, contract_column),
                "descricao": _cell(row, headers.get("CLIENTE")),
            }
            fields = (
                ("receita_prevista", "VALOR CONTRATO (R$)"),
                ("rob", "REALIZADO (R$)"),
                ("outros_faturamentos", "OUTROS FATURAMENTOS (R$)"),
                ("repactuacao", "REPACTUACAO (R$)"),
                ("faturamento_fora_prazo", "FAT. FORA DE PRAZO (R$)"),
                ("glosa_faturamento", "GLOSA (R$)"),
            )
            for category, label in fields:
                if label not in headers:
                    continue
                value = _as_decimal(_cell(row, headers[label]), label)
                if category == "glosa_faturamento":
                    value = abs(value)
                result.append(cls._entry(
                    **common,
                    categoria=category,
                    valor=value,
                    chave_origem=cls._source_key("contratos", worksheet.title, line_number, category),
                ))
        return result

    @classmethod
    def _employee_map(cls, departments, company):
        """Vincula a matrícula da folha ao departamento da empresa selecionada."""
        rows = (
            db.session.query(Employees, CostCenters.departamento)
            .join(CostCenters, CostCenters.id == Employees.centro_id)
            .filter(
                CostCenters.departamento.in_(departments),
                CostCenters.empresa_id == company.id,
                Employees.empresa_id == company.id,
            )
            .all()
        )
        return {employee.matricula: (employee, department) for employee, department in rows}

    @classmethod
    def _department_snapshot_by_competence(cls, competence, departments, branch, company):
        """Recupera o departamento salvo no benefício da mesma competência."""
        rows = (
            db.session.query(
                Employees.matricula,
                DreEntry.departamento,
                DreEntry.centro_custo_id,
            )
            .join(Employees, Employees.id == DreEntry.colaborador_id)
            .join(DreImport, DreImport.id == DreEntry.importacao_id)
            .filter(
                DreImport.ativo.is_(True),
                DreImport.filial_id == branch.id,
                DreImport.empresa_id == company.id,
                DreImport.tipo.in_({"beneficios", "beneficios_tmh"}),
                DreEntry.competencia == competence,
                DreEntry.categoria == "valor_va",
                DreEntry.departamento.in_(departments),
            )
            .all()
        )
        return {
            registration: (department, center_id)
            for registration, department, center_id in rows
        }

    @classmethod
    def _parse_analytical_payroll(cls, workbook, requested_competence, employees, source_name):
        """Mantém compatibilidade com a Folha Analítica que possui Liq. Folha."""
        worksheet = workbook.active
        result = []
        headers = None
        liquid_column = None

        for line_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            labels = _header_index(row)
            liquid_header = _header_value(
                labels,
                "LIQ. FOLHA",
                "LIQ FOLHA",
                "LIQUIDO FOLHA",
            )
            if "MATRICULA" in labels and liquid_header is not None:
                headers = labels
                liquid_column = liquid_header
                continue
            if not headers:
                continue

            registration = _as_int(_cell(row, headers["MATRICULA"]))
            employee_context = employees.get(registration)
            if not employee_context:
                continue
            employee, department = employee_context
            # O líquido já consolida salário, adicionais e descontos que
            # compõem o pagamento real ao colaborador no mês de competência.
            value = _as_decimal(_cell(row, liquid_column), "Líquido da folha")
            result.append(cls._entry(
                competencia=requested_competence,
                departamento=department,
                centro_custo_id=employee.centro_id,
                colaborador_id=employee.id,
                categoria="salarios",
                valor=value,
                descricao=_cell(row, headers.get("EMPREGADO")),
                chave_origem=cls._source_key("folha", source_name, line_number, "salarios"),
            ))

        if liquid_column is None:
            raise ValueError("Não localizamos a coluna 'Liq. Folha' na Folha Analítica.")
        return result

    @classmethod
    def _parse_payroll_extract(cls, workbook, requested_competence, employees, source_name):
        """Calcula o líquido do Extrato de Faturamento: proventos menos descontos."""
        worksheet = workbook.active
        result = []
        current_registration = None
        employee_totals = defaultdict(Decimal)
        employee_names = {}

        for row in worksheet.iter_rows(values_only=True):
            registration = _as_int(_cell(row, 0))
            if registration is not None:
                current_registration = registration
                employee_names[registration] = _cell(row, 1)

            event_type = _normalise(_cell(row, 14))
            if current_registration is None or event_type not in {"P", "D"}:
                continue

            value = _as_decimal(_cell(row, 15), "Valor da rubrica")
            employee_totals[current_registration] += value if event_type == "P" else -value

        for registration, value in employee_totals.items():
            employee_context = employees.get(registration)
            if not employee_context:
                continue
            employee, department = employee_context
            result.append(cls._entry(
                competencia=requested_competence,
                departamento=department,
                centro_custo_id=employee.centro_id,
                colaborador_id=employee.id,
                categoria="salarios",
                valor=value,
                descricao=employee_names.get(registration) or employee.nome,
                chave_origem=cls._source_key("folha", source_name, registration, "salarios"),
            ))
        return result

    @classmethod
    def _parse_payroll(cls, uploads, requested_competence, departments, company):
        """Importa a folha do RH por extratos, sem duplicar matrículas entre arquivos."""
        employees = cls._employee_map(departments, company)
        entries = []
        registrations = set()

        for upload in uploads:
            workbook = cls._read_workbook(upload)
            worksheet = workbook.active
            is_analytical = any(
                "LIQ. FOLHA" in _header_index(row)
                for row in worksheet.iter_rows(values_only=True)
            )
            source_name = f"{Path(upload.filename or '').name}:{worksheet.title}"
            file_entries = (
                cls._parse_analytical_payroll(workbook, requested_competence, employees, source_name)
                if is_analytical
                else cls._parse_payroll_extract(workbook, requested_competence, employees, source_name)
            )
            for entry in file_entries:
                registration = entry["colaborador_id"]
                if registration in registrations:
                    raise ValueError("A mesma matrícula foi encontrada em mais de um extrato de folha.")
                registrations.add(registration)
            entries.extend(file_entries)

        if not entries:
            raise ValueError("Nenhum colaborador da empresa selecionada foi encontrado nos extratos de folha.")
        return entries

    @classmethod
    def _parse_historic_dre(cls, workbook, requested_competence, departments):
        """Importa somente as DREs históricas fechadas de maio a julho de 2026."""
        if requested_competence not in HISTORIC_DRE_COMPETENCES:
            raise ValueError("O importador histórico aceita somente maio, junho e julho de 2026.")
        worksheet = workbook["DRE"] if "DRE" in workbook.sheetnames else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        header_row = next(
            (
                index
                for index, row in enumerate(rows)
                if "CONTA (R$)" in _header_index(row)
            ),
            None,
        )
        if header_row is None:
            raise ValueError("Não localizamos a linha 'CONTA (R$)' na planilha histórica da DRE.")

        headers = rows[header_row]
        department_columns = {
            department: index
            for index, value in enumerate(headers)
            for department in (_department_from_value(value),)
            if department in departments
        }
        if not department_columns:
            raise ValueError("A planilha histórica não possui departamentos válidos da DRE de Londrina.")

        category_by_label = (
            ("RECEITA OPERACIONAL BRUTA", "rob"),
            ("RECEITA PREVISTA", "receita_prevista"),
            ("FATURAMENTO FORA", "faturamento_fora_prazo"),
            ("GLOSAS", "glosa_faturamento"),
            ("IMPOSTOS", "impostos"),
            ("HORAS EXTRAS", "horas_extras"),
            # A folha salarial é importada exclusivamente da coluna
            # "Liq. Folha" da Folha Analítica, por competência e departamento.
            # O total retroativo serve para os demais itens já fechados da DRE.
            ("ENCARGOS", "encargos"),
            ("VALE ALIMENTACAO", "valor_va"),
            ("VALE TRANSPORTE", "valor_vt"),
            ("MATERIAIS", "materiais"),
            ("UNIFORMES", "uniformes"),
            ("EPI", "epis"),
            ("COMBUSTIVEL", "combustivel"),
            ("LOCACAO", "locacao"),
            ("MANUTENCAO", "manutencao"),
            ("MAQUINARIO", "maquinario"),
        )
        result = []
        imported_categories = set()
        for line_number, row in enumerate(rows[header_row + 1:], start=header_row + 2):
            label = _normalise(_cell(row, 0))
            category = next(
                (value for token, value in category_by_label if token in label),
                None,
            )
            # A planilha possui linhas de conferência que repetem impostos e
            # margem. Cada categoria financeira é importada somente uma vez.
            if not category or category in imported_categories:
                continue
            imported_categories.add(category)
            for department, column in department_columns.items():
                result.append(cls._entry(
                    competencia=requested_competence,
                    departamento=department,
                    centro_custo_id=departments[department],
                    categoria=category,
                    valor=_as_decimal(_cell(row, column), label),
                    descricao=_cell(row, 0),
                    chave_origem=cls._source_key(
                        "historico_dre",
                        f"{worksheet.title}-{department}",
                        line_number,
                        category,
                    ),
                ))
        if not result:
            raise ValueError("Não localizamos valores financeiros na planilha histórica da DRE.")
        return result

    @classmethod
    def _parse_disallowances(cls, workbook, requested_competence, departments, branch):
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        header_row = next(
            (index for index, row in enumerate(rows) if "DPTO" in _header_index(row) and "FILIAL" in _header_index(row)),
            None,
        )
        if header_row is None:
            raise ValueError("Não localizamos os cabeçalhos DPTO e FILIAL na planilha de glosas.")

        headers = _header_index(rows[header_row])
        result = []
        for line_number, row in enumerate(rows[header_row + 1:], start=header_row + 2):
            source_branch = _normalise(_cell(row, headers.get("FILIAL")))
            if source_branch != _normalise(branch.nome):
                continue
            department = _as_int(_cell(row, headers.get("DPTO")))
            if department not in departments:
                continue

            competence = _as_date(_cell(row, headers.get("COMPETENCIA SERVICO PRESTADO")))
            competence = _first_day(competence) if competence else None
            if competence != requested_competence:
                continue

            value = _as_decimal(_cell(row, headers.get("VALOR QUE A FILIAL ENTENDE COMO PERDA")))
            if not value:
                value = _as_decimal(_cell(row, headers.get("PREJUIZO")))
            result.append(cls._entry(
                competencia=competence,
                departamento=department,
                centro_custo_id=departments[department],
                categoria="glosa_analitica",
                valor=abs(value),
                descricao=_cell(row, headers.get("OBS G.C")),
                documento=_cell(row, headers.get("ORDEM")),
                chave_origem=cls._source_key("glosas", worksheet.title, line_number, "glosa_analitica"),
            ))
        return result

    @classmethod
    def _parse_costs(cls, workbook, requested_competence, departments):
        result = []
        found_sheets = 0
        for worksheet in workbook.worksheets:
            category = COST_SHEETS.get(_normalise(worksheet.title))
            if not category:
                continue
            found_sheets += 1
            rows = worksheet.iter_rows(values_only=True)
            headers = _header_index(next(rows, ()))
            required = ("DATA ENT", "VLR TOTAL", "DEPARTAMENTO", "SITUACAO")
            if any(label not in headers for label in required):
                raise ValueError(f"A aba '{worksheet.title}' não possui as colunas de custo esperadas.")

            for line_number, row in enumerate(rows, start=2):
                competence_date = _as_date(_cell(row, headers["DATA ENT"]))
                if not competence_date or _first_day(competence_date) != requested_competence:
                    continue
                if _normalise(_cell(row, headers["SITUACAO"])) != "FECHADA":
                    continue
                department = _department_from_value(_cell(row, headers["DEPARTAMENTO"]))
                if department not in departments:
                    continue

                result.append(cls._entry(
                    competencia=requested_competence,
                    departamento=department,
                    centro_custo_id=departments[department],
                    categoria=category,
                    valor=_as_decimal(_cell(row, headers["VLR TOTAL"]), "Valor total"),
                    quantidade=_as_decimal(_cell(row, headers.get("QTD")), "Quantidade"),
                    documento=_cell(row, headers.get("N NF")),
                    descricao=_cell(row, headers.get("ITEM")),
                    fornecedor=_cell(row, headers.get("FORNECEDOR")),
                    ordem_compra=_cell(row, headers.get("ORDEM DE COMPRA")),
                    chave_origem=cls._source_key("custos", worksheet.title, line_number, category),
                ))
        if not found_sheets:
            raise ValueError("Não localizamos as abas de custos esperadas.")
        return result

    @classmethod
    def _parse(cls, source_type, uploads, requested_competence, departments, branch, company):
        if source_type == "folha":
            entries = cls._parse_payroll(uploads, requested_competence, departments, company)
            if len(entries) > MAX_IMPORT_ROWS:
                raise ValueError(f"As planilhas excedem o limite de {MAX_IMPORT_ROWS:,} lançamentos.")
            return entries

        if len(uploads) != 1:
            raise ValueError("Este tipo de fonte aceita somente uma planilha por importação.")
        workbook = cls._read_workbook(uploads[0])
        if source_type == "contratos":
            entries = cls._parse_contracts(workbook, requested_competence, departments, branch)
        elif source_type == "glosas":
            entries = cls._parse_disallowances(workbook, requested_competence, departments, branch)
        else:
            parser = {
                "historico_dre": cls._parse_historic_dre,
                "custos": cls._parse_costs,
            }[source_type]
            entries = parser(workbook, requested_competence, departments)
        if len(entries) > MAX_IMPORT_ROWS:
            raise ValueError(f"A planilha excede o limite de {MAX_IMPORT_ROWS:,} lançamentos.")
        return entries

    @classmethod
    def _current_benefit_entries(cls, competence, departments, company):
        """Gera a fotografia mensal de benefícios a partir dos cadastros do TMHub."""
        rows = (
            db.session.query(Employees, CostCenters.departamento)
            .join(CostCenters, CostCenters.id == Employees.centro_id)
            .filter(
                CostCenters.departamento.in_(departments),
                CostCenters.empresa_id == company.id,
                Employees.empresa_id == company.id,
                Employees.situacao.in_(ACTIVE_EMPLOYEE_STATUSES),
            )
            .all()
        )
        entries = []
        for employee, department in rows:
            gross_va = Decimal(str(employee.valor_va or 0)).quantize(MONEY)
            company_cost = _va_company_cost(gross_va)
            vt_value = Decimal(str(employee.valor_vt or 0)).quantize(MONEY)

            common = {
                "competencia": competence,
                "departamento": department,
                "centro_custo_id": employee.centro_id,
                "colaborador_id": employee.id,
                "descricao": employee.nome,
            }
            for category, value in (("valor_va", company_cost), ("valor_vt", vt_value)):
                entries.append(cls._entry(
                    **common,
                    categoria=category,
                    valor=value,
                    chave_origem=cls._source_key(
                        "beneficios_tmh",
                        str(employee.id),
                        competence.isoformat(),
                        category,
                    ),
                ))
        return entries

    @staticmethod
    def _parse_requested_competence():
        raw = str(request.form.get("competencia") or "").strip()
        try:
            return dt.strptime(raw, "%Y-%m").date().replace(day=1)
        except ValueError:
            raise ValueError("Informe a competência no formato aaaa-mm.") from None

    @classmethod
    def _preview(cls, source_type, uploads, competence, departments, branch, company):
        entries = cls._parse(
            source_type,
            uploads,
            competence,
            departments,
            branch,
            company,
        )
        by_category = defaultdict(Decimal)
        by_department = defaultdict(Decimal)
        for entry in entries:
            by_category[entry["categoria"]] += entry["valor"]
            by_department[entry["departamento"]] += entry["valor"]
        return entries, {
            "competencia": competence.isoformat(),
            "tipo": source_type,
            "lancamentos": len(entries),
            "departamentos": len(by_department),
            "categorias": [
                {"nome": category, "valor": float(value)}
                for category, value in sorted(by_category.items())
            ],
            "amostra": [
                {
                    "departamento": entry["departamento"],
                    "categoria": entry["categoria"],
                    "valor": float(entry["valor"]),
                    "descricao": entry.get("descricao"),
                }
                for entry in entries[:12]
            ],
        }

    @safe_route
    def preview_import(self, token_data):
        if not has_permission(token_data, "controle_dre", "create"):
            return jsonify("Você não possui permissão para importar fontes da DRE."), 403
        if not is_admin(token_data):
            return jsonify("A importação da DRE é permitida somente para administradores."), 403

        source_type = str(request.form.get("tipo") or "").strip().lower()
        uploads = [upload for upload in request.files.getlist("files") if upload.filename]
        if not uploads and request.files.get("file"):
            uploads = [request.files["file"]]
        if source_type not in IMPORT_TYPES:
            return jsonify("Tipo de importação inválido."), 400
        if not uploads:
            return jsonify("Selecione ao menos uma planilha para importar."), 400
        if request.content_length and request.content_length > MAX_IMPORT_SIZE:
            return jsonify("As planilhas devem somar até 100 MB."), 413
        try:
            competence = self._parse_requested_competence()
            branch, _, company, _ = self._selected_context(token_data)
            if not branch or not company:
                return jsonify("Selecione uma filial e empresa disponíveis para a DRE."), 404
            departments = self._allowed_departments(token_data, branch, company)
            _, preview = self._preview(
                source_type,
                uploads,
                competence,
                departments,
                branch,
                company,
            )
        except ValueError as error:
            return jsonify(str(error)), 400
        except Exception as error:
            return jsonify(f"Não foi possível ler a planilha: {error}"), 400
        return jsonify(preview), 200

    @safe_route
    def import_source(self, token_data):
        if not has_permission(token_data, "controle_dre", "create"):
            return jsonify("Você não possui permissão para importar fontes da DRE."), 403
        if not is_admin(token_data):
            return jsonify("A importação da DRE é permitida somente para administradores."), 403

        source_type = str(request.form.get("tipo") or "").strip().lower()
        uploads = [upload for upload in request.files.getlist("files") if upload.filename]
        if not uploads and request.files.get("file"):
            uploads = [request.files["file"]]
        if source_type not in IMPORT_TYPES:
            return jsonify("Tipo de importação inválido."), 400
        if not uploads:
            return jsonify("Selecione ao menos uma planilha para importar."), 400
        if request.content_length and request.content_length > MAX_IMPORT_SIZE:
            return jsonify("As planilhas devem somar até 100 MB."), 413
        try:
            competence = self._parse_requested_competence()
            branch, _, company, _ = self._selected_context(token_data)
            if not branch or not company:
                return jsonify("Selecione uma filial e empresa disponíveis para a DRE."), 404
            departments = self._allowed_departments(token_data, branch, company)
            entries, preview = self._preview(
                source_type,
                uploads,
                competence,
                departments,
                branch,
                company,
            )
            if not entries:
                return jsonify("Nenhum lançamento foi encontrado para a filial e empresa selecionadas."), 400
            source_hash = _uploads_hash(uploads)
        except ValueError as error:
            return jsonify(str(error)), 400
        except Exception as error:
            return jsonify(f"Não foi possível importar a planilha: {error}"), 400

        duplicate = DreImport.query.filter_by(
            tipo=source_type,
            competencia=competence,
            filial_id=branch.id,
            empresa_id=company.id,
            arquivo_hash=source_hash,
            ativo=True,
        ).first()
        if duplicate:
            return jsonify("Esta mesma planilha já está ativa para a competência selecionada."), 409

        # Um novo arquivo da mesma fonte substitui a versão anterior, mas a
        # mantém armazenada para auditoria e nunca mistura valores duplicados.
        DreImport.query.filter_by(
            tipo=source_type,
            competencia=competence,
            filial_id=branch.id,
            empresa_id=company.id,
            ativo=True,
        ).update({"ativo": False}, synchronize_session=False)

        imported = DreImport(
            tipo=source_type,
            competencia=competence,
            filial_id=branch.id,
            empresa_id=company.id,
            arquivo_original=" · ".join(Path(upload.filename).name for upload in uploads)[:255],
            arquivo_hash=source_hash,
            registros_lidos=preview["lancamentos"],
            registros_importados=len(entries),
            importado_por_usuario_id=token_data.get("id"),
        )
        db.session.add(imported)
        db.session.flush()
        for entry in entries:
            db.session.add(DreEntry(
                importacao_id=imported.id,
                empresa_id=company.id,
                criado_por_usuario_id=token_data.get("id"),
                **entry,
            ))
        db.session.commit()
        return jsonify({
            "message": "Fonte da DRE importada com sucesso.",
            **preview,
        }), 201

    @safe_route
    def generate_current_benefits(self, token_data):
        """Fecha os benefícios da competência sem depender de uma planilha externa."""
        if not has_permission(token_data, "controle_dre", "create"):
            return jsonify("Você não possui permissão para gerar benefícios da DRE."), 403
        if not is_admin(token_data):
            return jsonify("A geração de benefícios da DRE é permitida somente para administradores."), 403

        try:
            competence = self._parse_requested_competence()
            branch, _, company, _ = self._selected_context(token_data)
            if not branch or not company:
                return jsonify("Selecione uma filial e empresa disponíveis para a DRE."), 404
            departments = self._allowed_departments(token_data, branch, company)
            entries = self._current_benefit_entries(competence, departments, company)
            if not entries:
                return jsonify("Nenhum colaborador ativo com benefício foi encontrado no contexto selecionado."), 400
        except ValueError as error:
            return jsonify(str(error)), 400
        except Exception as error:
            return jsonify(f"Não foi possível gerar os benefícios: {error}"), 400

        DreImport.query.filter_by(
            tipo="beneficios_tmh",
            competencia=competence,
            filial_id=branch.id,
            empresa_id=company.id,
            ativo=True,
        ).update({"ativo": False}, synchronize_session=False)

        generated = DreImport(
            tipo="beneficios_tmh",
            competencia=competence,
            filial_id=branch.id,
            empresa_id=company.id,
            arquivo_original="Cadastro de benefícios do TMHub",
            arquivo_hash=sha256(
                f"beneficios_tmh|{competence.isoformat()}|{dt.now().isoformat()}".encode("utf-8")
            ).hexdigest(),
            registros_lidos=len(entries),
            registros_importados=len(entries),
            importado_por_usuario_id=token_data.get("id"),
        )
        db.session.add(generated)
        db.session.flush()
        for entry in entries:
            db.session.add(DreEntry(
                importacao_id=generated.id,
                empresa_id=company.id,
                criado_por_usuario_id=token_data.get("id"),
                **entry,
            ))
        db.session.commit()
        return jsonify({
            "message": "Benefícios do TMHub gerados para a competência.",
            "competencia": competence.isoformat(),
            "lancamentos": len(entries),
        }), 201

    @safe_route
    def create_manual_entry(self, token_data):
        """Registra um ajuste ou substituição auditável para a DRE."""
        if not has_permission(token_data, "controle_dre", "create"):
            return jsonify("Você não possui permissão para lançar valores na DRE."), 403
        if not is_admin(token_data):
            return jsonify("Os lançamentos manuais da DRE são permitidos somente para administradores."), 403

        body = request.get_json(silent=True) or {}
        try:
            competence = dt.strptime(str(body.get("competencia") or ""), "%Y-%m").date().replace(day=1)
            department = _as_int(body.get("departamento"))
            category = str(body.get("categoria") or "").strip().lower()
            value = _as_decimal(body.get("valor"), "Valor")
        except ValueError:
            return jsonify("Informe competência, departamento, categoria e valor válidos."), 400
        if category not in MANUAL_CATEGORIES:
            return jsonify("A categoria manual informada não pode ser lançada diretamente."), 400
        replaces_import = str(body.get("substitui_importacao", True)).strip().lower() in {
            "1", "true", "sim", "yes", "on",
        }

        # VA e VT possuem regras próprias. O VA recebe o valor nominal e a
        # DRE armazena somente o custo empresarial após 20% de desconto. O VT
        # já representa o custo total e, portanto, não sofre conversão.
        if category == "valor_va":
            value = _va_company_cost(value)

        branch, _, company, _ = self._selected_context(
            token_data,
            branch_id=body.get("filial_id"),
            company_id=body.get("empresa_id"),
        )
        if not branch or not company:
            return jsonify("Selecione uma filial e empresa disponíveis para a DRE."), 404
        departments = self._allowed_departments(token_data, branch, company)
        # Competências passadas podem conter departamentos que já não estão
        # vinculados ao centro de custo atual. Mantemos esses registros
        # editáveis usando o centro salvo no próprio histórico da DRE.
        departments = {
            **self._historic_department_map(token_data, branch, company, competence),
            **departments,
        }
        if department not in departments:
            return jsonify("O departamento não pertence à filial e empresa selecionadas para esta competência."), 400
        center_id = departments[department]
        created_at = dt.now()
        manual_import = DreImport(
            tipo="manual",
            competencia=competence,
            filial_id=branch.id,
            empresa_id=company.id,
            arquivo_original="Lançamento manual",
            arquivo_hash=sha256(
                f"manual|{token_data.get('id')}|{created_at.isoformat()}".encode("utf-8")
            ).hexdigest(),
            registros_lidos=1,
            registros_importados=1,
            importado_por_usuario_id=token_data.get("id"),
        )
        db.session.add(manual_import)
        db.session.flush()
        db.session.add(DreEntry(
            importacao_id=manual_import.id,
            competencia=competence,
            empresa_id=company.id,
            departamento=department,
            centro_custo_id=center_id,
            categoria=category,
            valor=value,
            descricao=str(body.get("descricao") or "").strip()[:500] or None,
            chave_origem=sha256(f"manual-entry|{manual_import.id}".encode("utf-8")).hexdigest(),
            substitui_importacao=replaces_import,
            criado_por_usuario_id=token_data.get("id"),
        ))
        db.session.commit()
        message = (
            "O valor manual substituirá a importação desta categoria no demonstrativo."
            if replaces_import
            else "O ajuste manual foi somado ao demonstrativo."
        )
        if category == "valor_va":
            message = f"{message} O VA foi registrado com o desconto empresarial de 20%."
        elif category == "valor_vt":
            message = f"{message} O VT foi registrado pelo valor integral informado."
        return jsonify({"message": message, "valor_registrado": float(value)}), 201

    @safe_route
    def delete_competence(self, token_data, competencia):
        """Remove exclusivamente as fontes e os lançamentos de uma competência."""
        if not has_permission(token_data, "controle_dre", "edit") or not is_admin(token_data):
            return jsonify("A exclusão de competências da DRE é permitida somente para administradores."), 403
        try:
            competence = dt.strptime(str(competencia), "%Y-%m").date().replace(day=1)
        except ValueError:
            return jsonify("Informe a competência no formato aaaa-mm."), 400

        branch, _, company, _ = self._selected_context(token_data)
        if not branch or not company:
            return jsonify("Nenhuma filial e empresa com dados de DRE está disponível para exclusão."), 404

        imports = DreImport.query.filter_by(
            competencia=competence,
            filial_id=branch.id,
            empresa_id=company.id,
        ).all()
        if not imports:
            return jsonify("Não há dados importados para esta competência, filial e empresa."), 404

        for imported in imports:
            db.session.delete(imported)
        db.session.commit()
        return jsonify({
            "message": f"Os dados de {competence.strftime('%m/%Y')} foram excluídos.",
            "competencia": competence.isoformat(),
            "filial_id": branch.id,
            "empresa_id": company.id,
            "importacoes_excluidas": len(imports),
        }), 200

    @classmethod
    def _available_branches(cls, token_data):
        """Lista as filiais ativas que o usuário pode selecionar na DRE."""
        query = (
            Branch.query
            .filter(Branch.ativa.is_(True))
            .order_by(Branch.nome)
        )
        if not is_admin(token_data):
            query = query.join(
                filial_usuarios,
                filial_usuarios.c.filial_id == Branch.id,
            ).filter(filial_usuarios.c.usuario_id == token_data.get("id"))
        return query.all()

    @classmethod
    def _selected_branch(cls, token_data, branch_id=None):
        """Valida a filial solicitada antes de usá-la na consulta ou exclusão."""
        branches = cls._available_branches(token_data)
        if not branches:
            return None, []
        requested_id = _as_int(branch_id if branch_id is not None else request.values.get("filial_id"))
        selected = next((branch for branch in branches if branch.id == requested_id), None)
        if not selected:
            global_scope = requested_branch_ids()
            selected = next(
                (branch for branch in branches if global_scope and branch.id in global_scope),
                None,
            )
        if not selected:
            branches_with_dre = {
                branch_id
                for branch_id, in db.session.query(DreImport.filial_id)
                .filter(DreImport.ativo.is_(True))
                .distinct()
                .all()
            }
            selected = next((branch for branch in branches if branch.id in branches_with_dre), branches[0])
        return selected, branches

    @classmethod
    def _available_companies(cls, token_data, branch):
        """Lista somente as empresas com centros de custo no escopo da filial."""
        if not branch:
            return []
        query = (
            Company.query
            .join(CostCenters, CostCenters.empresa_id == Company.id)
            .outerjoin(
                filial_centros_custo,
                filial_centros_custo.c.centro_custo_id == CostCenters.id,
            )
            .outerjoin(
                filial_departamentos,
                filial_departamentos.c.departamento == CostCenters.departamento,
            )
            .filter(
                Company.ativa.is_(True),
                or_(
                    filial_centros_custo.c.filial_id == branch.id,
                    filial_departamentos.c.filial_id == branch.id,
                ),
            )
        )
        allowed_centers = allowed_cost_center_ids(token_data)
        if allowed_centers is not None:
            query = query.filter(CostCenters.id.in_(allowed_centers))
        return query.distinct().order_by(Company.nome).all()

    @classmethod
    def _selected_context(cls, token_data, branch_id=None, company_id=None):
        """Resolve a filial e empresa antes de consultar ou gravar dados financeiros."""
        branch, branches = cls._selected_branch(token_data, branch_id)
        companies = cls._available_companies(token_data, branch)
        if not branch or not companies:
            return branch, branches, None, companies
        requested_id = _as_int(
            company_id if company_id is not None else request.values.get("empresa_id"),
        )
        company = next((item for item in companies if item.id == requested_id), companies[0])
        return branch, branches, company, companies

    @classmethod
    def _records(cls, token_data, branch, company):
        departments = cls._allowed_departments(token_data, branch, company)
        if not departments or not branch or not company:
            return []
        competence_values = {
            value.strip()
            for raw in request.args.getlist("competencia")
            for value in str(raw).split(",")
            if value.strip()
        }
        query = (
            DreEntry.query
            .join(DreImport, DreImport.id == DreEntry.importacao_id)
            .filter(
                DreImport.ativo.is_(True),
                DreImport.filial_id == branch.id,
                DreImport.empresa_id == company.id,
                DreEntry.empresa_id == company.id,
                DreEntry.departamento.in_(departments),
            )
        )
        if competence_values:
            values = []
            for value in competence_values:
                try:
                    values.append(dt.strptime(value, "%Y-%m").date().replace(day=1))
                except ValueError:
                    continue
            if values:
                query = query.filter(DreEntry.competencia.in_(values))
        return query.order_by(DreEntry.competencia.desc(), DreEntry.departamento, DreEntry.id).all()

    @classmethod
    def _june_forecast_by_department(cls, token_data, branch, company):
        """Obtém a previsão de junho para completar apenas maio/2026 sem previsão."""
        departments = cls._allowed_departments(token_data, branch, company)
        if not departments or not branch or not company:
            return {}

        june_forecasts = defaultdict(Decimal)
        entries = (
            DreEntry.query
            .join(DreImport, DreImport.id == DreEntry.importacao_id)
            .filter(
                DreImport.ativo.is_(True),
                DreImport.filial_id == branch.id,
                DreImport.empresa_id == company.id,
                DreEntry.empresa_id == company.id,
                DreEntry.departamento.in_(departments),
                DreEntry.competencia == date(2026, 6, 1),
                DreEntry.categoria == "receita_prevista",
            )
            .all()
        )
        for entry in entries:
            june_forecasts[entry.departamento] += Decimal(str(entry.valor or 0))
        return june_forecasts

    @staticmethod
    def _fill_missing_may_forecast(records, june_forecasts):
        """Usa a previsão de junho quando maio não possui valor importado."""
        for record in records:
            if record["competencia"] != "2026-05-01" or record["receita_prevista"]:
                continue
            forecast = june_forecasts.get(record["departamento"], Decimal("0"))
            record["receita_prevista"] = float(forecast)
            record["faturamento_fora_competencia"] = float(
                forecast - Decimal(str(record["rob"])),
            )

    @staticmethod
    def _calculate(rows):
        """Calcula somente indicadores derivados; as entradas continuam auditáveis."""
        grouped = {}
        replacements = {}
        manual_replacements = set()
        for row in rows:
            # A linha principal consolida todas as fontes da competência em
            # um departamento. O detalhamento é exibido ao expandir a linha.
            key = (row.competencia, row.empresa_id, row.departamento)
            item = grouped.setdefault(key, defaultdict(Decimal))
            item["empresa_id"] = row.empresa_id
            item["centro_custo_id"] = item.get("centro_custo_id") or row.centro_custo_id
            item["contrato_codigo"] = row.contrato_codigo or item.get("contrato_codigo")
            if row.substitui_importacao:
                replacement_key = (*key, row.categoria)
                previous = replacements.get(replacement_key)
                # O registro mais recente prevalece, mantendo os anteriores
                # disponíveis para auditoria e sem somar valores substituídos.
                if previous is None or row.id > previous.id:
                    replacements[replacement_key] = row
                continue

            # O encargo importado nos demonstrativos antigos não representa a
            # base completa utilizada pela DRE. Ele será recalculado a partir
            # de salários, horas extras, VA e VT.
            if row.categoria == "encargos" and row.importacao.tipo != "manual":
                continue
            item[row.categoria] += Decimal(str(row.valor or 0))

        for replacement_key, row in replacements.items():
            competence, company_id, department, category = replacement_key
            key = (competence, company_id, department)
            grouped[key][category] = Decimal(str(row.valor or 0))
            manual_replacements.add(replacement_key)

        records = []
        for (competence, company_id, department), values in grouped.items():
            glosa = values["glosa_analitica"] or values["glosa_faturamento"]
            # Importações antigas podem ter a insalubridade registrada à parte.
            # A Folha Analítica atual já a consolida em "Liq. Folha".
            folha_salarial = values["salarios"] + values["insalubridade"]
            horas_extras = values["horas_extras"]
            salarios = folha_salarial + horas_extras
            custos_operacionais = sum(
                values[category]
                for category in ("materiais", "uniformes", "epis", "combustivel", "locacao", "manutencao", "maquinario", "ajuste_custos")
            )
            # Os encargos incidem sobre toda a base de RH: salários líquidos,
            # horas extras, VA e VT. Uma substituição manual de encargos é a
            # única exceção, pois foi informada deliberadamente pelo RH.
            base_encargos = (
                folha_salarial
                + horas_extras
                + values["valor_va"]
                + values["valor_vt"]
            )
            encargos_key = (competence, company_id, department, "encargos")
            encargos = (
                values["encargos"]
                if encargos_key in manual_replacements
                else (base_encargos * Decimal("0.67")).quantize(MONEY)
            )
            impostos = (
                values["impostos"]
                if "impostos" in values
                else (values["rob"] * DEFAULT_TAX_RATE).quantize(MONEY)
            )
            total_custos = salarios + encargos + values["valor_va"] + values["valor_vt"] + custos_operacionais
            rol = values["rob"] - impostos
            # A glosa já vem abatida no faturamento informado pela operação.
            # Ela permanece no retorno apenas para demonstrativo e não pode
            # reduzir o resultado uma segunda vez.
            margem = rol - total_custos
            percent_margin = (margem / rol * 100) if rol else Decimal("0")
            records.append({
                "competencia": competence.isoformat(),
                "empresa_id": company_id,
                "departamento": department,
                "centro_custo_id": values["centro_custo_id"],
                "contrato_codigo": values["contrato_codigo"],
                "receita_prevista": float(values["receita_prevista"]),
                "rob": float(values["rob"]),
                "outros_faturamentos": float(values["outros_faturamentos"]),
                "repactuacao": float(values["repactuacao"]),
                "faturamento_fora_competencia": float(values["receita_prevista"] - values["rob"]),
                "glosas": float(glosa),
                "impostos": float(impostos),
                "rol": float(rol),
                "salarios": float(salarios),
                "encargos": float(encargos),
                "va": float(values["valor_va"]),
                "vt": float(values["valor_vt"]),
                "custos_operacionais": float(custos_operacionais),
                "composicao": {
                    "salarios": float(folha_salarial),
                    "horas_extras": float(horas_extras),
                    "encargos": float(encargos),
                    "va": float(values["valor_va"]),
                    "vt": float(values["valor_vt"]),
                    "materiais": float(values["materiais"]),
                    "uniformes": float(values["uniformes"]),
                    "epis": float(values["epis"]),
                    "combustivel": float(values["combustivel"]),
                    "locacao": float(values["locacao"]),
                    "manutencao": float(values["manutencao"]),
                    "maquinario": float(values["maquinario"]),
                    "ajuste_custos": float(values["ajuste_custos"]),
                },
                "custos": {
                    "materiais": float(values["materiais"]),
                    "uniformes": float(values["uniformes"]),
                    "epis": float(values["epis"]),
                    "combustivel": float(values["combustivel"]),
                    "locacao": float(values["locacao"]),
                    "manutencao": float(values["manutencao"]),
                    "maquinario": float(values["maquinario"]),
                },
                "total_custos": float(total_custos),
                "margem": float(margem),
                "percentual_margem": float(percent_margin.quantize(Decimal("0.01"))),
            })
        return sorted(
            records,
            key=lambda item: (item["competencia"], item["empresa_id"], item["departamento"]),
            reverse=True,
        )

    @safe_route
    def read(self, token_data):
        if not has_permission(token_data, "controle_dre", "view"):
            return jsonify("Você não possui acesso ao controle DRE."), 403
        branch, branches, company, companies = self._selected_context(token_data)
        records = self._calculate(self._records(token_data, branch, company))
        self._fill_missing_may_forecast(
            records,
            self._june_forecast_by_department(token_data, branch, company),
        )

        total_rol = sum(Decimal(str(item["rol"])) for item in records)
        total_margin = sum(Decimal(str(item["margem"])) for item in records)
        summary_margin_percent = (
            (total_margin / total_rol * Decimal("100")).quantize(MONEY)
            if total_rol
            else Decimal("0")
        )

        summary = {
            "contratos": len(records),
            "previsto": sum(item["receita_prevista"] for item in records),
            "rob": sum(item["rob"] for item in records),
            "rol": sum(item["rol"] for item in records),
            "glosas": sum(item["glosas"] for item in records),
            "custos": sum(item["total_custos"] for item in records),
            "margem": float(total_margin),
            "percentual_margem": float(summary_margin_percent),
        }
        return jsonify({
            "registros": records,
            "resumo": summary,
            "filtros": {
                "competencias": sorted({item["competencia"][:7] for item in records}, reverse=True),
                "departamentos": sorted({item["departamento"] for item in records}),
            },
            "filiais": [
                {"id": branch_item.id, "nome": branch_item.nome}
                for branch_item in branches
            ],
            "filial_selecionada": branch.id if branch else None,
            "empresas": [
                {"id": company_item.id, "nome": company_item.nome}
                for company_item in companies
            ],
            "empresa_selecionada": company.id if company else None,
            "pode_excluir": bool(is_admin(token_data)),
        }), 200

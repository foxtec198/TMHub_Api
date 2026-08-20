"""Leitor normalizado dos relatórios XLS/XLSX de colaboradores.

O relatório corporativo pode reunir mais de uma empresa na mesma aba. A
empresa é identificada pelo cabeçalho do arquivo e, nos blocos seguintes, pelo
prefixo do departamento (por exemplo ``FACILITIES - ADMINISTRATIVO``).
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from re import match, sub
from typing import Any, Iterable
from unicodedata import combining, normalize

from import_col.cols import _cidade_para
from import_col.date_normalization import normalize_import_date


SUPPORTED_EXTENSIONS = {".xls", ".xlsx"}
COMPANY_TOKENS = ("COSTA OESTE", "FACILITIES", "GRABIN", "MAG")
# Índices do relatório legado. Arquivos novos são resolvidos pelo texto do
# cabeçalho para não dependerem da posição escolhida na exportação.
HEADER_COLUMNS = {
    "codigo": 0,
    "nome": 4,
    "cargo": 11,
    "centro_custo_num": 18,
    "centro_custo": 19,
    "hor": 20,
    "admissao": 22,
    "situacao": 26,
    "cpf": 28,
    "salario": 32,
}

HEADER_ALIASES = {
    "codigo": {"CODIGO", "MATRICULA", "MATRICULA COLABORADOR", "CODIGO COLABORADOR"},
    "nome": {"NOME", "NOME COLABORADOR", "COLABORADOR", "NOME FUNCIONARIO"},
    "cargo": {"CARGO", "FUNCAO", "FUNCAO COLABORADOR"},
    "centro_custo_num": {
        "CODIGO CENTRO DE CUSTO", "COD CENTRO DE CUSTO", "CODIGO C CENTRO",
        "C CUSTO", "CCUSTO", "CENTRO CUSTO CODIGO",
    },
    "centro_custo": {
        "CENTRO DE CUSTO", "CENTRO CUSTO", "NOME CENTRO DE CUSTO",
        "CONTRATO", "LOCAL", "LOCAL DE TRABALHO",
    },
    "empresa": {"EMPRESA", "RAZAO SOCIAL", "EMPREGADOR"},
    "hor": {"HOR", "CARGA HORARIA", "CARGA HORARIA MENSAL"},
    "admissao": {"ADMISSAO", "DATA ADMISSAO", "DT ADMISSAO"},
    "situacao": {"SITUACAO", "STATUS", "SITUACAO COLABORADOR"},
    "cpf": {"CPF"},
    "salario": {"SALARIO", "SALARIO BASE"},
    "departamento": {"DEPARTAMENTO", "DPTO", "DEPARTAMENTO CODIGO"},
}


def _clean(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _normalized(value: Any) -> str:
    decomposed = normalize("NFKD", _clean(value).upper())
    return "".join(char for char in decomposed if not combining(char))


def _header_key(value: Any) -> str:
    """Normaliza títulos de planilha sem confundir texto de linhas de dados."""
    return sub(r"[^A-Z0-9]+", " ", _normalized(value)).strip()


def _header_mapping(row: tuple[Any, ...]) -> dict[str, int] | None:
    """Detecta a linha de cabeçalho da exportação nova sem perder o legado."""
    available = {_header_key(value): index for index, value in enumerate(row) if _header_key(value)}
    mapped = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in available:
                mapped[field] = available[alias]
                break

    required = {"codigo", "nome"}
    has_center = "centro_custo" in mapped or "centro_custo_num" in mapped
    return mapped if required.issubset(mapped) and has_center else None


def _known_company(value: Any) -> str | None:
    normalized = _normalized(value)
    for token in COMPANY_TOKENS:
        # Cabeçalho corporativo ("COSTA OESTE SERVIÇOS") ou o prefixo do
        # departamento ("1 - FACILITIES - ..."). Não basta procurar o texto
        # em qualquer ponto: um contrato pode citar "MAG" no próprio nome.
        if normalized.startswith(token) or match(rf"^\d+\s*-\s*{token}(?:\s|-|$)", normalized):
            return token
    return None


def _company_name(value: Any, fallback: str) -> str:
    known_company = _known_company(value)
    if known_company:
        return known_company
    cleaned = _clean(value)
    return cleaned or fallback


def _filename_company(filename: str) -> str:
    name = Path(filename).stem.replace("_", " ")
    return _company_name(name, "COSTA OESTE")


def _department_details(row: tuple[Any, ...]) -> tuple[str, int | None] | None:
    joined = " ".join(_clean(value) for value in row if _clean(value))
    if "DEPARTAMENTO:" not in _normalized(joined):
        return None
    # O layout padrão possui a descrição depois do rótulo. Mantemos também
    # uma busca no texto completo para formatos exportados com colunas unidas.
    values = [_clean(value) for value in row if _clean(value)]
    description = next(
        (value for value in values if " - " in value and "DEPARTAMENTO" not in _normalized(value)),
        joined,
    )
    number_match = match(r"\s*(\d+)", description)
    code = int(number_match.group(1)) if number_match else None
    return description, code


def _as_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        digits = sub(r"\D", "", str(value))
        return int(digits) if digits else None


def _cell(row: tuple[Any, ...], key: str, columns: dict[str, int] | None = None) -> Any:
    index = (columns or {}).get(key, HEADER_COLUMNS.get(key))
    if index is None:
        return None
    return row[index] if len(row) > index else None


def _is_employee_row(row: tuple[Any, ...], columns: dict[str, int] | None = None) -> bool:
    return (
        _as_int(_cell(row, "codigo", columns)) is not None
        and bool(_clean(_cell(row, "nome", columns)))
    )


def _center_code(value: Any) -> int | None:
    """Extrai o código somente quando ele abre o nome do centro de custo."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = int(value)
        return numeric if numeric == value else None
    cleaned = _clean(value)
    if match(r"^\d+$", cleaned):
        return int(cleaned)
    prefix = match(r"\s*(\d+)\s*(?:-|/|$)", cleaned)
    return int(prefix.group(1)) if prefix else None


def _department_code(value: Any) -> int | None:
    direct = _as_int(value)
    if direct is not None:
        return direct
    prefix = match(r"\s*(\d+)", _clean(value))
    return int(prefix.group(1)) if prefix else None


def _center_details(row: tuple[Any, ...]) -> tuple[str, int | None] | None:
    """Lê o cabeçalho de grupo do relatório: ``Centro de Custo: 87 - ...``."""
    values = [_clean(value) for value in row]
    label_index = next(
        (index for index, value in enumerate(values) if "CENTRO DE CUSTO" in _normalized(value)),
        None,
    )
    if label_index is None:
        return None
    center_name = next((value for value in values[label_index + 1:] if value), "")
    if not center_name:
        return None
    return center_name, _center_code(center_name)


def _xlsx_rows(path: Path) -> Iterable[tuple[Any, ...]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            yield from sheet.iter_rows(values_only=True)
    finally:
        workbook.close()


def _xls_rows(path: Path) -> Iterable[tuple[Any, ...]]:
    """Lê relatórios XLS legados, inclusive exports OLE aceitos pelo Excel.

    Alguns relatórios ``RELAÇÃO DE EMPREGADOS II`` usam uma variação de XLS
    que o ``xlrd`` não consegue abrir, apesar de ser aberta normalmente pelo
    Excel. O Calamine cobre essa variação sem depender de Office no servidor.
    Mantemos o xlrd como contingência para os XLS convencionais.
    """
    calamine_error: Exception | None = None
    try:
        from python_calamine import load_workbook

        workbook = load_workbook(path)
        for sheet_name in workbook.sheet_names:
            sheet = workbook.get_sheet_by_name(sheet_name)
            for row in sheet.to_python():
                yield tuple(row)
        return
    except (ImportError, OSError, ValueError, RuntimeError) as error:
        calamine_error = error

    try:
        import xlrd
    except ModuleNotFoundError as error:
        raise ValueError(
            "Leitura de .xls indisponível. Instale python-calamine ou xlrd no servidor."
        ) from error

    try:
        workbook = xlrd.open_workbook(path)
        for sheet in workbook.sheets():
            for row_index in range(sheet.nrows):
                values = []
                for column_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, column_index)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        values.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
                    else:
                        values.append(cell.value)
                yield tuple(values)
    except Exception as error:
        raise ValueError(
            "Não foi possível abrir o XLS. Exporte novamente o relatório "
            "RELAÇÃO DE EMPREGADOS II ou envie em XLSX."
        ) from (calamine_error or error)


def _rows(path: Path) -> Iterable[tuple[Any, ...]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _xlsx_rows(path)
    if suffix == ".xls":
        return _xls_rows(path)
    raise ValueError("Formato não suportado. Envie .xls ou .xlsx.")


def parse_employee_spreadsheet(
    path: str | Path,
    filename: str | None = None,
    centro_forcado: int | None = None,
) -> dict[str, Any]:
    """Converte uma planilha de relatório em registros de importação.

    Retorna colaboradores já identificados por empresa e uma lista de avisos;
    não persiste dados nem cria empresas nesta camada.
    """

    source_path = Path(path)
    if source_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("Formato não suportado. Envie .xls ou .xlsx.")

    fallback_company = _filename_company(filename or source_path.name)
    if centro_forcado is not None and int(centro_forcado) <= 0:
        raise ValueError("O centro forçado deve ser um código positivo.")
    forced_center = int(centro_forcado) if centro_forcado is not None else None
    current_company = fallback_company
    current_department = ""
    current_department_code: int | None = None
    current_columns: dict[str, int] | None = None
    current_center_name: str | None = None
    current_center_code: int | None = None
    employees: list[dict[str, Any]] = []
    invalid: list[str] = []
    companies: set[str] = set()

    for row_number, raw_row in enumerate(_rows(source_path), start=1):
        row = tuple(raw_row)
        detected_header = _header_mapping(row)
        if detected_header:
            current_columns = detected_header
            continue
        first_cell = _clean(row[0] if row else None)
        identified_company = _known_company(first_cell)
        if identified_company:
            current_company = identified_company
        elif first_cell and "SERVICOS" in _normalized(first_cell):
            current_company = _company_name(first_cell, fallback_company)

        center = _center_details(row)
        if center:
            current_center_name, current_center_code = center
            continue

        department = _department_details(row)
        if department:
            current_department, current_department_code = department
            current_company = _known_company(current_department) or current_company
            continue
        if not _is_employee_row(row, current_columns):
            continue

        registration = _as_int(_cell(row, "codigo", current_columns))
        center_name = (
            _clean(_cell(row, "centro_custo", current_columns))
            or current_center_name
        )
        center_code = forced_center or _center_code(
            _cell(row, "centro_custo_num", current_columns)
        ) or _center_code(center_name) or current_center_code
        admission_value = _cell(row, "admissao", current_columns)
        # O RELAÇÃO DE EMPREGADOS II mescla visualmente o cabeçalho de
        # admissão uma coluna à direita do valor. Quando isso ocorre, a célula
        # apontada pelo cabeçalho fica vazia e o índice legado é o valor real.
        if admission_value in (None, "") and current_columns is not None:
            admission_value = _cell(row, "admissao")
        try:
            admission = normalize_import_date(
                admission_value,
                field="data de admissão",
            )
        except ValueError:
            admission = None
        if not registration or (not center_code and not center_name) or admission is None:
            invalid.append(f"Linha {row_number}: matrícula, centro/nome do centro ou admissão inválidos.")
            continue

        company = _company_name(
            _cell(row, "empresa", current_columns) or current_company,
            fallback_company,
        )
        department_value = _clean(_cell(row, "departamento", current_columns))
        department_code = _department_code(department_value) or current_department_code
        mapped_department, city_id, city_name = _cidade_para(center_code)
        if department_code is None:
            department_code = mapped_department
        companies.add(company)
        employees.append(
            {
                "codigo": registration,
                "nome": _clean(_cell(row, "nome", current_columns)),
                "cargo": _clean(_cell(row, "cargo", current_columns)),
                "centro_custo_num": center_code,
                "centro_custo": center_name,
                "hor": _cell(row, "hor", current_columns),
                "admissao": admission,
                "situacao": _as_int(_cell(row, "situacao", current_columns)),
                "cpf": _clean(_cell(row, "cpf", current_columns)) or None,
                "salario": _cell(row, "salario", current_columns),
                "departamento": department_value or current_department,
                "departamento_codigo": department_code,
                "cidade_id": city_id,
                "cidade": city_name,
                "empresa_nome": company,
            }
        )

    centers: dict[tuple[str, int], dict[str, Any]] = {}
    for employee in employees:
        key = (employee["empresa_nome"], employee["centro_custo_num"])
        centers.setdefault(
            key,
            {
                "empresa": employee["empresa_nome"],
                "centro_id": employee["centro_custo_num"],
                "nome": employee["centro_custo"],
                "empregados": [],
            },
        )["empregados"].append(employee)

    return {
        "employees": employees,
        "invalid": invalid,
        "companies": sorted(companies),
        "source": source_path.name,
        "centros_de_custo": list(centers.values()),
        "empresa": fallback_company if len(companies) == 1 else None,
        "relatorio": "RELAÇÃO DE EMPREGADOS II",
        "centro_forcado": forced_center,
    }

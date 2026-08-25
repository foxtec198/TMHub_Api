# Regras de negócio do controle de exames periódicos.
# Biblioteca padrão.
from datetime import date, datetime, timezone
from io import BytesIO
from uuid import uuid4
import re
import unicodedata

# Dependências externas.
from flask import jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import String, case, cast, or_
from sqlalchemy.orm import aliased

# Módulos internos da aplicação.
from models.centros_de_custo import CostCenters
from models.colaboradores import Employees
from models.exames_periodicos import PeriodicExam
from models.supervisores import Supervisors
from models.usuarios import Users
from utils.db import db
from utils.filial_scope import apply_cost_center_scope, is_admin
from utils.permissions import has_permission
from utils.safe_route import safe_route
from utils.socket import socketio


ACTIVE_EMPLOYEE_STATUS = 1
ALLOWED_STATUSES = ("a_vencer", "pendente", "em_andamento", "concluido")
MANUAL_STATUSES = ("pendente", "em_andamento", "concluido")
MAX_IMPORT_SIZE = 25 * 1024 * 1024
MAX_IMPORT_ROWS = 100_000


def _normalise(value):
    """Normaliza textos para comparar cabeçalhos e tipos de exame."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().upper()


def _as_int(value):
    """Converte códigos de planilha que podem chegar como 101001.0."""
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _as_date(value):
    """Converte datas do Excel ou textos brasileiros em um objeto date."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _first_day_next_month(reference_date=None):
    """Entrega o primeiro dia do mês seguinte sem usar uma quantidade fixa de dias."""
    reference_date = reference_date or date.today()
    if reference_date.month == 12:
        return date(reference_date.year + 1, 1, 1)
    return date(reference_date.year, reference_date.month + 1, 1)


def _first_day_month_after_next(reference_date=None):
    """Define o limite exclusivo que abrange todo o mês subsequente."""
    return _first_day_next_month(_first_day_next_month(reference_date))


class PeriodicExamService:
    """Importa, acompanha e conclui exames periódicos por colaborador."""

    @staticmethod
    def _status_for_due_date(due_date, reference_date=None):
        """Deixa pendentes todos os exames do mês atual ou do mês seguinte."""
        if due_date < _first_day_month_after_next(reference_date):
            return "pendente"
        return "a_vencer"

    @staticmethod
    def _selected_values(name):
        """Aceita filtros repetidos e valores separados por vírgula."""
        return {
            value.strip()
            for raw in request.args.getlist(name)
            for value in str(raw).split(",")
            if value.strip() and value.strip() != "__all__"
        }

    @classmethod
    def refresh_pending_statuses(cls, reference_date=None, commit=True):
        """Promove a pendente toda tarefa cujo vencimento é no mês subsequente."""
        rows = (
            PeriodicExam.query.join(Employees, Employees.id == PeriodicExam.colaborador_id)
            .filter(
                Employees.situacao == ACTIVE_EMPLOYEE_STATUS,
                PeriodicExam.status == "a_vencer",
                PeriodicExam.data_vencimento < _first_day_month_after_next(reference_date),
            )
            .all()
        )
        for exam in rows:
            exam.status = "pendente"

        if rows and commit:
            db.session.commit()
            socketio.emit("periodic_exam_update", {
                "action": "pending_statuses_refreshed",
                "ids": [exam.id for exam in rows],
            })
        return len(rows)

    @staticmethod
    def _value(row, column):
        """Lê a célula do cabeçalho e tolera uma coluna visual vazia no relatório SST."""
        if column is None:
            return None
        current = row[column] if column < len(row) else None
        if current is not None and str(current).strip() != "":
            return current
        following = column + 1
        return row[following] if following < len(row) else None

    @classmethod
    def _find_header(cls, rows):
        """Localiza os rótulos da planilha SST mesmo com títulos mesclados."""
        for row_index, row in enumerate(rows[:25]):
            labels = [_normalise(value) for value in row]
            if not any("NOME FUNCION" in label for label in labels):
                continue
            if not any("VENCIMENTO" in label for label in labels):
                continue

            def find(predicate):
                return next((index for index, label in enumerate(labels) if predicate(label)), None)

            headers = {
                "empresa": find(lambda label: label == "EMPRESA"),
                "matricula": find(
                    lambda label: label.startswith("COD") and "CUST" not in label
                ),
                "nome": find(lambda label: "NOME FUNCION" in label),
                "data_exame": find(lambda label: "DATA" in label and "EXAME" in label),
                "tipo_exame": find(lambda label: label == "TIPO"),
                "resultado": find(lambda label: "RESULTADO" in label),
                "data_vencimento": find(lambda label: "VENCIMENTO" in label),
                "centro_codigo": find(lambda label: "CUST" in label and "COD" in label),
                "centro_nome": find(
                    lambda label: label == "NOME" and labels.index(label) > 0
                ),
                "departamento": find(lambda label: label.startswith("DEPTO")),
            }
            required = ("matricula", "nome", "tipo_exame", "data_vencimento")
            if all(headers[field] is not None for field in required):
                return row_index, headers
        return None, None

    @staticmethod
    def _read_workbook(upload):
        """Lê XLSX com baixo consumo de memória e XLS pelo leitor já listado no projeto."""
        filename = (upload.filename or "").lower()
        if filename.endswith(".xlsx"):
            workbook = load_workbook(upload.stream, read_only=True, data_only=True)
            worksheet = workbook.active
            return [list(row) for row in worksheet.iter_rows(values_only=True)]

        # O formato XLS legado exige xlrd, já declarado no requirements.txt.
        import pandas as pd

        upload.stream.seek(0)
        frame = pd.read_excel(upload.stream, header=None, dtype=object)
        return frame.where(frame.notna(), None).values.tolist()

    @staticmethod
    def _serialize(exam, employee, center, supervisor):
        """Monta o contrato estável que a tela e a exportação consomem."""
        days_to_due = (exam.data_vencimento - date.today()).days
        return {
            "id": exam.id,
            "colaborador": employee.nome,
            "matricula": employee.matricula,
            "empresa_id": employee.empresa_id,
            "cargo": None,
            "departamento": center.departamento if center else None,
            "centro_custo": center.local if center else None,
            "centro_codigo": center.centro_id if center else None,
            "supervisor": supervisor.nome if supervisor else None,
            "tipo_exame": exam.tipo_exame,
            "data_exame": exam.data_exame.isoformat() if exam.data_exame else None,
            "resultado": exam.resultado,
            "data_vencimento": exam.data_vencimento.isoformat(),
            "status": exam.status,
            "observacao": exam.observacao,
            "concluido_em": exam.concluido_em.isoformat() if exam.concluido_em else None,
            "dias_para_vencimento": days_to_due,
            "vencido": days_to_due < 0 and exam.status != "concluido",
        }

    @classmethod
    def _query(cls, token_data):
        """Centraliza os joins, a situação ativa e o escopo de filiais."""
        supervisor = aliased(Supervisors)
        query = (
            db.session.query(PeriodicExam, Employees, CostCenters, supervisor)
            .join(Employees, Employees.id == PeriodicExam.colaborador_id)
            .outerjoin(CostCenters, CostCenters.id == PeriodicExam.centro_custo_id)
            .outerjoin(supervisor, supervisor.id == CostCenters.supervisor_id)
            .filter(Employees.situacao == ACTIVE_EMPLOYEE_STATUS)
        )
        return apply_cost_center_scope(query, CostCenters.id, token_data), supervisor

    @classmethod
    def _filtered_rows(cls, token_data):
        """Aplica filtros antes de entregar registros ou gerar o XLSX."""
        statuses = cls._selected_values("status")
        departments = cls._selected_values("departamento")
        supervisors = cls._selected_values("supervisor")
        contracts = cls._selected_values("contrato")
        months = cls._selected_values("competencia")
        search = (request.args.get("search") or "").strip().upper()

        query, supervisor = cls._query(token_data)
        if statuses:
            query = query.filter(PeriodicExam.status.in_(statuses))
        if departments:
            query = query.filter(cast(CostCenters.departamento, String).in_(departments))
        if supervisors:
            query = query.filter(supervisor.nome.in_(supervisors))
        if contracts:
            query = query.filter(CostCenters.local.in_(contracts))
        if months:
            query = query.filter(
                db.func.to_char(PeriodicExam.data_vencimento, "YYYY-MM").in_(months)
            )
        if search:
            term = f"%{search}%"
            query = query.filter(
                or_(
                    db.func.upper(Employees.nome).like(term),
                    cast(Employees.matricula, String).like(term),
                    db.func.upper(CostCenters.local).like(term),
                )
            )
        return query.order_by(
            case((PeriodicExam.status == "pendente", 0), else_=1),
            PeriodicExam.data_vencimento.asc(),
            Employees.nome.asc(),
        ).all()

    @safe_route
    def read(self, token_data):
        if not has_permission(token_data, "controle_exames_periodicos", "view"):
            return jsonify("Você não possui acesso ao Controle de Exames Periódicos."), 403

        self.refresh_pending_statuses()
        rows = self._filtered_rows(token_data)
        records = [self._serialize(*row) for row in rows]
        filters = {
            "status": list(ALLOWED_STATUSES),
            "departamento": sorted({str(row[2].departamento) for row in rows if row[2]}),
            "supervisor": sorted({row[3].nome for row in rows if row[3] and row[3].nome}),
            "contrato": sorted({row[2].local for row in rows if row[2] and row[2].local}),
            "competencia": sorted({record["data_vencimento"][:7] for record in records}),
        }
        summary = {
            "total": len(records),
            "a_vencer": sum(record["status"] == "a_vencer" for record in records),
            "pendentes": sum(record["status"] == "pendente" for record in records),
            "em_andamento": sum(record["status"] == "em_andamento" for record in records),
            "concluidos": sum(record["status"] == "concluido" for record in records),
            "vencidos": sum(record["vencido"] for record in records),
        }
        return jsonify({"registros": records, "resumo": summary, "filtros": filters}), 200

    @safe_route
    def import_spreadsheet(self, token_data):
        if not has_permission(token_data, "controle_exames_periodicos", "create"):
            return jsonify("Você não possui permissão para importar exames."), 403

        upload = request.files.get("file")
        filename = (upload.filename or "").lower() if upload else ""
        if not upload or not filename.endswith((".xlsx", ".xls")):
            return jsonify("Envie uma planilha SST nos formatos .xlsx ou .xls."), 400

        upload.stream.seek(0, 2)
        size = upload.stream.tell()
        upload.stream.seek(0)
        if not size or size > MAX_IMPORT_SIZE:
            return jsonify("A planilha deve ter até 25 MB."), 400

        try:
            rows = self._read_workbook(upload)
        except Exception as error:
            return jsonify(f"Não foi possível ler a planilha: {error}"), 400
        if len(rows) > MAX_IMPORT_ROWS:
            return jsonify(f"A planilha excede o limite de {MAX_IMPORT_ROWS:,} linhas."), 400

        header_row, headers = self._find_header(rows)
        if headers is None:
            return jsonify("Não localizamos os cabeçalhos Matrícula, Nome Funcionário, Tipo e Data Vencimento."), 400

        parsed = []
        errors = []
        for line_number, row in enumerate(rows[header_row + 1:], start=header_row + 2):
            matricula = _as_int(self._value(row, headers["matricula"]))
            exam_type = _normalise(self._value(row, headers["tipo_exame"]))
            due_date = _as_date(self._value(row, headers["data_vencimento"]))
            if not any(value is not None and str(value).strip() for value in row):
                continue
            if not matricula or not exam_type or not due_date:
                if len(errors) < 100:
                    errors.append(f"Linha {line_number}: matrícula, tipo e vencimento são obrigatórios.")
                continue
            parsed.append({
                "line": line_number,
                "empresa_id": _as_int(self._value(row, headers["empresa"])),
                "matricula": matricula,
                "nome": str(self._value(row, headers["nome"]) or "").strip(),
                "tipo_exame": exam_type,
                "data_exame": _as_date(self._value(row, headers["data_exame"])),
                "resultado": str(self._value(row, headers["resultado"]) or "").strip() or None,
                "data_vencimento": due_date,
            })

        if not parsed:
            return jsonify({"message": "Nenhuma linha válida foi encontrada.", "errors": errors}), 400

        matriculas = {item["matricula"] for item in parsed}
        employees = Employees.query.filter(Employees.matricula.in_(matriculas)).all()
        by_company_and_registration = {
            (employee.empresa_id, employee.matricula): employee for employee in employees
        }
        by_registration = {}
        for employee in employees:
            by_registration.setdefault(employee.matricula, []).append(employee)

        active_employee_ids = [
            employee.id for employee in employees if employee.situacao == ACTIVE_EMPLOYEE_STATUS
        ]
        existing = {
            (exam.colaborador_id, exam.tipo_exame, exam.data_vencimento): exam
            for exam in PeriodicExam.query.filter(PeriodicExam.colaborador_id.in_(active_employee_ids)).all()
        } if active_employee_ids else {}

        created = updated = skipped_inactive = skipped_missing = 0
        batch_id = str(uuid4())
        for item in parsed:
            employee = by_company_and_registration.get((item["empresa_id"], item["matricula"]))
            if not employee:
                matches = by_registration.get(item["matricula"], [])
                employee = matches[0] if len(matches) == 1 else None
            if not employee:
                skipped_missing += 1
                if len(errors) < 100:
                    errors.append(f"Linha {item['line']}: matrícula {item['matricula']} não encontrada ou ambígua.")
                continue
            if employee.situacao != ACTIVE_EMPLOYEE_STATUS:
                skipped_inactive += 1
                continue

            key = (employee.id, item["tipo_exame"], item["data_vencimento"])
            exam = existing.get(key)
            if exam:
                if exam.status != "concluido":
                    exam.data_exame = item["data_exame"]
                    exam.resultado = item["resultado"]
                    exam.centro_custo_id = employee.centro_id
                    updated += 1
                continue

            exam = PeriodicExam(
                colaborador_id=employee.id,
                centro_custo_id=employee.centro_id,
                tipo_exame=item["tipo_exame"],
                data_exame=item["data_exame"],
                resultado=item["resultado"],
                data_vencimento=item["data_vencimento"],
                status=self._status_for_due_date(item["data_vencimento"]),
                lote_importacao=batch_id,
            )
            db.session.add(exam)
            existing[key] = exam
            created += 1

        if not created and not updated:
            db.session.rollback()
            return jsonify({
                "message": "A importação não encontrou exames novos ou atualizáveis.",
                "errors": errors,
                "ignorados_inativos": skipped_inactive,
                "ignorados_sem_vinculo": skipped_missing,
            }), 400

        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise

        socketio.emit("periodic_exam_update", {
            "action": "imported",
            "created": created,
            "updated": updated,
        })
        return jsonify({
            "message": "Importação concluída.",
            "criados": created,
            "atualizados": updated,
            "ignorados_inativos": skipped_inactive,
            "ignorados_sem_vinculo": skipped_missing,
            "errors": errors,
        }), 201

    @safe_route
    def update(self, exam_id, token_data):
        if not has_permission(token_data, "controle_exames_periodicos", "edit"):
            return jsonify("Você não possui permissão para editar exames."), 403

        exam = PeriodicExam.query.get(exam_id)
        if not exam:
            return jsonify("Exame não encontrado."), 404
        scoped_query, _ = self._query(token_data)
        scoped = scoped_query.filter(PeriodicExam.id == exam_id).first()
        if not scoped:
            return jsonify("Você não possui acesso a este exame."), 403

        body = request.get_json(silent=True) or {}
        status = body.get("status")
        if status is not None:
            if status not in MANUAL_STATUSES:
                return jsonify("Estado de exame inválido."), 400
            exam.status = status
            if status == "concluido":
                exam.concluido_em = datetime.now(timezone.utc)
                exam.concluido_por_usuario_id = token_data.get("id")
            else:
                exam.concluido_em = None
                exam.concluido_por_usuario_id = None
        if "observacao" in body:
            exam.observacao = str(body.get("observacao") or "").strip()[:500] or None
        db.session.commit()
        socketio.emit("periodic_exam_update", {"action": "updated", "id": exam.id})
        return jsonify({"message": "Exame atualizado."}), 200

    @safe_route
    def update_bulk_status(self, token_data):
        if not has_permission(token_data, "controle_exames_periodicos", "edit"):
            return jsonify("Você não possui permissão para concluir exames."), 403
        body = request.get_json(silent=True) or {}
        status = body.get("status")
        if status not in MANUAL_STATUSES:
            return jsonify("Estado de exame inválido."), 400

        ids = [int(item) for item in body.get("ids", []) if str(item).isdigit()]
        query, _ = self._query(token_data)
        if ids:
            query = query.filter(PeriodicExam.id.in_(ids))
        else:
            department = str(body.get("departamento") or "").strip()
            competence = str(body.get("competencia") or "").strip()
            if not department or not re.fullmatch(r"\d{4}-\d{2}", competence):
                return jsonify("Selecione exames ou informe departamento e mês de vencimento."), 400
            query = query.filter(
                cast(CostCenters.departamento, String) == department,
                db.func.to_char(PeriodicExam.data_vencimento, "YYYY-MM") == competence,
            )

        rows = query.all()
        if not rows:
            return jsonify("Nenhum exame foi encontrado para a conclusão em lote."), 404

        now = datetime.now(timezone.utc)
        for exam, *_ in rows:
            exam.status = status
            exam.concluido_em = now if status == "concluido" else None
            exam.concluido_por_usuario_id = token_data.get("id") if status == "concluido" else None
        db.session.commit()
        socketio.emit("periodic_exam_update", {"action": "bulk_updated", "total": len(rows)})
        return jsonify({"message": "Exames atualizados.", "total": len(rows)}), 200

    @safe_route
    def delete_all(self, token_data):
        if not is_admin(token_data):
            return jsonify("Apenas administradores podem excluir todos os exames."), 403
        affected = PeriodicExam.query.delete(synchronize_session=False)
        db.session.commit()
        socketio.emit("periodic_exam_update", {"action": "deleted_all", "total": affected})
        return jsonify({"message": "Todos os exames foram excluídos.", "total": affected}), 200

    @safe_route
    def export_spreadsheet(self, token_data):
        if not has_permission(token_data, "controle_exames_periodicos", "view"):
            return jsonify("Você não possui acesso à exportação."), 403
        self.refresh_pending_statuses()
        records = [self._serialize(*row) for row in self._filtered_rows(token_data)]
        if not records:
            return jsonify("Não há exames para exportar com os filtros atuais."), 404

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Exames periódicos"
        sheet.sheet_view.showGridLines = False
        dark, teal, light, white = "173925", "20A65A", "EAF6EF", "FFFFFF"
        thin = Side(style="thin", color="DDE7E1")
        sheet.merge_cells("A1:K2")
        sheet["A1"] = "CONTROLE DE EXAMES PERIÓDICOS"
        sheet["A1"].font = Font(size=20, bold=True, color=white)
        sheet["A1"].fill = PatternFill("solid", fgColor=dark)
        sheet["A1"].alignment = Alignment(vertical="center")
        headers = [
            "Colaborador", "Matrícula", "Departamento", "Contrato", "Supervisor",
            "Tipo", "Data do exame", "Vencimento", "Situação", "Resultado", "Observação",
        ]
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(4, column, label)
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_index, record in enumerate(records, 5):
            values = [
                record["colaborador"], record["matricula"], record["departamento"],
                record["centro_custo"], record["supervisor"], record["tipo_exame"],
                _as_date(record["data_exame"]), _as_date(record["data_vencimento"]),
                record["status"].replace("_", " ").title(), record["resultado"], record["observacao"],
            ]
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_index, column, value)
                cell.fill = PatternFill("solid", fgColor="FFFFFF" if row_index % 2 else "F5F9F7")
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=column in {1, 4, 5, 6, 11})
            for column in (7, 8):
                sheet.cell(row_index, column).number_format = "dd/mm/yyyy"
        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:K{sheet.max_row}"
        for index, width in enumerate([34, 14, 15, 42, 28, 22, 16, 16, 18, 18, 42], 1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[4].height = 32

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="controle_exames_periodicos.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

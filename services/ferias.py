# Biblioteca padrão.
from collections import defaultdict
from datetime import date, datetime as dt, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from unicodedata import normalize

# Dependências externas.
from dateutil.relativedelta import relativedelta
from flask import jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import String, cast, func, or_

# Módulos internos da aplicação.
from models.centros_de_custo import CostCenters
from models.colaboradores import Employees
from models.controle_faltas import AbsenceControl
from models.ferias import VacationLeave, VacationPeriod
from models.supervisores import Supervisors
from models.usuarios import Users
from utils.db import db
from utils.filial_scope import allowed_cost_center_ids, apply_cost_center_scope, is_admin
from utils.permissions import has_permission
from utils.safe_route import safe_route
from utils.socket import socketio

# Constantes de valores de VA para Londrina.
VA_FAIXAS_LONDRINA = (
    (0, 0, Decimal("900.00")),
    (1, 3, Decimal("810.00")),
    (4, 5, Decimal("720.00")),
    (6, None, Decimal("0.00")),
)
MONEY = Decimal("0.01")
# Situações elegíveis somente para o provisionamento.
# 1 = Trabalhando, 9 = Férias e 18 = afastamento inferior a 15 dias.
# Demitidos e afastamentos sem retorno previsto não entram na projeção.
ACTIVE_VACATION_STATUSES = (1, 9, 18)


def _normalize(value):
    """Normaliza textos da planilha sem depender de acentos ou maiúsculas."""
    raw = str(value or "").strip()
    return "".join(
        character
        for character in normalize("NFD", raw)
        if ord(character) < 0x300 or ord(character) > 0x36F
    ).upper()


def _cell(row, index):
    return row[index] if index < len(row) else None


def _parse_date(value, label):
    if isinstance(value, dt):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        raise ValueError(f"{label} não informada")
    raw = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return dt.strptime(raw, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"{label} inválida")


def _optional_date(value, label):
    return None if value in (None, "") else _parse_date(value, label)


def _money(value, label):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value.quantize(MONEY)
    try:
        raw = str(value).strip().replace("R$", "").replace(" ", "")
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        return Decimal(raw).quantize(MONEY)
    except (InvalidOperation, ValueError):
        raise ValueError(f"{label} inválido") from None


def _employee_code(value):
    try:
        return int(str(value).strip().replace(".0", ""))
    except (TypeError, ValueError):
        return None


def _is_vacation_continuation(row):
    """Confirma se a linha contém o complemento de um lançamento de férias."""
    return (
        _employee_code(_cell(row, 0)) is None
        and _cell(row, 10) not in (None, "")
        and _cell(row, 13) not in (None, "")
    )


def _filter_values(name, caster=str):
    raw = str(request.args.get(name) or "").strip()
    if not raw:
        return []
    return [caster(value) for value in raw.split(",") if str(value).strip()]


def _date_range_days(start, end):
    return max(1, (end - start).days + 1)


def _br_date(value):
    """Converte a data serializada da API para o padrão exibido no Excel."""
    parsed = _optional_date(value, "Data")
    return parsed.strftime("%d/%m/%Y") if parsed else ""


def _official_export_record(record):
    """Oculta ajustes manuais no relatório oficial de férias da empresa."""
    exported = dict(record)
    if exported.get("ajustado_manual"):
        exported["dias_gozados"] = int(exported.get("dias_direito") or 30)
        exported["dias_a_gozar"] = 0
        exported["situacao"] = "concluida"
    return exported


class VacationService:
    """Centraliza histórico, provisionamento e gestão manual de férias."""

    @staticmethod
    def _base_query():
        return (
            db.session.query(
                VacationPeriod,
                Employees.id.label("colaborador_id"),
                Employees.nome.label("colaborador_nome"),
                Employees.matricula.label("matricula"),
                Employees.situacao.label("situacao_colaborador"),
                Employees.data_admissao.label("data_admissao"),
                Employees.centro_id.label("centro_custo_id"),
                CostCenters.local.label("centro_custo"),
                CostCenters.departamento.label("departamento"),
                Supervisors.id.label("supervisor_id"),
                Supervisors.nome.label("supervisor"),
                Users.nome.label("ajustado_por"),
            )
            .join(Employees, Employees.id == VacationPeriod.colaborador_id)
            .outerjoin(CostCenters, CostCenters.id == Employees.centro_id)
            .outerjoin(Supervisors, Supervisors.id == CostCenters.supervisor_id)
            .outerjoin(Users, Users.id == VacationPeriod.ajustado_por_usuario_id)
        )

    @staticmethod
    def _apply_history_scope(query, center_column, token_data):
        """Aplica apenas o escopo de filial ao histórico já consolidado."""
        allowed_centers = allowed_cost_center_ids(token_data)
        return query if allowed_centers is None else query.filter(center_column.in_(allowed_centers))

    @staticmethod
    def _absence_count(period):
        """Conta faltas lançadas no período aquisitivo usado para o VA."""
        if not period.colaborador_id:
            return 0
        return (
            db.session.query(func.count(AbsenceControl.id))
            .filter(
                AbsenceControl.colaborador_id == period.colaborador_id,
                func.date(AbsenceControl.data_falta) >= period.periodo_aquisitivo_inicio,
                func.date(AbsenceControl.data_falta) <= period.periodo_aquisitivo_fim,
            )
            .scalar()
            or 0
        )

    @staticmethod
    def _va_value(absence_count):
        for minimum, maximum, value in VA_FAIXAS_LONDRINA:
            if absence_count >= minimum and (maximum is None or absence_count <= maximum):
                return value
        return Decimal("0")

    @staticmethod
    def _status(period, days_gozados, reference_date=None):
        """Calcula a situação atual sem gravar estados que envelhecem no banco."""
        today = reference_date or date.today()
        remaining = max(0, int(period.dias_direito or 30) - days_gozados)
        if remaining == 0:
            return "concluida"

        # Um ajuste manual indica que a folha pagou o período completo, mas o
        # colaborador não gozou todos os dias. Esse saldo precisa ser tratado
        # pelo RH como pendência de gozo, e não como férias em dobro.
        if getattr(period, "ajustado_manual", False):
            if today >= period.limite_concessivo - timedelta(days=90):
                return "a_vencer"
            return "parcial"

        if today > period.limite_concessivo:
            return "em_dobro"
        if today >= period.limite_concessivo - timedelta(days=30):
            return "critica"
        if today >= period.limite_concessivo - timedelta(days=90):
            return "a_vencer"
        return "parcial" if days_gozados else "disponivel"

    @staticmethod
    def _leave_data(leave):
        return {
            "id": leave.id,
            "data_inicio": leave.data_inicio.isoformat(),
            "data_fim": leave.data_fim.isoformat(),
            "dias_gozados": leave.dias_gozados,
            "dias_calculados_pagos": leave.dias_calculados_pagos,
            "pagamento_realizado": bool(leave.pagamento_realizado),
            "observacao": leave.observacao,
        }

    def _serialize(self, row):
        period = row.VacationPeriod
        leaves = list(period.gozos)
        days_gozados = sum(int(leave.dias_gozados or 0) for leave in leaves)
        days_right = int(period.dias_direito or 30)
        absence_count = self._absence_count(period)
        return {
            "id": period.id,
            "colaborador_id": row.colaborador_id,
            "matricula": row.matricula,
            "nome": row.colaborador_nome,
            "data_admissao": row.data_admissao.date().isoformat() if isinstance(row.data_admissao, dt) else row.data_admissao.isoformat() if row.data_admissao else None,
            "centro_custo_id": row.centro_custo_id,
            "centro_custo": row.centro_custo or "Sem centro de custo",
            "departamento": row.departamento,
            "supervisor_id": row.supervisor_id,
            "supervisor": row.supervisor or "Sem supervisor",
            "periodo_aquisitivo_inicio": period.periodo_aquisitivo_inicio.isoformat(),
            "periodo_aquisitivo_fim": period.periodo_aquisitivo_fim.isoformat(),
            "limite_concessivo": period.limite_concessivo.isoformat(),
            "dias_direito": days_right,
            "dias_gozados": days_gozados,
            "dias_a_gozar": max(0, days_right - days_gozados),
            # A situação 8 indica desligamento. O histórico é preservado,
            # mas não deve permanecer com pendência de programação.
            "situacao": "concluida" if row.situacao_colaborador == 8 else self._status(period, days_gozados),
            "pagamento_ferias_integral": bool(period.pagamento_ferias_integral),
            "va_ferias_integral_pago": bool(period.va_ferias_integral_pago),
            "faltas_va": absence_count,
            "valor_va_ferias": float(self._va_value(absence_count)),
            "valores": {
                "ferias": float(period.valor_ferias or 0),
                "terco_ferias": float(period.valor_terco_ferias or 0),
                "complementar": float(period.valor_complementar or 0),
                "descontos": float(period.valor_descontos or 0),
                "liquido": float(period.valor_liquido or 0),
            },
            "gozos": [self._leave_data(leave) for leave in leaves],
            "arquivo_origem": period.arquivo_origem,
            "observacao_manual": period.observacao_manual,
            "ajustado_manual": bool(period.ajustado_manual),
            "ajustado_por": row.ajustado_por,
            "ajustado_em": period.ajustado_em.isoformat() if period.ajustado_em else None,
        }

    def _period_rows(self, token_data):
        query = self._apply_history_scope(
            self._base_query(),
            Employees.centro_id,
            token_data,
        )
        try:
            center_ids = _filter_values("centro_custo_id", int)
            department_ids = _filter_values("departamento", int)
            supervisor_ids = _filter_values("supervisor_id", int)
            if center_ids:
                query = query.filter(Employees.centro_id.in_(center_ids))
            if department_ids:
                query = query.filter(CostCenters.departamento.in_(department_ids))
            if supervisor_ids:
                query = query.filter(Supervisors.id.in_(supervisor_ids))
            if request.args.get("inicio"):
                query = query.filter(VacationPeriod.periodo_aquisitivo_inicio >= _parse_date(request.args["inicio"], "Data inicial"))
            if request.args.get("fim"):
                query = query.filter(VacationPeriod.periodo_aquisitivo_inicio <= _parse_date(request.args["fim"], "Data final"))
        except ValueError as error:
            return None, str(error)

        search = str(request.args.get("busca") or "").strip()
        if search:
            pattern = f"%{search}%"
            query = query.filter(or_(
                Employees.nome.ilike(pattern),
                cast(Employees.matricula, String).ilike(pattern),
                CostCenters.local.ilike(pattern),
                Supervisors.nome.ilike(pattern),
            ))
        return query.order_by(VacationPeriod.periodo_aquisitivo_inicio.desc()).all(), None

    def _provisioning(self, token_data, existing_records):
        """Monta períodos futuros e pendentes sem persistir uma fotografia diária."""
        existing = {
            (item["colaborador_id"], item["periodo_aquisitivo_inicio"]): item
            for item in existing_records
        }
        employees_query = apply_cost_center_scope(Employees.query, Employees.centro_id, token_data).outerjoin(
            CostCenters,
            CostCenters.id == Employees.centro_id,
        ).outerjoin(
            Supervisors,
            Supervisors.id == CostCenters.supervisor_id,
        ).filter(
            Employees.situacao.in_(ACTIVE_VACATION_STATUSES),
            Employees.data_admissao.isnot(None),
        )
        center_ids = _filter_values("centro_custo_id", int)
        department_ids = _filter_values("departamento", int)
        supervisor_ids = _filter_values("supervisor_id", int)
        if center_ids:
            employees_query = employees_query.filter(Employees.centro_id.in_(center_ids))
        if department_ids:
            employees_query = employees_query.filter(CostCenters.departamento.in_(department_ids))
        if supervisor_ids:
            employees_query = employees_query.filter(Supervisors.id.in_(supervisor_ids))

        search = str(request.args.get("busca") or "").strip()
        if search:
            pattern = f"%{search}%"
            employees_query = employees_query.filter(or_(
                Employees.nome.ilike(pattern),
                cast(Employees.matricula, String).ilike(pattern),
                CostCenters.local.ilike(pattern),
                Supervisors.nome.ilike(pattern),
            ))
        today = date.today()
        rows = []
        for employee in employees_query.all():
            admission = employee.data_admissao.date() if isinstance(employee.data_admissao, dt) else employee.data_admissao
            if not admission:
                continue
            period_start = admission
            while period_start <= today:
                acquisition_end = period_start + relativedelta(years=1) - timedelta(days=1)
                if acquisition_end > today:
                    break
                deadline = period_start + relativedelta(years=2) - timedelta(days=1)
                key = (employee.id, period_start.isoformat())
                saved = existing.get(key)
                days_gozados = saved["dias_gozados"] if saved else 0
                remaining = max(0, 30 - days_gozados)
                virtual_period = type("ProvisionPeriod", (), {
                    "dias_direito": 30,
                    "limite_concessivo": deadline,
                    "ajustado_manual": bool(saved and saved["ajustado_manual"]),
                })()
                status = self._status(virtual_period, days_gozados, today)
                if remaining:
                    rows.append({
                        "colaborador_id": employee.id,
                        "matricula": employee.matricula,
                        "nome": employee.nome,
                        "centro_custo_id": employee.centro_id,
                        "periodo_aquisitivo_inicio": period_start.isoformat(),
                        "periodo_aquisitivo_fim": acquisition_end.isoformat(),
                        "limite_concessivo": deadline.isoformat(),
                        "dias_direito": 30,
                        "dias_gozados": days_gozados,
                        "dias_a_gozar": remaining,
                        "situacao": status,
                        "periodo_id": saved["id"] if saved else None,
                        "ajustado_manual": bool(saved and saved["ajustado_manual"]),
                    })
                period_start += relativedelta(years=1)
        return sorted(rows, key=lambda item: (item["limite_concessivo"], item["nome"] or ""))

    @safe_route
    def read(self, token_data):
        if not has_permission(token_data, "controle_ferias", "view"):
            return jsonify("Você não possui acesso ao Controle de Férias."), 403
        rows, error = self._period_rows(token_data)
        if error:
            return jsonify(error), 400
        all_records = [self._serialize(row) for row in rows]
        status_filter = _filter_values("situacao")
        provisioning = self._provisioning(token_data, all_records)
        records = all_records
        if status_filter:
            records = [item for item in all_records if item["situacao"] in status_filter]
            provisioning = [item for item in provisioning if item["situacao"] in status_filter]
        filters = {
            "departamentos": sorted({item["departamento"] for item in all_records if item["departamento"] is not None}),
            "centros": sorted(
                [{"label": item["centro_custo"], "value": item["centro_custo_id"]} for item in {item["centro_custo_id"]: item for item in all_records}.values()],
                key=lambda item: item["label"].casefold(),
            ),
            # Somente situações presentes no recorte atual podem ser
            # selecionadas; a lista não é um catálogo fixo de estados.
            "situacoes": sorted({item["situacao"] for item in all_records if item.get("situacao")}),
            "supervisores": sorted(
                [{"label": item["supervisor"], "value": item["supervisor_id"]} for item in {item["supervisor_id"]: item for item in all_records if item["supervisor_id"]}.values()],
                key=lambda item: item["label"].casefold(),
            ),
        }
        summary = {
            "historico": len(records),
            "provisionamento": len(provisioning),
            "a_vencer": sum(item["situacao"] == "a_vencer" for item in provisioning),
            "criticas": sum(item["situacao"] in {"critica", "em_dobro"} for item in provisioning),
            "custo_pago": round(sum(item["valores"]["liquido"] for item in records), 2),
        }
        return jsonify({
            "registros": records,
            "provisionamento": provisioning,
            "resumo": summary,
            "filtros": filters,
        }), 200

    @staticmethod
    def _parse_import_rows(uploaded):
        """Extrai apenas períodos completos do relatório, sem consultar ou gravar no banco."""
        try:
            workbook = load_workbook(uploaded.stream, read_only=True, data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
        except (OSError, ValueError):
            return [], ["Não foi possível ler a planilha."]

        header_index = next(
            (
                index for index, row in enumerate(rows)
                if _normalize(_cell(row, 0)) == "CODIGO" and "EMPREGADO" in _normalize(_cell(row, 2))
            ),
            None,
        )
        if header_index is None:
            return [], ["Planilha fora do padrão: cabeçalho Código/Nome do empregado não encontrado."]

        parsed = []
        errors = []
        for index in range(header_index + 1, len(rows)):
            row = rows[index]
            code = _employee_code(_cell(row, 0))
            if code is None:
                continue
            next_row = rows[index + 1] if index + 1 < len(rows) else ()

            # O relatório imprime cada lançamento em duas linhas. Uma quebra
            # de página não pode transformar uma linha isolada em férias
            # válidas, pois isso perderia dias e valores do período.
            if not _is_vacation_continuation(next_row):
                errors.append(
                    f"Linha {index + 1}, matrícula {code}: período incompleto antes do rodapé; informe a linha de fim no relatório de origem."
                )
                continue

            try:
                acquisition_start = _parse_date(_cell(row, 10), "Início do aquisitivo")
                acquisition_end = _parse_date(_cell(next_row, 10), "Fim do aquisitivo")
                leave_start = _parse_date(_cell(row, 13), "Início das férias")
                leave_end = _parse_date(_cell(next_row, 13), "Fim das férias")
                if acquisition_end < acquisition_start or leave_end < leave_start:
                    raise ValueError("as datas de início e fim são incompatíveis")
                vacation_item = {
                    "linha": index + 1,
                    "matricula": code,
                    "nome": str(_cell(row, 2) or "").strip(),
                    "aquisitivo_inicio": acquisition_start,
                    "aquisitivo_fim": acquisition_end,
                    "ferias_inicio": leave_start,
                    "ferias_fim": leave_end,
                    "valor_ferias": _money(_cell(row, 17), "Valor de férias"),
                    "valor_complementar": _money(_cell(row, 19), "Valor complementar"),
                    "valor_terco": _money(_cell(row, 21), "1/3 de férias"),
                    "valor_descontos": _money(_cell(next_row, 21), "Desconto previdenciário") + _money(_cell(next_row, 25), "Desconto de IRRF") + _money(_cell(next_row, 28), "Outros descontos"),
                    "valor_liquido": _money(_cell(next_row, 33), "Líquido de férias"),
                    "origem_gozo": "ferias",
                }
                parsed.append(vacation_item)

                # No relatório utilizado pela operação, a coluna "Abono"
                # representa uma segunda fração disponível para gozo.
                abono_start = _optional_date(_cell(row, 15), "Início do período complementar")
                abono_end = _optional_date(_cell(next_row, 15), "Fim do período complementar")
                if bool(abono_start) != bool(abono_end):
                    raise ValueError("o período complementar está incompleto")
                if abono_start and abono_end:
                    if abono_end < abono_start:
                        raise ValueError("as datas do período complementar são incompatíveis")
                    parsed.append({
                        **vacation_item,
                        "ferias_inicio": abono_start,
                        "ferias_fim": abono_end,
                        # Os valores já pertencem ao lançamento principal;
                        # zerá-los impede que sejam somados duas vezes.
                        "valor_ferias": Decimal("0"),
                        "valor_complementar": Decimal("0"),
                        "valor_terco": Decimal("0"),
                        "valor_descontos": Decimal("0"),
                        "valor_liquido": Decimal("0"),
                        "origem_gozo": "periodo_complementar",
                    })
            except ValueError as error:
                errors.append(f"Linha {index + 1}: {error}.")

        return parsed, errors[:50]

    @staticmethod
    def _resolve_import_employees(parsed, token_data):
        """Vincula as matrículas em lote e preserva apenas o escopo de filial."""
        matriculas = {item["matricula"] for item in parsed}
        employees_by_registration = defaultdict(list)
        for employee in Employees.query.filter(Employees.matricula.in_(matriculas)).all():
            employees_by_registration[employee.matricula].append(employee)

        errors = []
        employees = {}
        for registration in matriculas:
            options = employees_by_registration.get(registration, [])
            if len(options) != 1:
                errors.append(f"Matrícula {registration}: colaborador não encontrado ou ambíguo.")
                continue
            employees[registration] = options[0]

        allowed_centers = allowed_cost_center_ids(token_data)
        if allowed_centers is not None:
            errors.extend(
                f"Matrícula {registration}: filial fora do seu acesso."
                for registration, employee in employees.items()
                if not employee.centro_id or employee.centro_id not in allowed_centers
            )
        return employees, errors[:50]

    @staticmethod
    def _group_import_rows(parsed, employees):
        grouped = defaultdict(list)
        for item in parsed:
            employee = employees[item["matricula"]]
            grouped[(employee.id, item["aquisitivo_inicio"])].append(item)
        return grouped

    @staticmethod
    def _existing_periods(grouped):
        employee_ids = {employee_id for employee_id, _ in grouped}
        acquisition_starts = {start for _, start in grouped}
        return {
            (period.colaborador_id, period.periodo_aquisitivo_inicio): period
            for period in VacationPeriod.query.filter(
                VacationPeriod.colaborador_id.in_(employee_ids),
                VacationPeriod.periodo_aquisitivo_inicio.in_(acquisition_starts),
            ).all()
        }

    @safe_route
    def preview_import(self, token_data):
        """Monta uma prévia da importação sem alterar períodos de férias."""
        if not has_permission(token_data, "controle_ferias", "create"):
            return jsonify("Você não possui permissão para importar férias."), 403
        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify("Envie a planilha padrão no formato .xlsx."), 400

        parsed, errors = self._parse_import_rows(uploaded)
        if errors:
            return jsonify({"message": "A prévia identificou dados inconsistentes na planilha.", "errors": errors}), 400
        if not parsed:
            return jsonify("Nenhuma férias completa foi encontrada na planilha."), 400

        employees, errors = self._resolve_import_employees(parsed, token_data)
        if errors:
            return jsonify({"message": "A prévia não pôde ser concluída.", "errors": errors}), 400

        grouped = self._group_import_rows(parsed, employees)
        existing = self._existing_periods(grouped)
        total_periods = len(grouped)
        total_updates = sum(1 for key in grouped if key in existing)
        leave_dates = [item["ferias_inicio"] for item in parsed]
        leave_end_dates = [item["ferias_fim"] for item in parsed]
        return jsonify({
            "total_periodos": total_periods,
            "total_lancamentos": len(parsed),
            "resumo": {
                "colaboradores": len({employee_id for employee_id, _ in grouped}),
                "novos_periodos": total_periods - total_updates,
                "periodos_atualizados": total_updates,
                "gozos_fracionados": sum(1 for items in grouped.values() if len(items) > 1),
                "primeiro_gozo": min(leave_dates).isoformat(),
                "ultimo_gozo": max(leave_end_dates).isoformat(),
            },
        }), 200

    @safe_route
    def import_xlsx(self, token_data):
        if not has_permission(token_data, "controle_ferias", "create"):
            return jsonify("Você não possui permissão para importar férias."), 403
        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify("Envie a planilha padrão no formato .xlsx."), 400
        parsed, errors = self._parse_import_rows(uploaded)
        if errors:
            return jsonify({"message": "A importação foi cancelada; corrija a planilha.", "errors": errors}), 400
        if not parsed:
            return jsonify("Nenhuma férias completa foi encontrada na planilha."), 400

        employees, errors = self._resolve_import_employees(parsed, token_data)
        if errors:
            return jsonify({"message": "A importação foi cancelada; nenhum registro foi gravado.", "errors": errors}), 400

        grouped = self._group_import_rows(parsed, employees)
        existing = self._existing_periods(grouped)
        source_name = Path(uploaded.filename).name[:255]
        created = updated = 0
        for key, items in grouped.items():
            employee = employees[items[0]["matricula"]]
            period = existing.get(key)
            if period and period.ajustado_manual:
                return jsonify(
                    f"A matrícula {employee.matricula} possui ajuste manual no período iniciado em {key[1].strftime('%d/%m/%Y')}."
                ), 409
            if not period:
                period = VacationPeriod(
                    colaborador_id=employee.id,
                    periodo_aquisitivo_inicio=items[0]["aquisitivo_inicio"],
                )
                db.session.add(period)
                created += 1
            else:
                period.gozos.clear()
                updated += 1
            first = items[0]
            period.periodo_aquisitivo_fim = first["aquisitivo_fim"]
            period.limite_concessivo = first["aquisitivo_inicio"] + relativedelta(years=2) - timedelta(days=1)
            period.dias_direito = 30
            period.pagamento_ferias_integral = True
            period.va_ferias_integral_pago = True
            period.valor_ferias = sum((item["valor_ferias"] for item in items), Decimal("0"))
            period.valor_terco_ferias = sum((item["valor_terco"] for item in items), Decimal("0"))
            period.valor_complementar = sum((item["valor_complementar"] for item in items), Decimal("0"))
            period.valor_descontos = sum((item["valor_descontos"] for item in items), Decimal("0"))
            period.valor_liquido = sum((item["valor_liquido"] for item in items), Decimal("0"))
            period.arquivo_origem = source_name
            period.importado_por_usuario_id = token_data.get("id")
            period.ajustado_manual = False
            period.observacao_manual = None
            period.ajustado_por_usuario_id = None
            period.ajustado_em = None

            for item in items:
                # Cada intervalo informado na planilha representa um gozo
                # independente. Dessa forma, três frações de 10 dias são
                # preservadas como três lançamentos, sem fechar o período
                # antecipadamente como se fossem 30 dias em um único gozo.
                days = _date_range_days(item["ferias_inicio"], item["ferias_fim"])
                period.gozos.append(VacationLeave(
                    data_inicio=item["ferias_inicio"],
                    data_fim=item["ferias_fim"],
                    dias_gozados=days,
                    dias_calculados_pagos=days,
                    pagamento_realizado=True,
                    observacao=(
                        "Importado como período complementar da relação de férias calculadas."
                        if item.get("origem_gozo") == "periodo_complementar"
                        else "Importado da relação de férias calculadas."
                    ),
                ))

        db.session.commit()
        socketio.emit("vacation_update", {"action": "imported", "created": created, "updated": updated})
        message = f"{created + updated} período(s) de férias processado(s)."
        return jsonify({
            "message": message,
            "criadas": created,
            "atualizadas": updated,
        }), 201

    def _period_in_scope(self, period_id, token_data):
        period = db.session.get(VacationPeriod, period_id)
        if not period:
            return None, (jsonify("Período de férias não encontrado."), 404)
        employee = db.session.get(Employees, period.colaborador_id)
        allowed_centers = allowed_cost_center_ids(token_data)
        if (
            not employee
            or not employee.centro_id
            or (allowed_centers is not None and employee.centro_id not in allowed_centers)
        ):
            return None, (jsonify("Você não possui acesso a este período de férias."), 403)
        return period, None

    @staticmethod
    def _mark_manual(period, token_data):
        period.ajustado_manual = True
        period.ajustado_por_usuario_id = token_data.get("id")
        period.ajustado_em = dt.now()

    @staticmethod
    def _remaining_days(period, ignore_leave_id=None):
        total = sum(
            int(leave.dias_gozados or 0)
            for leave in period.gozos
            if leave.id != ignore_leave_id
        )
        return max(0, int(period.dias_direito or 30) - total)

    @safe_route
    def update_period(self, period_id, token_data):
        if not has_permission(token_data, "controle_ferias", "edit"):
            return jsonify("Você não possui permissão para alterar férias."), 403
        period, error = self._period_in_scope(period_id, token_data)
        if error:
            return error
        body = request.get_json(silent=True) or {}
        observation = str(body.get("observacao_manual") or "").strip()
        if len(observation) > 1000:
            return jsonify("A observação pode ter no máximo 1000 caracteres."), 400
        period.observacao_manual = observation or None
        self._mark_manual(period, token_data)
        db.session.commit()
        socketio.emit("vacation_update", {"action": "updated", "period_id": period.id})
        return jsonify({"message": "Período atualizado."}), 200

    def _save_leave(self, leave, period, body, token_data, is_new=False):
        start = _parse_date(body.get("data_inicio") or leave.data_inicio, "Data inicial")
        raw_days = body.get("dias_gozados")
        end = _optional_date(body.get("data_fim"), "Data final")
        if raw_days not in (None, ""):
            try:
                days = int(raw_days)
            except (TypeError, ValueError):
                raise ValueError("Dias gozados deve ser um número inteiro") from None
            end = start + timedelta(days=days - 1)
        elif end:
            days = _date_range_days(start, end)
        else:
            days = leave.dias_gozados
            end = start + timedelta(days=days - 1)
        if days < 1 or days > 30:
            raise ValueError("Dias gozados deve estar entre 1 e 30")
        if days > self._remaining_days(period, None if is_new else leave.id):
            raise ValueError("Os dias informados ultrapassam o saldo disponível do período")
        leave.data_inicio = start
        leave.data_fim = end
        leave.dias_gozados = days
        leave.dias_calculados_pagos = int(body.get("dias_calculados_pagos") or leave.dias_calculados_pagos or 0)
        leave.pagamento_realizado = bool(body.get("pagamento_realizado", leave.pagamento_realizado))
        observation = str(body.get("observacao") or leave.observacao or "").strip()
        if len(observation) > 1000:
            raise ValueError("A observação pode ter no máximo 1000 caracteres")
        leave.observacao = observation or None
        self._mark_manual(period, token_data)

    @safe_route
    def create_leave(self, period_id, token_data):
        if not has_permission(token_data, "controle_ferias", "edit"):
            return jsonify("Você não possui permissão para registrar férias."), 403
        period, error = self._period_in_scope(period_id, token_data)
        if error:
            return error
        body = request.get_json(silent=True) or {}
        leave = VacationLeave(periodo_id=period.id)
        try:
            self._save_leave(leave, period, body, token_data, is_new=True)
        except ValueError as error:
            return jsonify(str(error)), 400
        db.session.add(leave)
        db.session.commit()
        socketio.emit("vacation_update", {"action": "leave_created", "period_id": period.id})
        return jsonify(self._leave_data(leave)), 201

    @safe_route
    def update_leave(self, leave_id, token_data):
        if not has_permission(token_data, "controle_ferias", "edit"):
            return jsonify("Você não possui permissão para alterar férias."), 403
        leave = db.session.get(VacationLeave, leave_id)
        if not leave:
            return jsonify("Lançamento de férias não encontrado."), 404
        period, error = self._period_in_scope(leave.periodo_id, token_data)
        if error:
            return error
        try:
            self._save_leave(leave, period, request.get_json(silent=True) or {}, token_data)
        except ValueError as error:
            return jsonify(str(error)), 400
        db.session.commit()
        socketio.emit("vacation_update", {"action": "leave_updated", "period_id": period.id})
        return jsonify(self._leave_data(leave)), 200

    @safe_route
    def delete_leave(self, leave_id, token_data):
        if not has_permission(token_data, "controle_ferias", "edit"):
            return jsonify("Você não possui permissão para excluir férias."), 403
        leave = db.session.get(VacationLeave, leave_id)
        if not leave:
            return jsonify("Lançamento de férias não encontrado."), 404
        period, error = self._period_in_scope(leave.periodo_id, token_data)
        if error:
            return error
        self._mark_manual(period, token_data)
        db.session.delete(leave)
        db.session.commit()
        socketio.emit("vacation_update", {"action": "leave_deleted", "period_id": period.id})
        return jsonify("Lançamento de férias excluído."), 200

    @safe_route
    def complete_periods(self, token_data):
        if not has_permission(token_data, "controle_ferias", "edit"):
            return jsonify("Você não possui permissão para concluir férias."), 403
        body = request.get_json(silent=True) or {}
        ids = body.get("periodo_ids") or []
        if not isinstance(ids, list) or not ids:
            return jsonify("Selecione pelo menos um período."), 400
        periods = []
        for raw_id in ids:
            try:
                period_id = int(raw_id)
            except (TypeError, ValueError):
                return jsonify("Um ou mais períodos são inválidos."), 400
            period, error = self._period_in_scope(period_id, token_data)
            if error:
                return error
            if self._remaining_days(period) > 0:
                return jsonify("Não é possível concluir um período que ainda possui dias a gozar."), 409
            periods.append(period)
        for period in periods:
            self._mark_manual(period, token_data)
        db.session.commit()
        socketio.emit("vacation_update", {"action": "completed", "total": len(periods)})
        return jsonify({"message": f"{len(periods)} período(s) confirmado(s)."}), 200

    @safe_route
    def delete_all(self, token_data):
        if not is_admin(token_data):
            return jsonify("Apenas administradores podem excluir todas as férias."), 403
        affected = VacationPeriod.query.delete(synchronize_session=False)
        db.session.commit()
        socketio.emit("vacation_update", {"action": "deleted_all", "total": affected})
        return jsonify({"message": f"{affected} período(s) de férias excluído(s)."}), 200

    @safe_route
    def export_xlsx(self, token_data):
        if not has_permission(token_data, "controle_ferias", "view"):
            return jsonify("Você não possui permissão para exportar férias."), 403
        rows, error = self._period_rows(token_data)
        if error:
            return jsonify(error), 400
        source_records = [self._serialize(row) for row in rows]
        all_records = [_official_export_record(item) for item in source_records]
        status_filter = _filter_values("situacao")
        provisioning = self._provisioning(token_data, source_records)
        provisioning = [item for item in provisioning if not item.get("ajustado_manual")]
        records = all_records
        if status_filter:
            records = [item for item in all_records if item["situacao"] in status_filter]
            provisioning = [item for item in provisioning if item["situacao"] in status_filter]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Férias"
        headers = [
            "Origem", "Matrícula", "Colaborador", "Centro de custo", "Departamento", "Supervisor",
            "Início aquisitivo", "Fim aquisitivo", "Limite concessivo", "Dias de direito",
            "Dias gozados", "Dias a gozar", "Situação", "Faltas para VA", "VA nas férias",
            "Valor férias", "1/3 de férias", "Descontos", "Líquido",
        ]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
            cell.alignment = Alignment(horizontal="center")
        for item in records + provisioning:
            values = item.get("valores", {})
            worksheet.append([
                "Histórico importado" if item in records else "Provisionamento",
                item.get("matricula"), item.get("nome"), item.get("centro_custo"), item.get("departamento"), item.get("supervisor"),
                _br_date(item["periodo_aquisitivo_inicio"]), _br_date(item["periodo_aquisitivo_fim"]), _br_date(item["limite_concessivo"]), item["dias_direito"],
                item["dias_gozados"], item["dias_a_gozar"], item["situacao"], item.get("faltas_va", 0), item.get("valor_va_ferias", 0),
                values.get("ferias", 0), values.get("terco_ferias", 0), values.get("descontos", 0), values.get("liquido", 0),
            ])
        for column in worksheet.columns:
            letter = column[0].column_letter
            worksheet.column_dimensions[letter].width = min(34, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="controle_de_ferias.xlsx",
        )

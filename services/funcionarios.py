"""Leitura e edição administrativa segura de colaboradores."""
from __future__ import annotations

from io import BytesIO

from flask import jsonify, request as rq, send_file
from sqlalchemy import String, cast, func, or_

from models.cargos import Cargos
from models.centros_de_custo import CostCenters
from models.colaboradores import Employees
from models.empresas import Company
from models.situacoes import Situations
from utils.db import db
from utils.filial_scope import apply_active_department_scope, apply_cost_center_scope, is_admin
from utils.safe_route import safe_route
from utils.socket import socketio


def _ids(value):
    values = []
    for item in str(value or "").split(","):
        try: values.append(int(item))
        except (TypeError, ValueError): pass
    return list(dict.fromkeys(values))


class EmployeesService:
    @staticmethod
    def _query(include_cpf=False):
        columns = [
            Employees.id, Employees.matricula, Employees.nome,
            Employees.data_admissao, Employees.carga_horaria, Employees.empresa_id,
            Company.nome.label("empresa_nome"), Employees.centro_id,
            CostCenters.centro_id.label("centro_numero"),
            CostCenters.local.label("centro_local"), CostCenters.departamento,
            Cargos.id.label("cargo_id"), Cargos.nome.label("cargo"),
            Situations.id.label("situacao_id"), Situations.tipo.label("situacao"),
        ]
        if include_cpf:
            columns.insert(3, Employees.cpf)
        return db.session.query(*columns).select_from(Employees).outerjoin(Company, Company.id == Employees.empresa_id).outerjoin(
            Cargos, Cargos.id == Employees.cargo).outerjoin(Situations, Situations.id == Employees.situacao).outerjoin(
            CostCenters, CostCenters.id == Employees.centro_id)

    @staticmethod
    def _apply_filters(query):
        args = rq.args
        search = " ".join(str(args.get("search") or "").split())
        if search:
            pattern = f"%{search}%"
            query = query.filter(or_(Employees.nome.ilike(pattern), cast(Employees.matricula, String).ilike(pattern),
                                     Cargos.nome.ilike(pattern), CostCenters.local.ilike(pattern), Company.nome.ilike(pattern)))
        centers = _ids(args.get("centro_id") or args.get("centro_ids") or args.get("centro_custo_id"))
        if centers: query = query.filter(Employees.centro_id.in_(centers))
        departments = _ids(args.get("departamento") or args.get("departamentos"))
        if departments: query = query.filter(CostCenters.departamento.in_(departments))
        companies = _ids(args.get("empresa_id") or args.get("empresa_ids"))
        if companies: query = query.filter(Employees.empresa_id.in_(companies))
        cargos = _ids(args.get("cargo_id") or args.get("cargo_ids"))
        if cargos: query = query.filter(Employees.cargo.in_(cargos))
        situations = _ids(args.get("situacao") or args.get("situacao_ids"))
        if situations: query = query.filter(Employees.situacao.in_(situations))
        if args.get("excluir_id", type=int): query = query.filter(Employees.id != args.get("excluir_id", type=int))
        if str(args.get("com_local") or "").lower() in {"1", "true", "sim"}: query = query.filter(CostCenters.id.isnot(None))
        return query

    @safe_route
    def read(self, token_data):
        if str(rq.args.get("fields") or "").lower() == "tm_ops": return self._read_tm_ops_lookup(token_data)
        include_cpf = str(rq.args.get("include_cpf") or "").lower() in {"1", "true"} and is_admin(token_data)
        query = apply_active_department_scope(self._query(include_cpf=include_cpf), Employees.centro_id)
        query = apply_cost_center_scope(query, Employees.centro_id, token_data)
        query = self._apply_filters(query).order_by(Employees.nome.asc(), Employees.id.asc())
        if str(rq.args.get("paginado") or "").lower() in {"1", "true"}:
            page = max(rq.args.get("page", 1, type=int), 1); per_page = min(max(rq.args.get("per_page", 25, type=int), 1), 100)
            pagination = query.paginate(page=page, per_page=per_page, error_out=False)
            return jsonify({"items": [row._asdict() for row in pagination.items], "page": pagination.page,
                            "per_page": pagination.per_page, "total": pagination.total, "pages": pagination.pages}), 200
        limit = rq.args.get("limit", type=int)
        return jsonify([row._asdict() for row in query.limit(min(max(limit, 1), 500)) .all()] if limit else [row._asdict() for row in query.all()]), 200

    @safe_route
    def filters(self, token_data):
        query = apply_cost_center_scope(self._query(), Employees.centro_id, token_data)
        rows = query.distinct().all()
        def options(values, key, label):
            pairs = {(row._mapping[key], row._mapping[label]) for row in values if row._mapping[key] is not None}
            return [{"value": value, "label": label} for value, label in sorted(pairs, key=lambda item: str(item[1]))]
        return jsonify({
            "empresas": options(rows, "empresa_id", "empresa_nome"),
            "departamentos": [{"value": value, "label": f"DPTO. {value}"} for value in sorted({row.departamento for row in rows if row.departamento is not None})],
            "cargos": options(rows, "cargo_id", "cargo"),
            "situacoes": options(rows, "situacao_id", "situacao"),
        }), 200

    @safe_route
    def export(self, token_data):
        """Gera XLSX somente com os campos operacionais visíveis na tela."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        query = apply_cost_center_scope(self._query(), Employees.centro_id, token_data)
        rows = self._apply_filters(query).order_by(Employees.nome.asc()).all()
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Colaboradores"
        headers = ["Matrícula", "Colaborador", "Situação", "Cargo", "Empresa", "Departamento", "Centro de custo", "Admissão"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="087A42")
        for row in rows:
            sheet.append([row.matricula, row.nome, row.situacao, row.cargo, row.empresa_nome,
                          row.departamento, f"{row.centro_numero or '—'} - {row.centro_local or 'Sem local'}",
                          row.data_admissao.date() if hasattr(row.data_admissao, "date") else row.data_admissao])
        for column, width in {"A": 16, "B": 38, "C": 20, "D": 28, "E": 24, "F": 16, "G": 56, "H": 15}.items(): sheet.column_dimensions[column].width = width
        sheet.auto_filter.ref = f"A1:H{max(len(rows) + 1, 1)}"; sheet.freeze_panes = "A2"
        stream = BytesIO(); workbook.save(stream); stream.seek(0)
        return send_file(stream, as_attachment=True, download_name="colaboradores.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @safe_route
    def update(self, employee_id, token_data):
        if not is_admin(token_data): return jsonify("Apenas administradores podem editar colaboradores."), 403
        employee = db.session.get(Employees, employee_id)
        if not employee: return jsonify("Colaborador não encontrado."), 404
        body = rq.get_json(silent=True) or {}
        try:
            if "nome" in body:
                name = " ".join(str(body["nome"] or "").split()).upper()
                if not name: raise ValueError
                employee.nome = name
            if "cargo_id" in body: employee.cargo = int(body["cargo_id"]) if body["cargo_id"] not in (None, "") else None
            if "situacao_id" in body: employee.situacao = int(body["situacao_id"]) if body["situacao_id"] not in (None, "") else None
            if "centro_id" in body:
                center = db.session.get(CostCenters, int(body["centro_id"])) if body["centro_id"] not in (None, "") else None
                if body["centro_id"] not in (None, "") and not center: return jsonify("Centro de custo não encontrado."), 404
                if center and employee.empresa_id and center.empresa_id != employee.empresa_id: return jsonify("O centro deve pertencer à mesma empresa do colaborador."), 409
                employee.centro_id = center.id if center else None
            db.session.commit()
        except (TypeError, ValueError): db.session.rollback(); return jsonify("Confira os dados do colaborador."), 400
        socketio.emit("data_changed", {"channel": "colaboradores", "resource": "colaboradores", "action": "updated", "id": employee.id})
        return jsonify({"message": "Colaborador atualizado."}), 200

    def _read_tm_ops_lookup(self, token_data):
        try: page = max(int(rq.args.get("page", 1)), 1); per_page = min(max(int(rq.args.get("per_page", 20)), 1), 50)
        except (TypeError, ValueError): return jsonify("Paginação inválida."), 400
        query = db.session.query(Employees.id, Employees.matricula, Employees.nome).filter(Employees.situacao == 1)
        query = apply_cost_center_scope(query, Employees.centro_id, token_data)
        requested = _ids(rq.args.get("ids"))
        if requested: query = query.filter(Employees.id.in_(requested))
        else:
            search = " ".join(str(rq.args.get("search") or "").split())
            if search:
                filters = [Employees.nome.ilike(f"{search}%")]
                if search.isdigit(): filters.append(Employees.matricula == int(search))
                query = query.filter(or_(*filters))
        pagination = query.order_by(Employees.nome).paginate(page=page, per_page=per_page, error_out=False)
        return jsonify({"items": [row._asdict() for row in pagination.items], "page": pagination.page, "per_page": pagination.per_page, "total": pagination.total, "pages": pagination.pages}), 200

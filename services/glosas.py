# Regras de negócio de glosas.
# Biblioteca padrão.
from collections import defaultdict
from datetime import date, datetime as dt
from decimal import Decimal, InvalidOperation
from io import BytesIO
from os import getenv
from pathlib import Path

# Dependências externas.
from flask import jsonify, request, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import aliased
from werkzeug.utils import secure_filename

# Módulos internos da aplicação.
from models.centros_de_custo import CostCenters
from models.colaboradores import Employees
from models.controle_faltas import AbsenceControl
from models.glosas import Disallowance
from models.rp_historico import History
from models.rp_requisicao import Requisicao
from models.usuarios import Users
from utils.db import db
from utils.filial_scope import apply_cost_center_scope, can_access_cost_center
from utils.permissions import has_permission
from utils.safe_route import safe_route
from utils.socket import socketio

VALID_COVERAGE = {"em_analise", "coberta", "parcial", "descoberta"}
DEFAULT_DAILY_VALUE = Decimal("180.00")
MAX_EVIDENCE_SIZE = 15 * 1024 * 1024
ALLOWED_EVIDENCE_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".webp"}
EVIDENCE_DIR = Path(
    getenv("GLOSA_EVIDENCE_DIR")
    or Path(__file__).resolve().parents[1] / "storage" / "glosas"
)

def _parse_date(value, field):
    try:
        return date.fromisoformat(str(value)[:10])
    except (TypeError, ValueError):
        raise ValueError(f"{field} inválida.")

def _parse_decimal(value, field, default=None, places="0.01", allow_zero=False):
    if value in (None, "") and default is not None:
        return default
    try:
        parsed = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field} inválido.")
    if parsed < 0 or (parsed == 0 and not allow_zero):
        comparison = "maior ou igual a zero" if allow_zero else "maior que zero"
        raise ValueError(f"{field} deve ser {comparison}.")
    return parsed.quantize(Decimal(places))

def _money(value):
    return f"R$ {float(value or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

class DisallowanceService:
    @safe_route
    def dashboard(self, token_data):
        if not has_permission(token_data, "dashboard_glosas", "view"):
            return jsonify("Você não possui acesso ao Dashboard de Glosas."), 403
        try:
            records, summary, filters = self._records_and_summary(token_data)
        except ValueError as error:
            return jsonify(str(error)), 400
        by_contract = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
        by_reason = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
        by_employee = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
        by_month = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
        by_status = defaultdict(int)
        for item in records:
            value = float(item.get("valor_total") or 0)
            by_contract[item.get("contrato") or "Não informado"]["quantidade"] += 1
            by_contract[item.get("contrato") or "Não informado"]["valor"] += value
            reason = item.get("justificativa") or "Não informado"
            by_reason[reason]["quantidade"] += 1
            by_reason[reason]["valor"] += value
            employee = item.get("colaborador") or item.get("colaborador_nome") or "Não identificado"
            by_employee[employee]["quantidade"] += 1
            by_employee[employee]["valor"] += value
            month = str(item.get("competencia") or "")[:7] or "Sem competência"
            by_month[month]["quantidade"] += 1
            by_month[month]["valor"] += value
            by_status[item.get("cobertura") or "em_analise"] += 1
        as_rows = lambda data, key: [
            {key: name, **values, "valor": round(values["valor"], 2)}
            for name, values in sorted(data.items(), key=lambda item: item[1]["valor"], reverse=True)
        ]
        return jsonify({
            "resumo": summary, "filtros": filters, "status": dict(by_status),
            "por_contrato": as_rows(by_contract, "contrato"), "por_motivo": as_rows(by_reason, "motivo"),
            "por_colaborador": as_rows(by_employee, "colaborador"),
            "evolucao_mensal": as_rows(by_month, "competencia"),
        })
    @classmethod
    def ensure_for_absence(cls, absence, user_id=None):
        """Cria a glosa preventiva de uma falta tratada, sem duplicar o vínculo."""
        if not absence or absence.status != "tratada":
            return None, False

        if not absence.id:
            db.session.flush()

        existing = Disallowance.query.filter_by(falta_id=absence.id).first()
        if existing:
            return existing, False

        requisition = (
            db.session.get(Requisicao, absence.requisicao_id)
            if absence.requisicao_id
            else None
        )
        history = (
            History.query.filter_by(requisicao_id=absence.requisicao_id)
            .order_by(History.id.desc())
            .first()
            if absence.requisicao_id
            else None
        )
        request_status = (history.status if history else getattr(requisition, "status", None)) or ""
        reserve_id = (
            history.reserva_id if history else getattr(requisition, "reserva_id", 0)
        ) or 0
        if request_status == "approved" and reserve_id:
            coverage = "coberta"
        elif request_status == "reproved":
            coverage = "descoberta"
        else:
            coverage = "em_analise"

        total_days = Decimal("1")
        if absence.tipo_ausencia == "parcial" and absence.quantidade_horas:
            total_days = (
                Decimal(str(absence.quantidade_horas)) / Decimal("8")
            ).quantize(Decimal("0.0001"))
        daily_value = cls.get_default_daily_value(absence.centro_custo_id).quantize(
            Decimal("0.01")
        )
        covered_days = total_days if coverage == "coberta" else Decimal("0")
        total_value = (total_days * daily_value).quantize(Decimal("0.01"))
        covered_value = (covered_days * daily_value).quantize(Decimal("0.01"))

        item = Disallowance(
            competencia=absence.data_falta.date().replace(day=1),
            data_falta=absence.data_falta.date(),
            centro_custo_id=absence.centro_custo_id,
            colaborador_id=absence.colaborador_id,
            colaborador_nome=absence.colaborador_nome,
            colaborador_matricula=absence.colaborador_matricula,
            falta_id=absence.id,
            requisicao_id=absence.requisicao_id,
            cobertura=coverage,
            quantidade_dias=total_days,
            quantidade_coberta_dias=covered_days,
            valor_diaria=daily_value,
            valor_total=total_value,
            valor_coberto=covered_value,
            valor_descoberto=(total_value - covered_value).quantize(Decimal("0.01")),
            justificativa="Glosa preventiva gerada ao tratar a falta.",
            observacao="Registro criado automaticamente pelo Controle de Faltas.",
            criado_por_usuario_id=user_id,
        )
        db.session.add(item)
        db.session.flush()
        return item, True

    @staticmethod
    def _evidence_url(item):
        if not item.evidencia_arquivo:
            return None
        return f"https://api.tmhub.hubbix.com.br/arquivos/glosas/{item.evidencia_arquivo}"

    @classmethod
    def _serialize(cls, row):
        item = row.Disallowance
        dias = float(item.quantidade_dias or 0)
        dias_cobertos = float(item.quantidade_coberta_dias or 0)
        return {
            "id": item.id,
            "competencia": item.competencia.isoformat(),
            "data_falta": item.data_falta.isoformat(),
            "centro_custo_id": item.centro_custo_id,
            "contrato": row.contrato,
            "departamento": row.departamento,
            "colaborador_id": item.colaborador_id,
            "colaborador": row.colaborador or item.colaborador_nome,
            "matricula": row.matricula or item.colaborador_matricula,
            "falta_id": item.falta_id,
            "requisicao_id": item.requisicao_id,
            "cobertura": item.cobertura,
            "quantidade_dias": round(dias, 4),
            "quantidade_horas": round(dias * 8, 2),
            "quantidade_coberta_dias": round(dias_cobertos, 4),
            "quantidade_coberta_horas": round(dias_cobertos * 8, 2),
            "valor_diaria": float(item.valor_diaria or 0),
            "valor_total": float(item.valor_total or 0),
            "valor_coberto": float(item.valor_coberto or 0),
            "valor_descoberto": float(item.valor_descoberto or 0),
            "justificativa": item.justificativa,
            "observacao": item.observacao,
            "evidencia_nome": item.evidencia_nome_original,
            "evidencia_mime": item.evidencia_mime,
            "evidencia_url": cls._evidence_url(item),
            "criado_por": row.criado_por,
            "alterado_por": row.alterado_por,
            "created_at": item.created_at,
            "updated_at": item.updated_at,
        }

    @staticmethod
    def _query():
        Creator = aliased(Users)
        Editor = aliased(Users)
        return (
            db.session.query(
                Disallowance,
                CostCenters.local.label("contrato"),
                CostCenters.departamento.label("departamento"),
                Employees.nome.label("colaborador"),
                Employees.matricula.label("matricula"),
                Creator.nome.label("criado_por"),
                Editor.nome.label("alterado_por"),
            )
            .join(CostCenters, CostCenters.id == Disallowance.centro_custo_id)
            .outerjoin(Employees, Employees.id == Disallowance.colaborador_id)
            .outerjoin(Creator, Creator.id == Disallowance.criado_por_usuario_id)
            .outerjoin(Editor, Editor.id == Disallowance.alterado_por_usuario_id)
        )

    def _records_and_summary(self, token_data):
        query = apply_cost_center_scope(self._query(), Disallowance.centro_custo_id, token_data)
        if request.args.get("inicio"):
            query = query.filter(
                Disallowance.competencia >= _parse_date(request.args["inicio"], "Competência inicial")
            )
        if request.args.get("fim"):
            query = query.filter(
                Disallowance.competencia <= _parse_date(request.args["fim"], "Competência final")
            )

        scoped_records = [
            self._serialize(row)
            for row in query.order_by(
                Disallowance.competencia.desc(), Disallowance.data_falta.desc()
            ).all()
        ]
        filter_options = {
            "departamentos": sorted(
                {str(item["departamento"]) for item in scoped_records if item["departamento"] not in (None, "")},
                key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value),
            ),
            "contratos": sorted(
                [
                    {"label": item["contrato"], "value": item["centro_custo_id"]}
                    for item in {
                        record["centro_custo_id"]: record
                        for record in scoped_records
                        if record["centro_custo_id"]
                    }.values()
                ],
                key=lambda item: (item["label"] or "").casefold(),
            ),
            "colaboradores": sorted(
                [
                    {
                        "label": item["colaborador"],
                        "value": item["colaborador_id"],
                        "matricula": item["matricula"],
                    }
                    for item in {
                        record["colaborador_id"]: record
                        for record in scoped_records
                        if record["colaborador_id"] and record["colaborador"]
                    }.values()
                ],
                key=lambda item: item["label"].casefold(),
            ),
        }

        def selected_values(name):
            return {
                value.strip()
                for raw in request.args.getlist(name)
                for value in str(raw).split(",")
                if value.strip() and value.strip() != "__all__"
            }

        coverage = selected_values("cobertura")
        department = selected_values("departamento")
        contract = selected_values("contrato")
        employee = selected_values("colaborador")
        search = str(request.args.get("busca") or "").strip().casefold()

        records = []
        for item in scoped_records:
            if coverage and item["cobertura"] not in coverage:
                continue
            if department and str(item["departamento"]) not in department:
                continue
            if contract and str(item["centro_custo_id"]) not in contract:
                continue
            if employee and str(item["colaborador_id"]) not in employee:
                continue
            if search and not any(
                search in str(value or "").casefold()
                for value in (
                    item["contrato"],
                    item["colaborador"],
                    item["matricula"],
                    item["justificativa"],
                    item["observacao"],
                )
            ):
                continue
            records.append(item)

        # As listas dos filtros acompanham o recorte atual, evitando escolhas
        # que levariam a uma tabela e indicadores zerados.
        filter_options = {
            "departamentos": sorted(
                {str(item["departamento"]) for item in records if item["departamento"] not in (None, "")},
                key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value),
            ),
            "contratos": sorted(
                [{"label": item["contrato"], "value": item["centro_custo_id"]} for item in {
                    record["centro_custo_id"]: record for record in records if record["centro_custo_id"]
                }.values()],
                key=lambda item: (item["label"] or "").casefold(),
            ),
            "colaboradores": sorted(
                [{"label": item["colaborador"], "value": item["colaborador_id"], "matricula": item["matricula"]} for item in {
                    record["colaborador_id"]: record for record in records if record["colaborador_id"] and record["colaborador"]
                }.values()],
                key=lambda item: item["label"].casefold(),
            ),
        }

        summary = {
            "total_registros": len(records),
            "dias": round(sum(item["quantidade_dias"] for item in records), 2),
            "valor_total": round(sum(item["valor_total"] for item in records), 2),
            "valor_coberto": round(sum(item["valor_coberto"] for item in records), 2),
            "valor_descoberto": round(
                sum(
                    item["valor_descoberto"]
                    for item in records
                    if item["cobertura"] in {"parcial", "descoberta"}
                ),
                2,
            ),
            "valor_em_analise": round(
                sum(item["valor_total"] for item in records if item["cobertura"] == "em_analise"),
                2,
            ),
        }
        return records, summary, filter_options

    @safe_route
    def read(self, token_data):
        if not has_permission(token_data, "controle_glosas", "view"):
            return jsonify("Você não possui acesso ao Controle de Glosas."), 403
        try:
            records, summary, filter_options = self._records_and_summary(token_data)
        except ValueError as error:
            return jsonify(str(error)), 400
        return jsonify(
            {
                "registros": records,
                "resumo": summary,
                "filtros": filter_options,
                "valor_diaria_padrao": float(DEFAULT_DAILY_VALUE),
            }
        ), 200

    @staticmethod
    def get_default_daily_value(cost_center_id):
        if not cost_center_id:
            return DEFAULT_DAILY_VALUE
        center = db.session.get(CostCenters, cost_center_id)
        if not center:
            return DEFAULT_DAILY_VALUE
        if center.valor_diaria_glosa:
            return Decimal(str(center.valor_diaria_glosa))
        if str(center.departamento) == "269" or str(center.id) == "269":
            return Decimal("182.02")
        return DEFAULT_DAILY_VALUE

    def _apply(self, item, body, token_data, creating=False):
        try:
            if creating or "competencia" in body:
                competence = _parse_date(body.get("competencia"), "Competência")
                item.competencia = competence.replace(day=1)
            if creating or "data_falta" in body:
                item.data_falta = _parse_date(body.get("data_falta"), "Data da falta")
            if "centro_custo_id" in body:
                center_id = int(body.get("centro_custo_id"))
                if not db.session.get(CostCenters, center_id):
                    return "Contrato não encontrado."
                if not can_access_cost_center(token_data, center_id):
                    return "Você não possui acesso à filial deste contrato."
                item.centro_custo_id = center_id
            if creating or "quantidade_dias" in body or "quantidade_horas" in body:
                if "quantidade_dias" in body:
                    item.quantidade_dias = _parse_decimal(
                        body.get("quantidade_dias"),
                        "Quantidade de dias",
                        Decimal("1"),
                        places="0.0001",
                    )
                else:
                    hours = _parse_decimal(
                        body.get("quantidade_horas"),
                        "Quantidade de horas",
                        Decimal("8"),
                        places="0.01",
                    )
                    item.quantidade_dias = (hours / Decimal("8")).quantize(Decimal("0.0001"))
        except (TypeError, ValueError) as error:
            return str(error)

        if "cobertura" in body or creating:
            coverage = str(body.get("cobertura") or "em_analise").strip().lower()
            if coverage not in VALID_COVERAGE:
                return "Informe se a glosa está em análise, coberta, parcialmente coberta ou descoberta."
            item.cobertura = coverage

        if "falta_id" in body:
            absence_id = body.get("falta_id") or None
            absence = db.session.get(AbsenceControl, absence_id) if absence_id else None
            if absence_id and not absence:
                return "Registro de falta não encontrado."
            if absence and not can_access_cost_center(token_data, absence.centro_custo_id):
                return "Você não possui acesso à falta informada."
            item.falta_id = absence.id if absence else None
            item.requisicao_id = absence.requisicao_id if absence else None
            if absence:
                item.colaborador_id = absence.colaborador_id
                item.colaborador_nome = absence.colaborador_nome
                item.colaborador_matricula = absence.colaborador_matricula
                item.data_falta = absence.data_falta.date()
                item.centro_custo_id = absence.centro_custo_id

        if "colaborador_id" in body and not body.get("falta_id"):
            employee_id = body.get("colaborador_id") or None
            employee = db.session.get(Employees, employee_id) if employee_id else None
            if employee_id and not employee:
                return "Colaborador não encontrado."
            item.colaborador_id = employee.id if employee else None
            item.colaborador_nome = (
                employee.nome
                if employee
                else str(body.get("colaborador_nome") or "").strip() or None
            )
            item.colaborador_matricula = (
                employee.matricula
                if employee
                else str(body.get("colaborador_matricula") or "").strip() or None
            )
            if employee and employee.centro_id:
                item.centro_custo_id = employee.centro_id
        elif "colaborador_nome" in body and not item.colaborador_id:
            item.colaborador_nome = str(body.get("colaborador_nome") or "").strip() or None
            item.colaborador_matricula = (
                str(body.get("colaborador_matricula") or "").strip() or None
            )

        if creating and not item.centro_custo_id:
            return "O colaborador selecionado não possui contrato/centro de custo vinculado."

        default_rate = self.get_default_daily_value(item.centro_custo_id)
        if "valor_diaria" in body:
            try:
                item.valor_diaria = _parse_decimal(
                    body.get("valor_diaria"), "Valor da diária", default_rate
                )
            except ValueError as error:
                return str(error)
        elif creating or not item.valor_diaria:
            item.valor_diaria = default_rate

        total_days = Decimal(item.quantidade_dias or 0)
        if item.cobertura == "coberta":
            covered_days = total_days
        elif item.cobertura == "parcial":
            try:
                if "quantidade_coberta_dias" in body:
                    covered_days = _parse_decimal(
                        body.get("quantidade_coberta_dias"),
                        "Quantidade coberta",
                        places="0.0001",
                    )
                elif "quantidade_coberta_horas" in body:
                    covered_hours = _parse_decimal(
                        body.get("quantidade_coberta_horas"),
                        "Horas cobertas",
                        places="0.01",
                    )
                    covered_days = (covered_hours / Decimal("8")).quantize(Decimal("0.0001"))
                else:
                    covered_days = Decimal(item.quantidade_coberta_dias or 0)
            except ValueError as error:
                return str(error)
            if covered_days <= 0 or covered_days >= total_days:
                return "Na cobertura parcial, informe um tempo coberto maior que zero e menor que o total."
        else:
            covered_days = Decimal("0")

        item.quantidade_coberta_dias = covered_days
        item.valor_total = (total_days * Decimal(item.valor_diaria)).quantize(Decimal("0.01"))
        item.valor_coberto = (covered_days * Decimal(item.valor_diaria)).quantize(Decimal("0.01"))
        item.valor_descoberto = (item.valor_total - item.valor_coberto).quantize(Decimal("0.01"))

        if "justificativa" in body or creating:
            item.justificativa = str(body.get("justificativa") or "").strip() or None
        if "observacao" in body or creating:
            item.observacao = str(body.get("observacao") or "").strip() or None
        return None

    @safe_route
    def create(self, token_data):
        if not has_permission(token_data, "controle_glosas", "create"):
            return jsonify("Você não possui permissão para criar glosas."), 403
        body = request.get_json(silent=True) or {}
        item = Disallowance(criado_por_usuario_id=token_data.get("id"))
        error = self._apply(item, body, token_data, creating=True)
        if error:
            return jsonify(error), 400
        db.session.add(item)
        db.session.commit()
        socketio.emit("disallowance_update", {"id": item.id, "action": "created"})
        return jsonify({"message": "Glosa registrada.", "id": item.id}), 201

    @safe_route
    def update(self, token_data):
        if not has_permission(token_data, "controle_glosas", "edit"):
            return jsonify("Você não possui permissão para alterar glosas."), 403
        body = request.get_json(silent=True) or {}
        item = db.session.get(Disallowance, body.get("id"))
        if not item:
            return jsonify("Glosa não encontrada."), 404
        if not can_access_cost_center(token_data, item.centro_custo_id):
            return jsonify("Você não possui acesso à filial desta glosa."), 403
        error = self._apply(item, body, token_data)
        if error:
            return jsonify(error), 400
        item.alterado_por_usuario_id = token_data.get("id")
        item.updated_at = dt.now()
        db.session.commit()
        socketio.emit("disallowance_update", {"id": item.id, "action": "updated"})
        return jsonify({"message": "Glosa atualizada.", "id": item.id}), 200

    @safe_route
    def upload_evidence(self, glosa_id, token_data):
        if not has_permission(token_data, "controle_glosas", "edit"):
            return jsonify("Você não possui permissão para alterar glosas."), 403
        item = db.session.get(Disallowance, glosa_id)
        if not item:
            return jsonify("Glosa não encontrada."), 404
        if not can_access_cost_center(token_data, item.centro_custo_id):
            return jsonify("Você não possui acesso à filial desta glosa."), 403

        uploaded = request.files.get("evidencia")
        if not uploaded or not uploaded.filename:
            return jsonify("Selecione uma imagem ou PDF como evidência."), 400
        original_name = secure_filename(uploaded.filename)
        extension = Path(original_name).suffix.lower()
        if extension not in ALLOWED_EVIDENCE_EXTENSIONS:
            return jsonify("Formato inválido. Envie PDF, PNG, JPG, JPEG ou WEBP."), 400

        uploaded.stream.seek(0, 2)
        size = uploaded.stream.tell()
        uploaded.stream.seek(0)
        if size > MAX_EVIDENCE_SIZE:
            return jsonify("A evidência deve ter no máximo 15 MB."), 400

        stored_name = f"{uuid4().hex}{extension}"
        stored_path = EVIDENCE_DIR / stored_name
        try:
            EVIDENCE_DIR.mkdir(parents=True, exist_ok=True)
            uploaded.save(stored_path)
        except OSError as error:
            return jsonify(
                f"Não foi possível salvar a evidência no servidor: {error.strerror or error}"
            ), 500

        old_name = item.evidencia_arquivo
        item.evidencia_arquivo = stored_name
        item.evidencia_nome_original = original_name
        item.evidencia_mime = uploaded.mimetype
        item.alterado_por_usuario_id = token_data.get("id")
        item.updated_at = dt.now()
        try:
            db.session.commit()
        except Exception:
            if stored_path.is_file():
                stored_path.unlink()
            raise

        if old_name:
            old_path = EVIDENCE_DIR / Path(old_name).name
            if old_path.is_file():
                old_path.unlink()

        socketio.emit("disallowance_update", {"id": item.id, "action": "evidence_updated"})
        return jsonify(
            {
                "message": "Evidência salva.",
                "nome": item.evidencia_nome_original,
                "url": self._evidence_url(item),
            }
        ), 200

    @safe_route
    def remove_evidence(self, glosa_id, token_data):
        if not has_permission(token_data, "controle_glosas", "edit"):
            return jsonify("Você não possui permissão para alterar glosas."), 403
        item = db.session.get(Disallowance, glosa_id)
        if not item:
            return jsonify("Glosa não encontrada."), 404
        if not can_access_cost_center(token_data, item.centro_custo_id):
            return jsonify("Você não possui acesso à filial desta glosa."), 403
        stored_name = item.evidencia_arquivo
        item.evidencia_arquivo = None
        item.evidencia_nome_original = None
        item.evidencia_mime = None
        item.alterado_por_usuario_id = token_data.get("id")
        item.updated_at = dt.now()
        db.session.commit()
        if stored_name:
            path = EVIDENCE_DIR / Path(stored_name).name
            if path.is_file():
                path.unlink()
        socketio.emit("disallowance_update", {"id": item.id, "action": "evidence_removed"})
        return jsonify("Evidência removida."), 200

    @staticmethod
    def serve_evidence(filename):
        safe_name = Path(filename).name
        item = Disallowance.query.filter_by(evidencia_arquivo=safe_name).first()
        if not item or safe_name != filename:
            return jsonify("Evidência não encontrada."), 404
        return send_from_directory(EVIDENCE_DIR, safe_name, as_attachment=False)

    @safe_route
    def export(self, token_data):
        if not has_permission(token_data, "controle_glosas", "view"): return jsonify("Você não possui acesso ao Controle de Glosas."), 403
        try: records, summary, _ = self._records_and_summary(token_data)
        except ValueError as error: return jsonify(str(error)), 400

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Controle de Glosas"
        sheet.sheet_view.showGridLines = False
        green = "20A65A"
        dark = "173925"
        light = "EAF6EF"
        red = "D64545"
        amber = "D99000"
        white = "FFFFFF"
        thin = Side(style="thin", color="DDE7E1")

        sheet.merge_cells("A1:M2")
        title = sheet["A1"]
        title.value = "CONTROLE DE GLOSAS"
        title.font = Font(size=20, bold=True, color=white)
        title.fill = PatternFill("solid", fgColor=dark)
        title.alignment = Alignment(vertical="center", horizontal="left")

        cards = [
            ("REGISTROS", summary["total_registros"], dark),
            ("VALOR APONTADO", _money(summary["valor_total"]), green),
            ("VALOR COBERTO", _money(summary["valor_coberto"]), "2E8B57"),
            ("SALDO DESCOBERTO", _money(summary["valor_descoberto"]), red),
            ("EM ANÁLISE", _money(summary["valor_em_analise"]), amber),
        ]
        columns = (1, 4, 7, 10, 12)
        widths = (3, 3, 3, 2, 2)
        for (label, value, color), start, width in zip(cards, columns, widths):
            end = start + width - 1
            sheet.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
            sheet.merge_cells(start_row=5, start_column=start, end_row=6, end_column=end)
            label_cell = sheet.cell(4, start, label)
            value_cell = sheet.cell(5, start, value)
            for row in range(4, 7):
                for column in range(start, end + 1):
                    cell = sheet.cell(row, column)
                    cell.fill = PatternFill("solid", fgColor=light)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            label_cell.font = Font(size=9, bold=True, color=color)
            value_cell.font = Font(size=14, bold=True, color=dark)
            label_cell.alignment = value_cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

        headers = [
            "Competência",
            "Data da falta",
            "Departamento",
            "Contrato",
            "Colaborador",
            "Matrícula",
            "Situação",
            "Dias apontados",
            "Dias cobertos",
            "Valor apontado",
            "Valor coberto",
            "Saldo descoberto",
            "Evidência",
        ]
        header_row = 8
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(header_row, column, label)
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        coverage_labels = {
            "em_analise": "Em análise",
            "coberta": "Coberta",
            "parcial": "Parcialmente coberta",
            "descoberta": "Descoberta",
        }
        for row_index, item in enumerate(records, header_row + 1):
            values = [
                item["competencia"],
                item["data_falta"],
                item["departamento"],
                item["contrato"],
                item["colaborador"],
                item["matricula"],
                coverage_labels.get(item["cobertura"], item["cobertura"]),
                item["quantidade_dias"],
                item["quantidade_coberta_dias"],
                item["valor_total"],
                item["valor_coberto"],
                item["valor_descoberto"],
                "Abrir evidência" if item["evidencia_url"] else "",
            ]
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_index, column, value)
                cell.fill = PatternFill(
                    "solid", fgColor="FFFFFF" if row_index % 2 else "F5F9F7"
                )
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center")
            for column in (10, 11, 12):
                sheet.cell(row_index, column).number_format = 'R$ #,##0.00'
            if item["evidencia_url"]:
                evidence = sheet.cell(row_index, 13)
                evidence.hyperlink = item["evidencia_url"]
                evidence.style = "Hyperlink"

        sheet.freeze_panes = "A9"
        sheet.auto_filter.ref = f"A8:M{max(header_row, header_row + len(records))}"
        widths = [15, 15, 15, 42, 34, 14, 22, 16, 15, 18, 18, 19, 22]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[5].height = 24
        sheet.row_dimensions[8].height = 30

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"controle_glosas_{date.today().isoformat()}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @safe_route
    def delete(self, token_data):
        if not has_permission(token_data, "controle_glosas", "edit"):
            return jsonify("Você não possui permissão para excluir glosas."), 403
        body = request.get_json(silent=True) or request.args
        item = db.session.get(Disallowance, body.get("id"))
        if not item:
            return jsonify("Glosa não encontrada."), 404
        if not can_access_cost_center(token_data, item.centro_custo_id):
            return jsonify("Você não possui acesso à filial desta glosa."), 403
        item_id = item.id
        stored_name = item.evidencia_arquivo
        db.session.delete(item)
        db.session.commit()
        if stored_name:
            path = EVIDENCE_DIR / Path(stored_name).name
            if path.is_file():
                path.unlink()
        socketio.emit("disallowance_update", {"id": item_id, "action": "deleted"})
        return jsonify("Glosa excluída."), 200

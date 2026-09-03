from __future__ import annotations

from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
import re
import secrets
import unicodedata

from flask import jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text

from models.colaboradores import Employees
from utils.db import db
from utils.filial_scope import allowed_cost_center_ids, is_admin
from utils.permissions import has_permission
from routes.rpa import discard_command, send_command_for_capability, track_command
from utils.safe_route import safe_route
from utils.socket import socketio
from utils.token import create_token


MAX_IMPORT_SIZE = 30 * 1024 * 1024
INDICATORS = {"intrajornada": "Intrajornada", "interjornada": "Interjornada", "escala": "Escala 6x1 / 5x2"}
VALUE_SUFFIX = re.compile(r"\s*\[valor:\s*(?P<valor>[^\]]+)\]\s*$", re.I)


def _normal(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join("".join(char for char in value if not unicodedata.combining(char)).upper().split())


def _date(value):
    if isinstance(value, datetime): return value.date()
    if isinstance(value, date): return value
    text = str(value or "").strip()
    date_match = re.search(r"\b\d{2}/\d{2}/\d{4}\b", text)
    for candidate in (date_match.group(0) if date_match else text[:10],):
        for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try: return datetime.strptime(candidate, pattern).date()
            except ValueError: continue
    return None


def _registration(value):
    value = str(value or "").strip()
    if not value:
        return None
    return value[:-2] if value.endswith(".0") else value


def _indicator(value):
    value = _normal(value)
    if "INTRAJORNADA" in value or "INTRA JORNADA" in value: return "intrajornada"
    if "INTERJORNADA" in value or "INTER JORNADA" in value: return "interjornada"
    if "6X1" in value or "5X2" in value or "CONSECUT" in value or "ESCALA" in value: return "escala"
    return None


def _offense(value):
    value = _normal(value)
    return bool(value and value not in {"-", "N/A", "NA", "NAO", "NAO OCORREU", "CONFORME", "OK", "0", "0:00", "00:00"})


def _zero_value(value):
    normalized = _normal(value).replace(" ", "")
    return normalized in {"", "0", "0.0", "0,0", "0:00", "00:00", "00:00:00"}


def _report_value(value):
    if value is None: return None
    if isinstance(value, time): return value.strftime("%H:%M")
    if isinstance(value, timedelta):
        minutes = round(value.total_seconds() / 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    text = str(value).strip()
    return text or None


def _time_minutes(value):
    match = re.fullmatch(r"\s*(\d{1,2}):(\d{2})(?::\d{2})?\s*", str(value or ""))
    if not match: return None
    return int(match.group(1)) * 60 + int(match.group(2))


def _duration_label(minutes):
    hours, remainder = divmod(abs(minutes), 60)
    if hours and remainder: return f"{hours}h{remainder:02d}"
    if hours: return f"{hours}h"
    return f"{remainder} min"


def _split_result_value(result):
    text = str(result or "").strip()
    match = VALUE_SUFFIX.search(text)
    return (text[:match.start()].strip(), match.group("valor").strip()) if match else (text, None)


def _value_description(indicator, value):
    value = _report_value(value)
    if not value: return None
    minutes = _time_minutes(value)
    if indicator == "intrajornada" and minutes is not None:
        shortage = max(0, 60 - minutes)
        return f"Intervalo realizado: {value}" + (f" · {_duration_label(shortage)} abaixo de 1h" if shortage else "")
    if indicator == "interjornada" and minutes is not None:
        shortage = max(0, 11 * 60 - minutes)
        return f"Descanso realizado: {value}" + (f" · {_duration_label(shortage)} abaixo de 11h" if shortage else "")
    if indicator == "escala":
        days = re.search(r"\d+", value)
        if days: return f"{days.group(0)} dia(s) consecutivo(s) trabalhado(s)"
    return f"Valor reportado: {value}"


class JourneyControlService:
    @staticmethod
    def _read(upload):
        raw = upload.read()
        if not raw or len(raw) > MAX_IMPORT_SIZE: raise ValueError("A planilha deve ter entre 1 byte e 30 MB.")
        try:
            sheet = load_workbook(BytesIO(raw), read_only=True, data_only=True).active
            # O exportador do Ponto Mais declara, incorretamente, a área A:B no
            # metadata do XLSX, embora grave Data, Ocorrência e Valor em C:E.
            # Resetar a dimensão faz o openpyxl percorrer as células efetivas,
            # sem abrir a planilha inteira em memória.
            sheet.reset_dimensions()
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        except Exception as error:
            raise ValueError("O XLSX não pôde ser lido. Exporte novamente o relatório no PontoMais.") from error

    @staticmethod
    def _parse(rows):
        header_index = headers = None
        for index, row in enumerate(rows[:30]):
            labels = [_normal(value) for value in row]
            name = next((i for i, value in enumerate(labels) if value in {"NOME", "COLABORADOR", "FUNCIONARIO", "NOME DO COLABORADOR"}), None)
            day = next((i for i, value in enumerate(labels) if value in {"DATA", "DIA"}), None)
            direct = {kind: i for i, value in enumerate(labels) if (kind := _indicator(value))}
            generic = next((i for i, value in enumerate(labels) if value in {"RESULTADO", "INFRACAO", "INFRACAO APURADA", "OCORRENCIA", "MOTIVO", "DESCRICAO"}), None)
            if name is not None and day is not None and (direct or generic is not None):
                header_index, headers = index, {"nome": name, "data": day, "matricula": next((i for i, value in enumerate(labels) if value in {"MATRICULA", "REGISTRO", "CODIGO"}), None), "generico": generic, "valor": next((i for i, value in enumerate(labels) if value in {"VALOR", "VALOR APURADO"}), None), **direct}
                break
        if headers is None: raise ValueError("Não localizamos Colaborador, Data e Resultado/Infração no modelo Jornadas.")
        parsed, direct = [], {key: value for key, value in headers.items() if key in INDICATORS}
        for row in rows[header_index + 1:]:
            name = str(row[headers["nome"]] if headers["nome"] < len(row) else "").strip(); day = _date(row[headers["data"]] if headers["data"] < len(row) else None)
            if not name or not day: continue
            registration = _registration(row[headers["matricula"]]) if headers["matricula"] is not None and headers["matricula"] < len(row) else None
            found = [(kind, str(row[column]).strip()) for kind, column in direct.items() if column < len(row) and _offense(row[column])]
            generic = row[headers["generico"]] if headers["generico"] is not None and headers["generico"] < len(row) else None
            value = row[headers["valor"]] if headers["valor"] is not None and headers["valor"] < len(row) else None
            if not found and _offense(generic):
                found = [
                    (kind, part.strip())
                    for part in re.split(r"[;|\n]", str(generic))
                    if (kind := _indicator(part)) and not (kind == "intrajornada" and _zero_value(value))
                ]
            parsed.extend({"nome": name, "matricula": registration, "data": day, "indicador": kind, "resultado": result or INDICATORS[kind], "valor": _report_value(value)} for kind, result in found)
        return parsed

    @staticmethod
    def _employees_by_registration():
        registrations = {}
        for employee in Employees.query.all():
            registration = _registration(employee.matricula)
            if registration:
                registrations.setdefault(registration, []).append(employee)
        return registrations

    @staticmethod
    def _rows(token_data):
        rows = [dict(row) for row in db.session.execute(text("""
            SELECT i.id, i.colaborador_id, i.nome_colaborador, i.matricula, i.data_ocorrencia, i.indicador, i.resultado_relatorio,
                   imp.data_referencia, imp.arquivo_origem,
                   e.centro_id, c.local AS contrato, c.departamento,
                   CASE WHEN e.id IS NULL THEN 'pendente' ELSE 'vinculado' END AS vinculo_status
            FROM jornadas_infracoes i JOIN jornadas_importacoes imp ON imp.id = i.importacao_id
            LEFT JOIN colaboradores e ON e.id = i.colaborador_id
            LEFT JOIN centro_de_custo c ON c.id = e.centro_id ORDER BY i.data_ocorrencia DESC, i.nome_colaborador
        """)).mappings()]
        if is_admin(token_data): return rows
        allowed = allowed_cost_center_ids(token_data) or set()
        return [row for row in rows if row.get("centro_id") in allowed]

    @staticmethod
    def _filter(rows):
        values = lambda key: {value for value in str(request.args.get(key) or "").split(",") if value}
        types, links = values("tipo"), values("vinculo")
        contracts = {_normal(value) for value in values("contrato")}
        departments = {_normal(value) for value in values("departamento")}
        start, end, search = _date(request.args.get("inicio")), _date(request.args.get("fim")), _normal(request.args.get("search"))
        return [row for row in rows if (
            (not types or row["indicador"] in types)
            and (not links or row["vinculo_status"] in links)
            and (not contracts or _normal(row.get("contrato")) in contracts)
            and (not departments or _normal(row.get("departamento")) in departments)
            and (not start or row["data_ocorrencia"] >= start)
            and (not end or row["data_ocorrencia"] <= end)
            and (not search or search in _normal(" ".join(str(row.get(key) or "") for key in ("nome_colaborador", "matricula", "contrato", "departamento"))))
        )]

    @staticmethod
    def _sql_filters(token_data, include_filters=True):
        """Monta filtros parametrizados para o banco, sem carregar o histórico em memória."""
        conditions, params = ["1 = 1"], {}

        def values(key):
            return [value.strip() for value in str(request.args.get(key) or "").split(",") if value.strip()]

        def in_clause(column, prefix, items):
            if not items: return
            placeholders = []
            for index, item in enumerate(items):
                name = f"{prefix}_{index}"
                params[name] = item
                placeholders.append(f":{name}")
            conditions.append(f"{column} IN ({', '.join(placeholders)})")

        if not is_admin(token_data):
            allowed = list(allowed_cost_center_ids(token_data) or set())
            if not allowed:
                conditions.append("1 = 0")
            else:
                in_clause("e.centro_id", "centro", allowed)

        if include_filters:
            in_clause("i.indicador", "tipo", values("tipo"))
            links = set(values("vinculo"))
            if links == {"vinculado"}: conditions.append("e.id IS NOT NULL")
            elif links == {"pendente"}: conditions.append("e.id IS NULL")

            contracts = values("contrato")
            if contracts:
                in_clause("LOWER(COALESCE(c.local, ''))", "contrato", [item.lower() for item in contracts])
            departments = values("departamento")
            if departments:
                in_clause("CAST(c.departamento AS VARCHAR)", "departamento", departments)

            start, end = _date(request.args.get("inicio")), _date(request.args.get("fim"))
            if start:
                conditions.append("i.data_ocorrencia >= :inicio")
                params["inicio"] = start
            if end:
                conditions.append("i.data_ocorrencia <= :fim")
                params["fim"] = end
            search = str(request.args.get("search") or "").strip()
            if search:
                conditions.append("LOWER(CONCAT_WS(' ', i.nome_colaborador, COALESCE(i.matricula, ''), COALESCE(c.local, ''), COALESCE(CAST(c.departamento AS VARCHAR), ''))) LIKE :search")
                params["search"] = f"%{search.lower()}%"

        return " AND ".join(conditions), params

    @staticmethod
    def _sql_from():
        return """
            FROM jornadas_infracoes i
            JOIN jornadas_importacoes imp ON imp.id = i.importacao_id
            LEFT JOIN colaboradores e ON e.id = i.colaborador_id
            LEFT JOIN centro_de_custo c ON c.id = e.centro_id
        """

    @safe_route
    def filter_options(self, token_data):
        if not has_permission(token_data, "controle_jornadas", "view"):
            return jsonify("Você não possui acesso ao Controle de Jornadas."), 403
        where, params = self._sql_filters(token_data, include_filters=False)
        contracts = db.session.execute(text(f"SELECT DISTINCT c.local {self._sql_from()} WHERE {where} AND c.local IS NOT NULL ORDER BY c.local"), params).scalars().all()
        departments = db.session.execute(text(f"SELECT DISTINCT c.departamento {self._sql_from()} WHERE {where} AND c.departamento IS NOT NULL ORDER BY c.departamento"), params).scalars().all()
        return jsonify({
            "contratos": [{"label": value, "value": value} for value in contracts],
            "departamentos": [{"label": str(value), "value": str(value)} for value in departments],
        }), 200

    @safe_route
    def read(self, token_data):
        if not has_permission(token_data, "controle_jornadas", "view"): return jsonify("Você não possui acesso ao Controle de Jornadas."), 403
        where, params = self._sql_filters(token_data)
        page = max(request.args.get("page", 1, type=int) or 1, 1)
        per_page = min(max(request.args.get("per_page", 25, type=int) or 25, 10), 100)
        sort_fields = {
            "data": "i.data_ocorrencia", "indicador": "i.indicador", "colaborador": "i.nome_colaborador",
            "contrato": "c.local", "vinculo_status": "CASE WHEN e.id IS NULL THEN 'pendente' ELSE 'vinculado' END",
        }
        sort_field = sort_fields.get(str(request.args.get("ordenar") or "data"), "i.data_ocorrencia")
        direction = "ASC" if str(request.args.get("direcao") or "").lower() == "asc" else "DESC"
        aggregate = db.session.execute(text(f"""
            SELECT COUNT(*) AS total,
                   COALESCE(SUM(CASE WHEN i.indicador = 'intrajornada' THEN 1 ELSE 0 END), 0) AS intrajornada,
                   COALESCE(SUM(CASE WHEN i.indicador = 'interjornada' THEN 1 ELSE 0 END), 0) AS interjornada,
                   COALESCE(SUM(CASE WHEN i.indicador = 'escala' THEN 1 ELSE 0 END), 0) AS escala,
                   COUNT(DISTINCT COALESCE(CAST(i.colaborador_id AS VARCHAR), NULLIF(i.matricula, ''), i.nome_colaborador)) AS colaboradores,
                   COALESCE(SUM(CASE WHEN e.id IS NULL THEN 1 ELSE 0 END), 0) AS vinculos_pendentes
            {self._sql_from()} WHERE {where}
        """), params).mappings().one()
        query_params = {**params, "limit": per_page, "offset": (page - 1) * per_page}
        rows = [dict(row) for row in db.session.execute(text(f"""
            SELECT i.id, i.colaborador_id, i.nome_colaborador, i.matricula, i.data_ocorrencia, i.indicador, i.resultado_relatorio,
                   c.local AS contrato, c.departamento,
                   CASE WHEN e.id IS NULL THEN 'pendente' ELSE 'vinculado' END AS vinculo_status
            {self._sql_from()} WHERE {where}
            ORDER BY {sort_field} {direction}, i.id DESC
            LIMIT :limit OFFSET :offset
        """), query_params).mappings()]
        latest = db.session.execute(text(f"""
            SELECT imp.data_referencia, imp.arquivo_origem
            {self._sql_from()} WHERE {where}
            ORDER BY imp.data_referencia DESC, imp.id DESC LIMIT 1
        """), params).mappings().first()
        records = []
        for row in rows:
            detail, value = _split_result_value(row["resultado_relatorio"])
            description = _value_description(row["indicador"], value) or detail
            records.append({"id": row["id"], "tipo": row["indicador"], "indicador": INDICATORS[row["indicador"]], "detalhe": detail, "valor": value, "descricao_valor": description, "data": row["data_ocorrencia"].isoformat(), "colaborador": row["nome_colaborador"], "matricula": row["matricula"], "colaborador_id": row["colaborador_id"], "vinculo_status": row["vinculo_status"], "contrato": row["contrato"], "departamento": str(row["departamento"]) if row["departamento"] is not None else None})
        return jsonify({"registros": records, "pagina": page, "por_pagina": per_page, "total": int(aggregate["total"] or 0), "resumo": {key: int(aggregate[key] or 0) for key in ("total", "intrajornada", "interjornada", "escala", "colaboradores", "vinculos_pendentes")}, "ultima_importacao": {"data_referencia": latest["data_referencia"].isoformat(), "arquivo": latest["arquivo_origem"]} if latest else None}), 200

    @safe_route
    def import_spreadsheet(self, token_data):
        is_rpa = token_data.get("typ") == "pontomais_rpa"
        if is_rpa:
            if token_data.get("scope") != "jornadas_import":
                return jsonify("A credencial do agente não é válida para esta importação."), 403
        elif not has_permission(token_data, "controle_jornadas", "create"):
            return jsonify("Você não possui permissão para importar jornadas."), 403
        upload = request.files.get("file")
        if not upload or not str(upload.filename or "").lower().endswith(".xlsx"): return jsonify("Envie o relatório Auditoria/Jornadas em XLSX."), 400
        reference_date = _date(request.form.get("data_referencia")) or date.today()
        if is_rpa and reference_date.isoformat() != token_data.get("reference_date"):
            return jsonify("A credencial do agente só pode importar a data solicitada."), 403
        try: parsed = self._parse(self._read(upload))
        except Exception as error: return jsonify(f"Não foi possível ler o relatório Jornadas: {error}"), 400
        if not parsed: return jsonify("O relatório não contém infrações de intrajornada, interjornada ou escala."), 400
        reports_by_date = {reference_date: parsed}

        registrations, total_imported, linked = self._employees_by_registration(), 0, 0
        for import_date, report_items in reports_by_date.items():
            # Cada dia possui seu próprio lote. Reimportar um período substitui
            # somente aqueles dias, preservando os demais dias já consolidados.
            db.session.execute(text("DELETE FROM jornadas_importacoes WHERE data_referencia = :data"), {"data": import_date})
            import_id = db.session.execute(text("INSERT INTO jornadas_importacoes (data_referencia, arquivo_origem, importado_por_usuario_id) VALUES (:data, :arquivo, :usuario) RETURNING id"), {"data": import_date, "arquivo": Path(upload.filename).name[:255], "usuario": token_data.get("id")}).scalar_one()
            payload, keys = [], set()
            for item in report_items:
                key = (item["matricula"] or _normal(item["nome"]), item["data"], item["indicador"])
                if key in keys: continue
                keys.add(key); matches = registrations.get(item["matricula"], []); employee = matches[0] if len(matches) == 1 else None; linked += int(employee is not None)
                result = item["resultado"]
                if item.get("valor"):
                    result = f"{result} [valor: {item['valor']}]"
                payload.append({"importacao_id": import_id, "colaborador_id": employee.id if employee else None, "nome_colaborador": item["nome"][:255], "matricula": item["matricula"], "data_ocorrencia": item["data"], "indicador": item["indicador"], "resultado_relatorio": result[:1000]})
            if payload:
                db.session.execute(text("INSERT INTO jornadas_infracoes (importacao_id, colaborador_id, nome_colaborador, matricula, data_ocorrencia, indicador, resultado_relatorio) VALUES (:importacao_id, :colaborador_id, :nome_colaborador, :matricula, :data_ocorrencia, :indicador, :resultado_relatorio)"), payload)
                total_imported += len(payload)
        db.session.commit(); socketio.emit("journey_update", {"action": "imported"})
        return jsonify({"message": "Infrações do relatório Jornadas importadas com sucesso.", "importados": total_imported, "vinculados": linked, "pendentes_vinculo": total_imported - linked}), 201

    @safe_route
    def automate_import(self, token_data):
        if not has_permission(token_data, "controle_jornadas", "create"):
            return jsonify("Você não possui permissão para automatizar a importação de jornadas."), 403
        reference_date = date.today() - timedelta(days=1)
        import_token = create_token({
            "id": token_data.get("id"),
            "ver": token_data.get("ver"),
            "typ": "pontomais_rpa",
            "scope": "jornadas_import",
            "reference_date": reference_date.isoformat(),
        }, expires_in_minutes=30)
        command_id = secrets.token_urlsafe(18)
        track_command(command_id, token_data.get("id"), "pontomais_report_import")
        sent, error, agent_id = send_command_for_capability("pontomais_report_import", {
            "command_id": command_id,
            "type": "pontomais_report_import",
            "category": "Ponto Mais",
            "report": "jornadas",
            "reference_date": reference_date.isoformat(),
            "import_token": import_token,
        }, category="Ponto Mais")
        if not sent:
            discard_command(command_id)
            return jsonify("O agente Ponto Mais - Relatórios está offline. Inicie-o e tente novamente."), 409
        return jsonify({"message": "A automação Ponto Mais foi iniciada. O relatório Jornadas será importado ao terminar.", "command_id": command_id, "agent_id": agent_id, "data_referencia": reference_date.isoformat()}), 202

    @safe_route
    def update_record(self, record_id, token_data):
        if not has_permission(token_data, "controle_jornadas", "edit"): return jsonify("Você não possui permissão para atualizar jornadas."), 403
        registration = _registration((request.get_json(silent=True) or {}).get("matricula")); employee = Employees.query.filter_by(matricula=int(registration)).first() if registration.isdigit() else None
        if not employee: return jsonify("Matrícula não encontrada. Cadastre ou atualize o colaborador e tente novamente."), 404
        if not is_admin(token_data) and employee.centro_id not in (allowed_cost_center_ids(token_data) or set()): return jsonify("A matrícula não pertence ao seu escopo de filiais."), 403
        updated = db.session.execute(text("UPDATE jornadas_infracoes SET colaborador_id = :colaborador, matricula = :matricula WHERE id = :id"), {"id": record_id, "colaborador": employee.id, "matricula": str(employee.matricula)})
        if not updated.rowcount: return jsonify("Registro de jornada não encontrado."), 404
        db.session.commit(); socketio.emit("journey_update", {"action": "updated", "id": record_id}); return jsonify({"message": "Vínculo manual salvo."}), 200

    @safe_route
    def export_spreadsheet(self, token_data):
        if not has_permission(token_data, "controle_jornadas", "view"): return jsonify("Você não possui acesso à exportação."), 403
        rows = self._filter(self._rows(token_data))
        if not rows: return jsonify("Não há ofensores no recorte selecionado."), 404
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Ofensores de jornada"; headers = ["Indicador", "Data", "Colaborador", "Matrícula", "Contrato", "Departamento", "Resultado do relatório", "Vínculo"]
        sheet.merge_cells("A1:H1"); sheet["A1"] = "DEMONSTRATIVO DE OFENSORES DE JORNADA"; sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF"); sheet["A1"].fill = PatternFill("solid", fgColor="173925")
        for index, label in enumerate(headers, 1): sheet.cell(3, index, label).font = Font(bold=True, color="FFFFFF"); sheet.cell(3, index).fill = PatternFill("solid", fgColor="173925")
        for row_index, row in enumerate(rows, 4):
            for column, value in enumerate([INDICATORS[row["indicador"]], row["data_ocorrencia"], row["nome_colaborador"], row["matricula"], row["contrato"], row["departamento"], row["resultado_relatorio"], "Vinculado" if row["vinculo_status"] == "vinculado" else "Pendente"], 1): sheet.cell(row_index, column, value).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A4"; sheet.auto_filter.ref = f"A3:H{sheet.max_row}"
        for letter, width in zip("ABCDEFGH", (22, 14, 34, 16, 30, 16, 54, 16)): sheet.column_dimensions[letter].width = width
        output = BytesIO(); workbook.save(output); output.seek(0)
        return send_file(output, as_attachment=True, download_name="ofensores_de_jornada.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Regras de negócio de medidas disciplinares.
# Biblioteca padrão.
import hashlib
import re
import unicodedata
from datetime import date, datetime
from zipfile import BadZipFile

# Dependências externas.
from flask import jsonify, request
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import String, case, cast, func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

# Módulos internos da aplicação.
from models.colaboradores import Employees
from models.centros_de_custo import CostCenters
from models.medidas_disciplinares import DisciplinaryMeasure
from models.supervisores import Supervisors
from utils.db import db
from utils.filial_scope import allowed_cost_center_ids, apply_cost_center_scope, is_admin
from utils.permissions import has_permission
from utils.safe_route import safe_route


SCREEN = "controle_medidas_disciplinares"
MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000
MEASURE_TYPES = {
    "advertencia": "Advertência",
    "suspensao": "Suspensão",
}
LEGACY_REASONS = {
    "falta_injustificada": "Falta injustificada",
    "atraso": "Atraso",
    "insubordinacao": "Insubordinação",
    "conduta_inadequada": "Conduta inadequada",
    "descumprimento_norma": "Descumprimento de norma",
    "outro": "Outro",
}

ARTICLE_482_SHORT_LABELS = {
    "a": "Improbidade",
    "b": "Má conduta",
    "c": "Negociação sem permissão",
    "d": "Condenação criminal",
    "e": "Desídia",
    "f": "Embriaguez",
    "g": "Violação de segredo",
    "h": "Indisciplina/insubordinação",
    "i": "Abandono de emprego",
    "j": "Ofensa contra pessoa",
    "k": "Ofensa à chefia",
    "l": "Jogos de azar",
    "m": "Perda de habilitação",
}
KNOWN_ALINEA_COMBINATIONS = (("a", "e"), ("b", "h"), ("e", "h"), ("e", "k"))
HEADER_ALIASES = {
    "matricula": {"matricula", "registro", "codigo_colaborador", "codigo_do_colaborador", "codigo_funcionario", "codigo_do_funcionario"},
    "tipo": {"tipo", "medida", "tipo_medida", "medida_disciplinar"},
    "motivo": {"motivo", "razao", "ocorrencia"},
    "data_medida": {"data", "data_medida", "data_aplicacao", "data_ocorrencia"},
    "quantidade_dias": {"dias", "dias_suspensao", "quantidade_dias"},
    "observacao": {"observacao", "observacoes", "descricao", "detalhes"},
}
REQUIRED_COLUMNS = {"matricula", "tipo", "motivo", "data_medida"}


def normalize_text(value):
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def parse_measure_type(value):
    normalized = normalize_text(value)
    aliases = {
        "advertencia": "advertencia",
        "advertencia_escrita": "advertencia",
        "suspensao": "suspensao",
        "suspensao_disciplinar": "suspensao",
    }
    return aliases.get(normalized)


def parse_alinea_codes(value):
    """Extrai e padroniza as alíneas mesmo com as grafias encontradas no relatório."""
    raw = unicodedata.normalize("NFKD", str(value or "").strip())
    raw = "".join(char for char in raw if not unicodedata.combining(char))
    raw = re.sub(r"(?<=\s)e(?=\s)", " ", raw)  # conjunção em "B e H"
    normalized = normalize_text(raw)
    if len(normalized) == 1 and normalized in ARTICLE_482_SHORT_LABELS:
        return (normalized,)

    prefixes = ("alinea", "aliena", "aline", "alnea", "al_nea")
    prefix = next((item for item in prefixes if normalized.startswith(item)), None)
    if not prefix:
        return ()

    tail = normalized[len(prefix):].strip("_")
    if not tail:
        return ()
    tokens = [token for token in tail.split("_") if len(token) == 1 and token in ARTICLE_482_SHORT_LABELS]
    if not tokens and re.fullmatch(r"[a-m]+", tail):
        tokens = list(tail)
    return tuple(sorted(set(tokens)))


def alinea_reason(codes):
    return f"alinea_{'_'.join(codes)}" if codes else None


def parse_reason(value):
    codes = parse_alinea_codes(value)
    if codes:
        return alinea_reason(codes)
    normalized = normalize_text(value)
    aliases = {
        "falta": "falta_injustificada",
        "falta_injustificada": "falta_injustificada",
        "ausencia_injustificada": "falta_injustificada",
        "atraso": "atraso",
        "insubordinacao": "insubordinacao",
        "conduta_inadequada": "conduta_inadequada",
        "mau_comportamento": "conduta_inadequada",
        "descumprimento_de_norma": "descumprimento_norma",
        "descumprimento_norma": "descumprimento_norma",
        "outro": "outro",
        "outros": "outro",
    }
    if not normalized:
        return None
    return aliases.get(normalized, "outro")


def parse_reason_detail(value):
    reason = parse_reason(value)
    if not reason or (reason != "outro" and not reason.startswith("alinea_")):
        return None
    return str(value).strip()[:255] or None


def reason_label(reason, detail=None):
    canonical = parse_reason(detail) if reason == "outro" and detail else reason
    if canonical and canonical.startswith("alinea_"):
        codes = tuple(code for code in canonical.removeprefix("alinea_").split("_") if code)
        letters = "/".join(code.upper() for code in codes)
        descriptions = " + ".join(ARTICLE_482_SHORT_LABELS.get(code, code.upper()) for code in codes)
        prefix = "Alínea" if len(codes) == 1 else "Alíneas"
        return f"{prefix} {letters} — {descriptions}"
    return LEGACY_REASONS.get(canonical, detail or canonical or "Não informado")


def reason_filter_options():
    values = [f"alinea_{letter}" for letter in ARTICLE_482_SHORT_LABELS]
    values.extend(alinea_reason(codes) for codes in KNOWN_ALINEA_COMBINATIONS)
    return [{"value": value, "label": reason_label(value)} for value in values]


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    return None


def disciplinary_guidance(reason, warning_count, suspension_count):
    messages = []
    if reason == "falta_injustificada" and int(warning_count or 0) >= 3:
        messages.append("Verifique com o RH se é possível aplicar suspensão.")
    if int(suspension_count or 0) >= 3:
        messages.append("Verifique com o RH sobre possível justa causa.")
    return messages


def measure_fingerprint(employee_id, measure_type, reason, measure_date, days, observation, reason_detail=None):
    normalized_observation = " ".join(str(observation or "").lower().split())
    detail_key = "" if str(reason).startswith("alinea_") else normalize_text(reason_detail)
    source = "|".join([
        str(employee_id), measure_type, reason, measure_date.isoformat(),
        str(days or ""), detail_key, normalized_observation,
    ])
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def selected_values(args, name):
    return {
        value.strip()
        for raw in args.getlist(name)
        for value in str(raw).split(",")
        if value.strip() and value.strip() != "__all__"
    }


def parse_filters(args):
    raw_departments = selected_values(args, "departamento")
    raw_centers = selected_values(args, "centro_custo")
    try:
        departments = {int(value) for value in raw_departments}
        centers = {int(value) for value in raw_centers}
    except ValueError as error:
        raise ValueError("Filtro de departamento ou centro de custo inválido.") from error

    filters = {
        "tipos": selected_values(args, "tipo"),
        "motivos": selected_values(args, "motivo"),
        "colaboradores": selected_values(args, "colaborador_id"),
        "supervisores": selected_values(args, "supervisor"),
        "busca": str(args.get("busca") or "").strip(),
        "departamentos": departments,
        "centros": centers,
        "inicio": parse_date(args.get("inicio")) if args.get("inicio") else None,
        "fim": parse_date(args.get("fim")) if args.get("fim") else None,
    }
    if args.get("inicio") and not filters["inicio"]:
        raise ValueError("Data inicial inválida.")
    if args.get("fim") and not filters["fim"]:
        raise ValueError("Data final inválida.")
    if filters["inicio"] and filters["fim"] and filters["inicio"] > filters["fim"]:
        raise ValueError("A data inicial não pode ser posterior à data final.")
    return filters


def parse_pagination(args):
    try:
        page = int(args.get("page", 1))
        per_page = int(args.get("per_page", 10))
    except (TypeError, ValueError):
        raise ValueError("Paginação inválida.")
    if page < 1:
        raise ValueError("A página deve ser maior que zero.")
    if per_page < 1 or per_page > 100:
        raise ValueError("A quantidade por página deve estar entre 1 e 100.")
    return page, per_page


def parse_import_spreadsheet(stream):
    """Converte a planilha em linhas normalizadas; vínculos e permissões ficam no serviço."""
    try:
        stream.seek(0)
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as error:
        raise ValueError("O arquivo não é uma planilha .xlsx válida.") from error

    sheet = workbook.active
    first_rows = list(sheet.iter_rows(min_row=1, max_row=10, values_only=True))
    is_standard_report = any(
        "advertencias_e_suspensoes" in normalize_text(cell)
        for row in first_rows
        for cell in row
        if cell not in (None, "")
    )
    if is_standard_report:
        parsed, rejected = _parse_standard_report(sheet)
        workbook.close()
        return parsed, rejected

    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        raise ValueError("A planilha está vazia.")

    columns = {}
    for index, header in enumerate(headers):
        normalized = normalize_text(header)
        for canonical, aliases in HEADER_ALIASES.items():
            if normalized in aliases and canonical not in columns:
                columns[canonical] = index
                break
    missing = sorted(REQUIRED_COLUMNS - set(columns))
    if missing:
        labels = {"matricula": "Matrícula", "tipo": "Tipo", "motivo": "Motivo", "data_medida": "Data"}
        raise ValueError("Colunas obrigatórias ausentes: " + ", ".join(labels[item] for item in missing) + ".")

    parsed = []
    rejected = []
    fingerprints = set()
    for row_number, values in enumerate(iterator, start=2):
        if row_number > MAX_IMPORT_ROWS:
            raise ValueError(f"A planilha deve possuir no máximo {MAX_IMPORT_ROWS:,} linhas.")
        if not any(value not in (None, "") for value in values):
            continue

        def cell(name):
            index = columns.get(name)
            return values[index] if index is not None and index < len(values) else None

        errors = []
        try:
            employee_number = int(float(str(cell("matricula")).replace(",", ".")))
        except (TypeError, ValueError):
            employee_number = None
            errors.append("Matrícula inválida")
        measure_type = parse_measure_type(cell("tipo"))
        if not measure_type:
            errors.append("Tipo deve ser Advertência ou Suspensão")
        reason = parse_reason(cell("motivo"))
        if not reason:
            errors.append("Motivo não reconhecido")
        measure_date = parse_date(cell("data_medida"))
        if not measure_date:
            errors.append("Data inválida")

        days = None
        raw_days = cell("quantidade_dias")
        if raw_days not in (None, ""):
            try:
                days = int(raw_days)
            except (TypeError, ValueError):
                errors.append("Quantidade de dias inválida")
        if measure_type == "suspensao" and (days is None or days < 1):
            errors.append("Informe ao menos 1 dia para a suspensão")
        if measure_type == "advertencia":
            days = None
        observation = str(cell("observacao") or "").strip() or None
        reason_detail = parse_reason_detail(cell("motivo"))
        detail_key = None if reason and reason.startswith("alinea_") else reason_detail
        local_key = (employee_number, measure_type, reason, detail_key, measure_date, days, " ".join(str(observation or "").lower().split()))
        if not errors and local_key in fingerprints:
            errors.append("Linha duplicada dentro da planilha")
        fingerprints.add(local_key)

        if errors:
            rejected.append({"linha": row_number, "motivo": "; ".join(errors)})
        else:
            parsed.append({
                "linha": row_number,
                "matricula": employee_number,
                "tipo": measure_type,
                "motivo": reason,
                "motivo_detalhe": reason_detail,
                "reincidencia": None,
                "data_medida": measure_date,
                "quantidade_dias": days,
                "observacao": observation,
            })
    workbook.close()
    return parsed, rejected


def _yes_no(value):
    normalized = normalize_text(value)
    if normalized in {"sim", "s", "yes"}:
        return True
    if normalized in {"nao", "n", "no"}:
        return False
    return None


def _parse_standard_report(sheet):
    """Lê o relatório padrão Relação de Advertências e Suspensões."""
    parsed = []
    rejected = []
    seen = set()
    current_employee_number = None
    employee_pattern = re.compile(r"^\s*(\d+)\s*-\s*(.+?)\s*$")

    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > MAX_IMPORT_ROWS:
            raise ValueError(f"A planilha deve possuir no máximo {MAX_IMPORT_ROWS:,} linhas.")
        employee_cell = values[1] if len(values) > 1 else None  # coluna B
        employee_match = employee_pattern.match(str(employee_cell or ""))
        if employee_match:
            current_employee_number = int(employee_match.group(1))

        raw_date = values[3] if len(values) > 3 else None       # coluna D
        measure_date = parse_date(raw_date)
        if not measure_date:
            continue
        if current_employee_number is None:
            rejected.append({"linha": row_number, "motivo": "Ocorrência sem colaborador identificado nas linhas anteriores"})
            continue

        raw_reason = values[7] if len(values) > 7 else None     # coluna H
        reason = parse_reason(raw_reason)
        if not reason:
            rejected.append({"linha": row_number, "motivo": "Motivo não informado"})
            continue
        suspended = _yes_no(values[12] if len(values) > 12 else None)  # coluna M
        measure_type = "suspensao" if suspended is True else "advertencia"
        days = None
        if measure_type == "suspensao":
            try:
                days = int(values[15])  # coluna P
            except (TypeError, ValueError):
                rejected.append({"linha": row_number, "motivo": "Suspensão sem quantidade de dias válida"})
                continue
            if days < 1:
                rejected.append({"linha": row_number, "motivo": "A suspensão deve possuir ao menos 1 dia"})
                continue
        reason_detail = parse_reason_detail(raw_reason)
        recurrence = _yes_no(values[10] if len(values) > 10 else None)  # coluna K
        detail_key = None if reason.startswith("alinea_") else reason_detail
        local_key = (current_employee_number, measure_type, reason, detail_key, measure_date, days)
        if local_key in seen:
            rejected.append({"linha": row_number, "motivo": "Linha duplicada dentro da planilha"})
            continue
        seen.add(local_key)
        parsed.append({
            "linha": row_number,
            "matricula": current_employee_number,
            "tipo": measure_type,
            "motivo": reason,
            "motivo_detalhe": reason_detail,
            "reincidencia": recurrence,
            "data_medida": measure_date,
            "quantidade_dias": days,
            "observacao": None,
        })
    if not parsed and not rejected:
        raise ValueError("Nenhuma ocorrência foi encontrada no relatório padrão.")
    return parsed, rejected


class DisciplinaryMeasureService:
    @staticmethod
    def _permission(token_data, action):
        if has_permission(token_data, SCREEN, action):
            return None
        return jsonify("Você não possui permissão para esta operação."), 403

    @staticmethod
    def _counts_for_employees(employee_ids):
        employee_ids = {int(employee_id) for employee_id in employee_ids if employee_id is not None}
        if not employee_ids:
            return {}
        rows = (
            db.session.query(
                DisciplinaryMeasure.colaborador_id,
                func.sum(case((DisciplinaryMeasure.tipo == "advertencia", 1), else_=0)).label("advertencias"),
                func.sum(case((DisciplinaryMeasure.tipo == "suspensao", 1), else_=0)).label("suspensoes"),
            )
            .filter(
                DisciplinaryMeasure.colaborador_id.in_(employee_ids),
            )
            .group_by(DisciplinaryMeasure.colaborador_id)
            .all()
        )
        return {
            row.colaborador_id: {
                "advertencias": int(row.advertencias or 0),
                "suspensoes": int(row.suspensoes or 0),
            }
            for row in rows
        }

    @staticmethod
    def _supervisor_snapshots(center_ids):
        center_ids = {int(center_id) for center_id in center_ids if center_id is not None}
        if not center_ids:
            return {}
        rows = (
            db.session.query(
                CostCenters.id.label("centro_id"),
                Supervisors.id.label("supervisor_id"),
                Supervisors.nome.label("supervisor_nome"),
            )
            .outerjoin(Supervisors, Supervisors.id == CostCenters.supervisor_id)
            .filter(CostCenters.id.in_(center_ids))
            .all()
        )
        return {
            row.centro_id: {
                "supervisor_id": row.supervisor_id,
                "supervisor_nome": row.supervisor_nome or "Sem supervisor",
            }
            for row in rows
        }

    @staticmethod
    def _serialize(record, center=None):
        employee = record.colaborador
        if center is None and employee and employee.centro_id:
            center = db.session.get(CostCenters, employee.centro_id)
        canonical_reason = parse_reason(record.motivo_detalhe) if record.motivo == "outro" else record.motivo
        return {
            "id": record.id,
            "colaborador_id": record.colaborador_id,
            "matricula": employee.matricula if employee else None,
            "colaborador": employee.nome if employee else None,
            "centro_custo_id": center.id if center else None,
            "centro_custo": center.local if center else None,
            "departamento": center.departamento if center else None,
            "tipo": record.tipo,
            "tipo_label": MEASURE_TYPES.get(record.tipo, record.tipo),
            "supervisor_id": record.supervisor_id,
            "supervisor": record.supervisor_nome,
            "motivo": canonical_reason,
            "motivo_label": reason_label(canonical_reason, record.motivo_detalhe),
            "motivo_detalhe": record.motivo_detalhe,
            "data_medida": record.data_medida.isoformat(),
            "quantidade_dias": record.quantidade_dias,
            "observacao": record.observacao,
            "origem": record.origem,
            "arquivo_origem": record.arquivo_origem,
            "linha_origem": record.linha_origem,
            "criado_por": record.criado_por.nome if record.criado_por else None,
            "criado_em": record.criado_em.isoformat() if record.criado_em else None,
        }

    @safe_route
    def read(self, token_data):
        denied = self._permission(token_data, "view")
        if denied:
            return denied
        try:
            filters = parse_filters(request.args)
            page, per_page = parse_pagination(request.args)
        except ValueError as error:
            return jsonify(str(error)), 400

        query = DisciplinaryMeasure.query.join(Employees, Employees.id == DisciplinaryMeasure.colaborador_id)
        query = apply_cost_center_scope(query, Employees.centro_id, token_data)
        if filters["tipos"]:
            query = query.filter(DisciplinaryMeasure.tipo.in_(filters["tipos"]))
        if filters["motivos"]:
            query = query.filter(DisciplinaryMeasure.motivo.in_(filters["motivos"]))
        if filters["colaboradores"]:
            try:
                ids = {int(value) for value in filters["colaboradores"]}
            except ValueError:
                return jsonify("Filtro de colaborador inválido."), 400
            query = query.filter(DisciplinaryMeasure.colaborador_id.in_(ids))

        if filters["departamentos"]:
            query = query.join(CostCenters,CostCenters.id == Employees.centro_id,)               
            query = query.filter(CostCenters.departamento.in_(filters["departamentos"]))
        if filters["centros"]:
            query = query.filter(Employees.centro_id.in_(filters["centros"]))
        if filters["supervisores"]:
            query = query.filter(DisciplinaryMeasure.supervisor_nome.in_(filters["supervisores"]))
        if filters["inicio"]:
            query = query.filter(DisciplinaryMeasure.data_medida >= filters["inicio"])
        if filters["fim"]:
            query = query.filter(DisciplinaryMeasure.data_medida <= filters["fim"])
        if filters["busca"]:
            term = f"%{filters['busca']}%"
            query = query.filter(or_(
                Employees.nome.ilike(term),
                cast(Employees.matricula, String).ilike(term),
                DisciplinaryMeasure.supervisor_nome.ilike(term),
                DisciplinaryMeasure.motivo_detalhe.ilike(term),
                DisciplinaryMeasure.observacao.ilike(term),
            ))
        # As opções são facetas do mesmo conjunto filtrado, nunca um catálogo
        # fixo. Assim um tipo ou alínea sem ocorrência no recorte não aparece.
        option_values = query.with_entities(
            DisciplinaryMeasure.tipo,
            DisciplinaryMeasure.motivo,
            DisciplinaryMeasure.motivo_detalhe,
        ).distinct().all()
        available_types = {
            tipo for tipo, _motivo, _detail in option_values if tipo in MEASURE_TYPES
        }
        available_reasons = {
            (motivo, detail)
            for _tipo, motivo, detail in option_values
            if motivo
        }
        summary = query.with_entities(
            func.count(DisciplinaryMeasure.id).label("total"),
            func.sum(case((DisciplinaryMeasure.tipo == "advertencia", 1), else_=0)).label("advertencias"),
            func.sum(case((DisciplinaryMeasure.tipo == "suspensao", 1), else_=0)).label("suspensoes"),
        ).one()
        records = (
            query
            .options(joinedload(DisciplinaryMeasure.criado_por))
            .order_by(DisciplinaryMeasure.data_medida.desc(), DisciplinaryMeasure.id.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )
        center_ids = {record.colaborador.centro_id for record in records if record.colaborador and record.colaborador.centro_id}
        centers = {
            center.id: center
            for center in CostCenters.query.filter(CostCenters.id.in_(center_ids)).all()
        } if center_ids else {}
        return jsonify({
            "registros": [
                self._serialize(record, centers.get(record.colaborador.centro_id) if record.colaborador else None)
                for record in records
            ],
            "resumo": {
                "total": int(summary.total or 0),
                "advertencias": int(summary.advertencias or 0),
                "suspensoes": int(summary.suspensoes or 0),
            },
            "paginacao": {
                "page": page,
                "per_page": per_page,
                "total": int(summary.total or 0),
            },
            "opcoes": {
                "tipos": [
                    {"value": key, "label": label}
                    for key, label in MEASURE_TYPES.items()
                    if key in available_types
                ],
                "motivos": [
                    {"value": motivo, "label": reason_label(motivo, detail)}
                    for motivo, detail in sorted(
                        available_reasons,
                        key=lambda item: reason_label(item[0], item[1]).casefold(),
                    )
                ],
            },
        }),200

    @safe_route
    def filter_options(self, token_data):
        denied = self._permission(token_data, "view")
        if denied:
            return denied
        query = (
            db.session.query(
                Employees.id,
                Employees.matricula,
                Employees.nome,
            )
            .join(DisciplinaryMeasure, DisciplinaryMeasure.colaborador_id == Employees.id)
        )
        query = apply_cost_center_scope(query, Employees.centro_id, token_data)
        employees = query.distinct().order_by(Employees.nome).all()
        supervisor_query = (
            db.session.query(DisciplinaryMeasure.supervisor_nome)
            .join(Employees, Employees.id == DisciplinaryMeasure.colaborador_id)
            .filter(DisciplinaryMeasure.supervisor_nome.isnot(None))
        )
        supervisor_query = apply_cost_center_scope(supervisor_query, Employees.centro_id, token_data)
        supervisors = [name for name, in supervisor_query.distinct().order_by(DisciplinaryMeasure.supervisor_nome).all()]
        department_query = (
            db.session.query(CostCenters.departamento)
            .join(Employees, Employees.centro_id == CostCenters.id)
            .join(DisciplinaryMeasure, DisciplinaryMeasure.colaborador_id == Employees.id)
            .filter(CostCenters.departamento.isnot(None))
        )
        department_query = apply_cost_center_scope(department_query, CostCenters.id, token_data)
        departments = [department for department, in department_query.distinct().order_by(CostCenters.departamento).all()]
        center_query = (
            db.session.query(CostCenters.id, CostCenters.local)
            .join(Employees, Employees.centro_id == CostCenters.id)
            .join(DisciplinaryMeasure, DisciplinaryMeasure.colaborador_id == Employees.id)
        )
        center_query = apply_cost_center_scope(center_query, CostCenters.id, token_data)
        centers = center_query.distinct().order_by(CostCenters.local).all()
        
        return jsonify({
            "colaboradores": [{"id": employee.id, "matricula": employee.matricula, "nome": employee.nome}for employee in employees],
            "supervisores": [{"value": name, "label": name} for name in supervisors],
            "departamentos": [{"value": department, "label": f"DPTO. {department}"}for department in departments],        
            "centros": [{"value": center_id, "label": local or "Sem local"} for center_id, local in centers],
        }), 200

    @safe_route
    def delete_all(self, token_data):
        if not is_admin(token_data):
            return jsonify("Apenas administradores podem excluir todos os registros."), 403

        deleted = DisciplinaryMeasure.query.delete(synchronize_session=False)
        db.session.commit()
        return jsonify({
            "message": "Todos os registros de medidas disciplinares foram excluídos.",
            "excluidos": int(deleted or 0),
        }), 200

    @safe_route
    def import_xlsx(self, token_data):
        denied = self._permission(token_data, "create")
        if denied:
            return denied
        uploaded = request.files.get("arquivo")
        if not uploaded or not uploaded.filename:
            return jsonify("Selecione uma planilha .xlsx."), 400
        if not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify("Somente arquivos .xlsx são aceitos."), 400
        if request.content_length and request.content_length > MAX_IMPORT_BYTES:
            return jsonify("A planilha deve possuir no máximo 10 MB."), 413
        try:
            parsed, rejected = parse_import_spreadsheet(uploaded.stream)
        except ValueError as error:
            return jsonify(str(error)), 400

        employee_numbers = {row["matricula"] for row in parsed}
        employees = Employees.query.filter(Employees.matricula.in_(employee_numbers)).all()
        employee_by_number = {employee.matricula: employee for employee in employees}
        supervisors_by_center = self._supervisor_snapshots({employee.centro_id for employee in employees})
        allowed_ids = allowed_cost_center_ids(token_data)
        candidates = []
        for row in parsed:
            employee = employee_by_number.get(row["matricula"])
            if not employee:
                rejected.append({"linha": row["linha"], "motivo": "Colaborador não encontrado pela matrícula"})
                continue
            if not employee.centro_id or (allowed_ids is not None and employee.centro_id not in allowed_ids):
                rejected.append({"linha": row["linha"], "motivo": "Sem permissão para o colaborador desta linha"})
                continue
            fingerprint = measure_fingerprint(employee.id, row["tipo"], row["motivo"], row["data_medida"], row["quantidade_dias"], row["observacao"], row.get("motivo_detalhe"))
            legacy_fingerprint = None
            if row["motivo"].startswith("alinea_"):
                legacy_fingerprint = measure_fingerprint(
                    employee.id,
                    row["tipo"],
                    "outro",
                    row["data_medida"],
                    row["quantidade_dias"],
                    row["observacao"],
                    row.get("motivo_detalhe"),
                )
            supervisor = supervisors_by_center.get(
                employee.centro_id,
                {"supervisor_id": None, "supervisor_nome": "Sem supervisor"},
            )
            candidates.append((row, employee, fingerprint, legacy_fingerprint, supervisor))

        existing_fingerprints = set()
        candidate_fingerprints = [
            candidate_hash
            for _, _, fingerprint, legacy_fingerprint, _ in candidates
            for candidate_hash in (fingerprint, legacy_fingerprint)
            if candidate_hash
        ]
        for start in range(0, len(candidate_fingerprints), 1000):
            batch = candidate_fingerprints[start:start + 1000]
            existing_fingerprints.update(
                fingerprint
                for fingerprint, in (
                    db.session.query(DisciplinaryMeasure.fingerprint)
                    .filter(DisciplinaryMeasure.fingerprint.in_(batch))
                    .all()
                )
            )

        count_cache = self._counts_for_employees({employee.id for _, employee, _, _, _ in candidates})
        records_to_add = []
        pending_report = []
        stored_filename = (secure_filename(uploaded.filename) or "importacao.xlsx")[:255]
        for row, employee, fingerprint, legacy_fingerprint, supervisor in candidates:
            if fingerprint in existing_fingerprints or legacy_fingerprint in existing_fingerprints:
                rejected.append({"linha": row["linha"], "motivo": "Medida duplicada; já existe no sistema"})
                continue
            counts = count_cache.setdefault(employee.id, {"advertencias": 0, "suspensoes": 0})
            guidance = disciplinary_guidance(row["motivo"], counts["advertencias"], counts["suspensoes"])
            record = DisciplinaryMeasure(
                colaborador_id=employee.id,
                tipo=row["tipo"],
                motivo=row["motivo"],
                motivo_detalhe=row.get("motivo_detalhe"),
                reincidencia=row.get("reincidencia"),
                data_medida=row["data_medida"],
                quantidade_dias=row["quantidade_dias"],
                observacao=row["observacao"],
                supervisor_id=supervisor["supervisor_id"],
                supervisor_nome=supervisor["supervisor_nome"],
                fingerprint=fingerprint,
                origem="importacao",
                arquivo_origem=stored_filename,
                linha_origem=row["linha"],
                criado_por_usuario_id=token_data.get("id"),
            )
            records_to_add.append(record)
            pending_report.append((row, employee, record, guidance))
            existing_fingerprints.add(fingerprint)
            counts["advertencias" if row["tipo"] == "advertencia" else "suspensoes"] += 1

        imported = []
        if records_to_add:
            try:
                db.session.add_all(records_to_add)
                db.session.flush()
                imported = [
                    {
                        "linha": row["linha"],
                        "id": record.id,
                        "matricula": employee.matricula,
                        "avisos": guidance,
                    }
                    for row, employee, record, guidance in pending_report
                ]
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                return jsonify(
                    "Os dados foram alterados durante a importação. Tente novamente para que as duplicidades sejam reavaliadas."
                ), 409
        rejected.sort(key=lambda item: item["linha"])
        return jsonify({
            "message": "Importação concluída.",
            "importadas": len(imported),
            "rejeitadas": len(rejected),
            "linhas_importadas": imported,
            "linhas_rejeitadas": rejected,
        }), 200

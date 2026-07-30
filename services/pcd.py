from flask import jsonify, request as rq
from openpyxl import load_workbook

from models.cargos import Cargos
from models.centros_de_custo import CostCenters
from models.colaboradores import Employees, db
from models.filiais import Branch, filial_departamentos
from models.situacoes import Situations
from models.supervisores import Supervisors
from utils.filial_scope import apply_cost_center_scope, can_access_cost_center, is_admin
from utils.safe_route import safe_route

# Colunas de tipo de deficiência do relatório "Relação de Empregados - Cadastro".
TIPO_COLUNAS = ["motora", "visual", "auditiva", "intelectual", "outras", "reabilitado"]
TIPO_LABELS = {
    "motora": "Motora",
    "visual": "Visual",
    "auditiva": "Auditiva",
    "intelectual": "Intelectual",
    "outras": "Outras",
    "reabilitado": "Reabilitado",
}


def _filiais_by_departamento(departamentos):
    """Mapeia código de departamento -> lista de nomes de filial cadastradas em
    `filial_departamentos`. Um mesmo código pode pertencer a mais de uma filial (ex: Matriz
    e Londrina usando o mesmo número) — nesse caso todas aparecem, para exibição em tag."""
    departamentos = {d for d in departamentos if d is not None}
    if not departamentos:
        return {}

    rows = (
        db.session.query(filial_departamentos.c.departamento, Branch.nome)
        .join(Branch, Branch.id == filial_departamentos.c.filial_id)
        .filter(filial_departamentos.c.departamento.in_(departamentos))
        .distinct()
        .all()
    )
    result = {}
    for departamento, filial_nome in rows:
        result.setdefault(departamento, []).append(filial_nome)
    for departamento, nomes in result.items():
        nomes.sort(key=str)
    return result


class PcdService:
    @safe_route
    def read(self, token_data):
        """Retorna a lista plana de colaboradores PCD (departamento, centro de custo e
        supervisor), respeitando o mesmo escopo de acesso usado no Colab. por DPTO."""
        query = (
            db.session.query(
                Employees.id,
                Employees.matricula,
                Employees.nome,
                Employees.type_pcd,
                Employees.obs_pcd,
                Employees.centro_id,
                CostCenters.local.label("centro_local"),
                CostCenters.departamento,
                Supervisors.id.label("supervisor_id"),
                Supervisors.nome.label("supervisor_nome"),
                Cargos.nome.label("cargo"),
                Situations.id.label("situacao_id"),
                Situations.tipo.label("situacao"),
            )
            .select_from(Employees)
            .outerjoin(CostCenters, CostCenters.id == Employees.centro_id)
            .outerjoin(Supervisors, Supervisors.id == CostCenters.supervisor_id)
            .outerjoin(Cargos, Cargos.id == Employees.cargo)
            .outerjoin(Situations, Situations.id == Employees.situacao)
            .filter(Employees.pcd.is_(True))
            .order_by(Employees.nome.asc())
        )
        query = apply_cost_center_scope(query, Employees.centro_id, token_data)
        rows = query.all()

        colaboradores = [{
            "id": r.id,
            "matricula": r.matricula,
            "nome": r.nome,
            "cargo": r.cargo,
            "type_pcd": r.type_pcd,
            "obs_pcd": r.obs_pcd,
            "centro_id": r.centro_id,
            "centro_custo": r.centro_local or "Sem centro de custo",
            "departamento": r.departamento,
            "supervisor_id": r.supervisor_id,
            "supervisor": r.supervisor_nome or "Sem supervisor",
            "situacao_id": r.situacao_id,
            "situacao": r.situacao or "Não informada",
        } for r in rows]

        filiais_por_departamento = _filiais_by_departamento({r.departamento for r in rows})

        return jsonify({
            "total": len(colaboradores),
            "colaboradores": colaboradores,
            "filiais_por_departamento": filiais_por_departamento,
        }), 200

        return jsonify({"total": len(colaboradores), "colaboradores": colaboradores}), 200

    @safe_route
    def update(self, token_data):
        """Marca/desmarca um colaborador como PCD ou atualiza seus dados (tipo/observação)."""
        body = rq.get_json() or {}
        employee_id = body.get("id")
        if not employee_id:
            return jsonify("Informe o colaborador."), 400

        employee = db.session.get(Employees, employee_id)
        if not employee:
            return jsonify("Colaborador não encontrado."), 404
        if employee.centro_id and not can_access_cost_center(token_data, employee.centro_id):
            return jsonify("Você não possui acesso à filial deste colaborador."), 403

        if "pcd" in body:
            employee.pcd = bool(body.get("pcd"))

        if not employee.pcd:
            # Sem a condição de PCD, os detalhes não fazem mais sentido.
            employee.type_pcd = None
            employee.obs_pcd = None
        else:
            if "type_pcd" in body:
                employee.type_pcd = body.get("type_pcd") or None
            if "obs_pcd" in body:
                employee.obs_pcd = body.get("obs_pcd") or None

        db.session.commit()
        return jsonify(employee.to_dict()), 200

    @safe_route
    def import_xlsx(self, token_data):
        """
        Importa o relatório de PCD (.xlsx) e marca os colaboradores encontrados pela matrícula.
        A matrícula é a chave de correspondência: mais simples e confiável do que casar
        nome/departamento em texto livre.
        """
        uploaded = rq.files.get("file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify("Envie uma planilha no formato .xlsx."), 400

        try:
            workbook = load_workbook(uploaded.stream, read_only=True, data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
        except (ValueError, OSError):
            return jsonify("Não foi possível ler a planilha."), 400

        header_index = next(
            (i for i, row in enumerate(rows) if row and str(row[0] or "").strip().lower() == "cód. emp."),
            None,
        )
        if header_index is None:
            return jsonify("Planilha fora do padrão esperado (cabeçalho 'Cód. Emp.' não encontrado)."), 400

        header = [str(v).strip().lower() if v is not None else None for v in rows[header_index]]

        def col(label):
            return header.index(label) if label in header else None

        idx_matricula = col("matricula")
        idx_obs = col("observação")
        idx_tipos = {tipo: col(tipo) for tipo in TIPO_COLUNAS}
        if idx_matricula is None:
            return jsonify("A coluna 'Matricula' não foi encontrada na planilha."), 400

        updated, not_found, skipped, no_access = [], [], 0, 0
        for row in rows[header_index + 1:]:
            if not row or idx_matricula >= len(row):
                continue
            matricula = row[idx_matricula]
            if matricula in (None, ""):
                skipped += 1
                continue
            try:
                matricula = int(matricula)
            except (TypeError, ValueError):
                skipped += 1
                continue

            employee = Employees.query.filter_by(matricula=matricula).first()
            if not employee:
                not_found.append(matricula)
                continue
            if employee.centro_id and not can_access_cost_center(token_data, employee.centro_id):
                no_access += 1
                continue

            tipos_marcados = [
                TIPO_LABELS[tipo] for tipo, position in idx_tipos.items()
                if position is not None and position < len(row)
                and str(row[position] or "").strip().upper() == "SIM"
            ]
            employee.pcd = True
            if tipos_marcados:
                employee.type_pcd = ", ".join(tipos_marcados)
            observacao = row[idx_obs] if idx_obs is not None and idx_obs < len(row) else None
            if observacao:
                employee.obs_pcd = str(observacao).strip()
            updated.append(matricula)

        db.session.commit()
        return jsonify({
            "message": f"{len(updated)} colaborador(es) marcados como PCD.",
            "atualizados": len(updated),
            "nao_encontrados": not_found,
            "ignorados": skipped,
            "sem_acesso": no_access,
        }), 200

    @safe_route
    def delete_all(self, token_data):
        """Zera o indicador de PCD de todos os colaboradores. Somente ADMIN, para casos de erro na importação."""
        if not is_admin(token_data):
            return jsonify("Apenas administradores podem excluir todos os dados de PCD."), 403

        affected = Employees.query.filter(Employees.pcd.is_(True)).update(
            {Employees.pcd: False, Employees.type_pcd: None, Employees.obs_pcd: None},
            synchronize_session=False,
        )
        db.session.commit()
        return jsonify({
            "message": f"{affected} colaborador(es) tiveram o indicador de PCD removido.",
            "removidos": affected,
        }), 200

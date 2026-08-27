# Modelos
from models.centros_de_custo import CostCenters, db
from models.controle_faltas import AbsenceControl
from models.glosas import Disallowance
from models.rp_requisicao import Requisicao
from models.rp_timeline import Timeline
from models.supervisores import Supervisors
from models.colaboradores import Employees
from models.rp_historico import History
from models.cargos import Cargos
from models.usuarios import Users
from models.reservas_tecnicas import Floaters
from models.situacoes import Situations
from models.medidas_disciplinares import DisciplinaryMeasure

# Utilitários
from datetime import date, datetime as dt, timedelta
from decimal import Decimal
from dateutils import relativedelta
from flask import jsonify, request, send_file
import re
from utils.socket import socketio
from calendar import monthrange
from sqlalchemy import and_, case, func, or_
from utils.check_field import check_field
from utils.safe_route import safe_route
from sqlalchemy.orm import aliased
from io import BytesIO
from openpyxl import Workbook, load_workbook
from zoneinfo import ZoneInfo
from utils.filial_scope import (
    apply_active_department_scope,
    apply_cost_center_scope,
    can_access_cost_center,
    can_access_supervisor_user,
    is_admin,
    supervisor_users_query,
)
from utils.permissions import has_permission
from services.controle_faltas import AbsenceControlService
from services.medidas_disciplinares import disciplinary_guidance

def _emit_kds_update(action, request_id=None, status=None):
    """Notify TV dashboards without exposing requisition data over the socket."""
    socketio.emit("kds_update", {
        "action": action,
        "request_id": request_id,
        "status": status,
        "emitted_at": dt.now(ZoneInfo("America/Sao_Paulo")).isoformat(),
    })

def _can_access_employee(token_data, employee_id, allow_uncovered=False):
    if allow_uncovered and employee_id in (None, 0, "0"):
        return True
    try:
        employee = db.session.get(Employees, int(employee_id))
    except (TypeError, ValueError):
        return False
    return bool(
        employee
        and employee.centro_id
        and can_access_cost_center(token_data, employee.centro_id)
    )

class RequestService:
    """Owns requisition validation, date ranges, spreadsheet I/O and queue queries."""
    REASONS = {"AFASTAMENTO", "ATESTADO", "DECLARAÇÃO", "FÉRIAS", "FERIAS", "POSTO VAGO", "REMANEJAMENTO", "INJUSTIFICADA"}
    ISNOTFAULT = ["FÉRIAS", "FERIAS", "POSTO VAGO", "REMANEJAMENTO", "AFASTAMENTO"]
    OPERATIONAL_COVERAGE_REASON = "COBERTURA OPERACIONAL DE ADICIONAL"
    GAF_VALOR_DIARIO = Decimal("4.36")
    INSALUBRIDADE_VALOR_DIARIO = {
        "10": Decimal("5.40"),
        "20": Decimal("10.81"),
        "40": Decimal("21.61"),
    }
    INSALUBRIDADE_PATTERN = re.compile(r"INSALUB(?:RIDADE|\.)?\s*(10|20|40)\s*%")
    GAF_PATTERN = re.compile(r"G[\s./-]*A[\s./-]*F")

    @safe_route
    def requester(self, token_data):
        """Resolve o usuário supervisor responsável pela nova requisição."""
        user = db.session.get(Users, token_data.get("id"))
        if not user:
            return jsonify("Usuário autenticado não encontrado."), 401
        if not has_permission(token_data, "reposicoes", "create"):
            return jsonify("Você não possui permissão para criar requisições."), 403

        supervisors = supervisor_users_query(token_data).order_by(Users.nome).all()
        options = [{"id": item.id, "nome": item.nome} for item in supervisors]
        if is_admin(token_data):
            return jsonify({
                "pode_selecionar_supervisor": True,
                "supervisor": None,
                "supervisores": options,
            }), 200

        if str(user.role or "").upper() != "SUPERVISOR":
            return jsonify(
                "Seu usuário não possui a role SUPERVISOR para registrar requisições."
            ), 403
        return jsonify({
            "pode_selecionar_supervisor": False,
            "supervisor": {"id": user.id, "nome": user.nome},
            "supervisores": [],
        }), 200

    @classmethod
    def _additional_data(cls, cargo):
        """Calcula insalubridade pelo nome do cargo e preserva adicionais manuais."""
        if not cargo:
            return None, Decimal("0")

        additional_type = str(cargo.adicional_tipo or "").strip() or None
        cargo_name = str(cargo.nome or "").upper()
        match = cls.INSALUBRIDADE_PATTERN.search(cargo_name)
        has_gaf = bool(cls.GAF_PATTERN.search(cargo_name))
        manual_daily_value = Decimal(str(cargo.adicional_valor_diaria or 0))

        # Apenas cargos com percentual explícito são calculados automaticamente.
        # Nomes genéricos como "INSALUBRIDADE" permanecem para preenchimento manual.
        if not match:
            if has_gaf and "INSALUB" not in cargo_name:
                return "GAF", cls.GAF_VALOR_DIARIO
            return additional_type, manual_daily_value

        daily_value = cls.INSALUBRIDADE_VALOR_DIARIO[match.group(1)]
        if has_gaf:
            daily_value += cls.GAF_VALOR_DIARIO

        calculated_type = f"INSALUBRIDADE {match.group(1)}%"
        if has_gaf:
            calculated_type += " + GAF"
        return calculated_type, daily_value

    @classmethod
    def _cargo_data(cls, employee):
        """Lê os valores do cargo sem usar dados enviados pelo navegador."""
        cargo = (
            db.session.get(Cargos, employee.cargo)
            if employee and employee.cargo
            else None
        )
        additional_type, additional_daily_value = cls._additional_data(cargo)
        return {
            "cargo_id": cargo.id if cargo else None,
            "cargo": cargo.nome if cargo else None,
            "adicional_tipo": additional_type,
            "adicional_valor_diaria": additional_daily_value,
            "valor_diaria": Decimal(str(cargo.multa or 0)) if cargo else Decimal("0"),
        }

    @classmethod
    def _additional_context(cls, absent_employee, coverage_employee=None):
        """Define se a cobertura exige adicional e se desloca um posto menor."""
        absent_cargo = cls._cargo_data(absent_employee)
        coverage_cargo = cls._cargo_data(coverage_employee) if coverage_employee else None
        has_additional = bool(
            absent_cargo["adicional_tipo"]
            and absent_cargo["adicional_valor_diaria"] > 0
        )
        same_job_and_additional = bool(
            coverage_cargo
            and absent_cargo["cargo_id"] == coverage_cargo["cargo_id"]
            and absent_cargo["adicional_tipo"] == coverage_cargo["adicional_tipo"]
            and absent_cargo["adicional_valor_diaria"] == coverage_cargo["adicional_valor_diaria"]
        )
        requires_additional = (
            has_additional
            and bool(coverage_cargo)
            and not same_job_and_additional
        )
        coverage_has_lower_value = bool(
            requires_additional
            # Valor zero representa cargo ainda não parametrizado, não um
            # cargo comprovadamente menor para gerar glosa operacional.
            and coverage_cargo["valor_diaria"] > 0
            and coverage_cargo["valor_diaria"] < absent_cargo["valor_diaria"]
        )
        return {
            "has_additional": has_additional,
            "requires_additional": requires_additional,
            "coverage_has_lower_value": coverage_has_lower_value,
            "adicional_tipo": absent_cargo["adicional_tipo"] if requires_additional else None,
            "adicional_valor_diaria": (
                absent_cargo["adicional_valor_diaria"]
                if requires_additional
                else Decimal("0")
            ),
        }

    @staticmethod
    def _coverage_candidates(absent_employee):
        """Retorna somente colegas ativos do mesmo centro de custo do ausente."""
        if not absent_employee or not absent_employee.centro_id:
            return []
        rows = (
            db.session.query(Employees, Cargos.nome.label("cargo"))
            .outerjoin(Cargos, Cargos.id == Employees.cargo)
            .filter(
                Employees.centro_id == absent_employee.centro_id,
                Employees.id != absent_employee.id,
                Employees.situacao == 1,
            )
            .order_by(Employees.nome)
            .all()
        )
        return [
            {
                "id": employee.id,
                "nome": employee.nome,
                "matricula": employee.matricula,
                "centro_id": employee.centro_id,
                "cargo": cargo,
            }
            for employee, cargo in rows
        ]

    @safe_route
    def additional_context(self, token_data):
        """Prepara o item 3 dentro do escopo da sessão autenticada."""
        if not has_permission(token_data, "reposicoes", "create"):
            return jsonify("Você não possui permissão para consultar este contexto."), 403
        body = request.get_json(silent=True) or {}
        try:
            absent_id = int(body.get("ausente_id"))
            reserve_id = int(body.get("reserva_id") or 0)
        except (TypeError, ValueError):
            return jsonify("Informe um colaborador ausente válido."), 400

        absent_employee = db.session.get(Employees, absent_id)
        if not absent_employee:
            return jsonify("Colaborador ausente não encontrado."), 404
        if not absent_employee.centro_id:
            return jsonify("O colaborador ausente não possui centro de custo cadastrado."), 400
        if not can_access_cost_center(token_data, absent_employee.centro_id):
            return jsonify("Você não possui acesso à filial deste colaborador."), 403

        coverage_employee = db.session.get(Employees, reserve_id) if reserve_id else None
        if coverage_employee and not can_access_cost_center(token_data, coverage_employee.centro_id):
            return jsonify("Você não possui acesso à filial desta cobertura."), 403
        context = self._additional_context(absent_employee, coverage_employee)
        if not context["has_additional"]:
            return jsonify({
                "modo": "desabilitado",
                "motivo": "O cargo não possui adicional cadastrado.",
            }), 200
        if coverage_employee:
            return jsonify({
                "modo": "desabilitado",
                "motivo": "A cobertura já foi informada nas reservas técnicas.",
                "adicional": {
                    "tipo": context["adicional_tipo"],
                    "valor_diaria": float(context["adicional_valor_diaria"]),
                } if context["requires_additional"] else None,
                "beneficiario": {
                    "id": coverage_employee.id,
                    "nome": coverage_employee.nome,
                    "matricula": coverage_employee.matricula,
                },
            }), 200

        candidates = self._coverage_candidates(absent_employee)
        if not candidates:
            return jsonify({
                "modo": "desabilitado",
                "motivo": "Não há outro colaborador ativo neste centro de custo para cobrir o posto.",
            }), 200
        return jsonify({
            "modo": "selecionar_cobertura",
            "candidatos": candidates,
        }), 200

    @classmethod
    def _remove_operational_coverage(cls, req):
        """Remove a ausência operacional quando a cobertura original deixa de valer."""
        children = Requisicao.query.filter_by(requisicao_origem_id=req.id).all()
        for child in children:
            absence = AbsenceControl.query.filter_by(requisicao_id=child.id).first()
            if absence:
                Disallowance.query.filter_by(falta_id=absence.id).delete(synchronize_session=False)
                db.session.delete(absence)
            History.query.filter_by(requisicao_id=child.id).delete(synchronize_session=False)
            Timeline.query.filter_by(requisicao_id=child.id).delete(synchronize_session=False)
            db.session.delete(child)

    @classmethod
    def sync_operational_coverage(cls, req):
        """Gera a falta operacional do cargo menor somente após a aprovação."""
        cls._remove_operational_coverage(req)
        # Garante que um filho anterior seja removido antes de inserir outro
        # para a mesma requisição de origem.
        db.session.flush()
        if req.status != "approved" or not req.cobertura_colaborador_id:
            return None

        coverage_employee = db.session.get(Employees, req.cobertura_colaborador_id)
        return cls._ensure_operational_coverage(req, coverage_employee)

    @classmethod
    def _ensure_operational_coverage(cls, req, coverage_employee):
        """Cria uma única falta operacional quando a cobertura tem valor menor."""
        existing = Requisicao.query.filter_by(requisicao_origem_id=req.id).first()
        if existing:
            return existing

        absent_employee = db.session.get(Employees, req.ausente_id)
        if not absent_employee or not coverage_employee or not coverage_employee.centro_id:
            return None
        context = cls._additional_context(absent_employee, coverage_employee)
        if not context["coverage_has_lower_value"]:
            return None

        coverage_center = db.session.get(CostCenters, coverage_employee.centro_id)
        child = Requisicao(
            reserva_id=0,
            ausente_id=coverage_employee.id,
            cc=coverage_employee.centro_id,
            supervisor_id=None,
            supervisor_usuario_id=(
                coverage_center.supervisor_usuario_id
                if coverage_center and coverage_center.supervisor_usuario_id
                else req.supervisor_usuario_id
            ),
            warning=False,
            motivo=cls.OPERATIONAL_COVERAGE_REASON,
            obs=f"COBERTURA OPERACIONAL DA REQUISIÇÃO #{req.id}.",
            created_at=req.created_at,
            opened_at=dt.now(ZoneInfo("America/Sao_Paulo")),
            status="approved",
            requisicao_origem_id=req.id,
            origem="cobertura_operacional",
        )
        db.session.add(child)
        db.session.flush()
        AbsenceControlService.ensure_for_request(child)
        return child

    @classmethod
    def _historical_additional_rows(cls):
        """Retorna somente a última cobertura técnica aprovada por requisição."""
        latest_history = (
            db.session.query(
                History.requisicao_id,
                func.max(History.id).label("history_id"),
            )
            .group_by(History.requisicao_id)
            .subquery()
        )
        return (
            db.session.query(History, Requisicao)
            .join(latest_history, History.id == latest_history.c.history_id)
            .outerjoin(Requisicao, Requisicao.id == History.requisicao_id)
            .filter(
                History.status == "approved",
                History.reserva_id.isnot(None),
                History.reserva_id != 0,
            )
            .order_by(History.ended_at.desc(), History.id.desc())
            .all()
        )

    @classmethod
    def simulate_historical_additionals(cls, limit=50):
        """Simula vínculos históricos de adicional sem alterar qualquer registro."""
        histories = cls._historical_additional_rows()

        summary = {
            "historicos_analisados": len(histories),
            "sem_requisicao_origem": 0,
            "sem_colaborador": 0,
            "motivo_sem_falta": 0,
            "ausente_sem_adicional": 0,
            "cobertura_com_mesmo_cargo": 0,
            "cobertura_sem_valor_diario": 0,
            "candidatos_adicional": 0,
            "faltas_operacionais_possiveis": 0,
            "faltas_a_criar": 0,
        }
        candidates = []

        for history, requisition in histories:
            if not requisition:
                summary["sem_requisicao_origem"] += 1
                continue

            # O histórico é a fonte da cobertura antiga. A cobertura manual não
            # existia nesse período e, por isso, não participa da simulação.
            absent_employee = db.session.get(Employees, history.ausente_id)
            coverage_employee = db.session.get(Employees, history.reserva_id)
            if not absent_employee or not coverage_employee:
                summary["sem_colaborador"] += 1
                continue

            if not AbsenceControlService._is_absence_reason(history.motivo):
                summary["motivo_sem_falta"] += 1
                continue

            context = cls._additional_context(absent_employee, coverage_employee)
            if not context["has_additional"]:
                summary["ausente_sem_adicional"] += 1
                continue
            if not context["requires_additional"]:
                summary["cobertura_com_mesmo_cargo"] += 1
                continue

            absence = AbsenceControl.query.filter_by(
                requisicao_id=requisition.id,
            ).first()
            absent_cargo = cls._cargo_data(absent_employee)
            coverage_cargo = cls._cargo_data(coverage_employee)
            operational_coverage = Requisicao.query.filter_by(
                requisicao_origem_id=requisition.id,
            ).first()

            summary["candidatos_adicional"] += 1
            if coverage_cargo["valor_diaria"] <= 0:
                summary["cobertura_sem_valor_diario"] += 1
            if context["coverage_has_lower_value"]:
                summary["faltas_operacionais_possiveis"] += 1
            if not absence:
                summary["faltas_a_criar"] += 1

            # O limite controla apenas os detalhes impressos. Os totais sempre
            # consideram todo o histórico elegível para a conferência ser fiel.
            if len(candidates) >= limit:
                continue
            candidates.append({
                "historico_id": history.id,
                "requisicao_id": requisition.id,
                "data": history.created_at.isoformat() if history.created_at else None,
                "ausente": {
                    "id": absent_employee.id,
                    "nome": absent_employee.nome,
                    "matricula": absent_employee.matricula,
                    "cargo": absent_cargo["cargo"],
                    "adicional_tipo": absent_cargo["adicional_tipo"],
                    "adicional_valor_diaria": str(absent_cargo["adicional_valor_diaria"]),
                    "valor_diaria": str(absent_cargo["valor_diaria"]),
                },
                "cobertura_reserva": {
                    "id": coverage_employee.id,
                    "nome": coverage_employee.nome,
                    "matricula": coverage_employee.matricula,
                    "cargo": coverage_cargo["cargo"],
                    "adicional_tipo": coverage_cargo["adicional_tipo"],
                    "adicional_valor_diaria": str(coverage_cargo["adicional_valor_diaria"]),
                    "valor_diaria": str(coverage_cargo["valor_diaria"]),
                },
                "adicional_a_registrar": {
                    "tipo": context["adicional_tipo"],
                    "valor_diaria": str(context["adicional_valor_diaria"]),
                },
                "controle_faltas_existente": bool(absence),
                "falta_operacional_necessaria": context["coverage_has_lower_value"],
                "falta_operacional_existente": bool(operational_coverage),
            })

        return {"resumo": summary, "candidatos": candidates}

    @classmethod
    def apply_historical_additionals(cls):
        """Aplica adicionais históricos aprovados; a transação é controlada pelo script."""
        result = {
            "historicos_analisados": 0,
            "requisicoes_atualizadas": 0,
            "faltas_criadas": 0,
            "faltas_operacionais_criadas": 0,
            "faltas_operacionais_ignoradas_sem_valor": 0,
            "ignorados": 0,
        }

        for history, requisition in cls._historical_additional_rows():
            result["historicos_analisados"] += 1
            if not requisition:
                result["ignorados"] += 1
                continue

            absent_employee = db.session.get(Employees, history.ausente_id)
            coverage_employee = db.session.get(Employees, history.reserva_id)
            if (
                not absent_employee
                or not coverage_employee
                or not AbsenceControlService._is_absence_reason(history.motivo)
            ):
                result["ignorados"] += 1
                continue

            context = cls._additional_context(absent_employee, coverage_employee)
            if not context["requires_additional"]:
                result["ignorados"] += 1
                continue

            additional_value = context["adicional_valor_diaria"]
            current_value = Decimal(str(requisition.adicional_valor_diaria or 0))
            if (
                requisition.adicional_tipo != context["adicional_tipo"]
                or current_value != additional_value
            ):
                requisition.adicional_tipo = context["adicional_tipo"]
                requisition.adicional_valor_diaria = additional_value
                result["requisicoes_atualizadas"] += 1

            absence = AbsenceControl.query.filter_by(
                requisicao_id=requisition.id,
            ).first()
            if not absence:
                AbsenceControlService.ensure_for_request(requisition)
                result["faltas_criadas"] += 1

            coverage_cargo = cls._cargo_data(coverage_employee)
            if coverage_cargo["valor_diaria"] <= 0:
                result["faltas_operacionais_ignoradas_sem_valor"] += 1
                continue
            operational_coverage = Requisicao.query.filter_by(
                requisicao_origem_id=requisition.id,
            ).first()
            if not operational_coverage and cls._ensure_operational_coverage(
                requisition,
                coverage_employee,
            ):
                result["faltas_operacionais_criadas"] += 1

        return result

    @staticmethod
    def _disciplinary_context(employee_id):
        """Retorna totais e orientações persistentes para o colaborador selecionado."""
        summary = (
            db.session.query(
                func.sum(case((DisciplinaryMeasure.tipo == "advertencia", 1), else_=0)).label("advertencias"),
                func.sum(case((DisciplinaryMeasure.tipo == "suspensao", 1), else_=0)).label("suspensoes"),
            )
            .filter(DisciplinaryMeasure.colaborador_id == employee_id)
            .one()
        )
        counts = {
            "advertencias": int(summary.advertencias or 0),
            "suspensoes": int(summary.suspensoes or 0),
        }
        warnings = disciplinary_guidance(
            "falta_injustificada",
            counts["advertencias"],
            counts["suspensoes"],
        )
        return {"contagens": counts, "avisos": warnings}

    @safe_route
    def disciplinary_context(self, token_data):
        """Expõe totais e avisos somente no escopo da sessão autenticada."""
        if not has_permission(token_data, "reposicoes", "create"):
            return jsonify("Você não possui permissão para consultar este contexto."), 403
        try:
            body = request.get_json(silent=True) or {}
            employee_id = int(body.get("colaborador_id"))
        except (TypeError, ValueError):
            return jsonify("Colaborador inválido."), 400

        employee = db.session.get(Employees, employee_id)
        if not employee:
            return jsonify("Colaborador não encontrado."), 404
        if not can_access_cost_center(token_data, employee.centro_id):
            return jsonify("Você não possui acesso à filial deste colaborador."), 403

        return jsonify(self._disciplinary_context(employee_id)), 200

    @staticmethod
    def _parse_datetime(value):
        """Normalize browser ISO timestamps to a naive Sao Paulo database value."""
        if not value:
            return dt.now()
        if isinstance(value, dt):
            return value
        parsed = dt.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo:
            parsed = parsed.astimezone(ZoneInfo("America/Sao_Paulo")).replace(tzinfo=None)
        return parsed

    @staticmethod
    def _spreadsheet_datetime(value):
        """Accept Excel or textual dates while preserving the current submission time."""
        now = dt.now()
        if isinstance(value, dt):
            parsed = value
        elif isinstance(value, date):
            parsed = dt.combine(value, now.time())
        else:
            raw = str(value or "").strip()
            parsed = None
            for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    parsed = dt.strptime(raw, date_format)
                    break
                except ValueError:
                    continue
            if parsed is None:
                raise ValueError("data inválida; use dd/mm/aaaa")
        return parsed.replace(hour=now.hour, minute=now.minute, second=now.second, microsecond=now.microsecond)

    @safe_route
    def read(self, token_data):
        bd = request.args
        
        limit = bd.get("limit", None)
        id = bd.get("id", None)
        status = bd.get("status")

        Ausente = aliased(Employees)
        Reserva = aliased(Employees)
        SupervisorUsuario = aliased(Users)
        active_statuses = ["pending", "updated"]
        request_days = (
            db.session.query(
                Requisicao.ausente_id.label("ausente_id"),
                Requisicao.cc.label("cc"),
                Requisicao.motivo.label("motivo"),
                func.count(func.distinct(func.date(Requisicao.created_at))).label("dias"),
            )
            .filter(Requisicao.status.in_(active_statuses))
            .group_by(Requisicao.ausente_id, Requisicao.cc, Requisicao.motivo)
            .subquery()
        )

        reqs = (
            db.session.query(
                Requisicao.id,
                Requisicao.reserva_id,
                Requisicao.created_at.label("data"),
                request_days.c.dias,
                Ausente.nome.label("ausencia"),
                case(
                    (Requisicao.reserva_id == 0, "SEM COBERTURA"), else_=Reserva.nome
                ).label("reserva"),
                Floaters.id.label("reserva_floater_id"),
                CostCenters.local,
                func.coalesce(
                    SupervisorUsuario.nome,
                    Supervisors.nome,
                    "SEM SUPERVISOR",
                ).label("supervisor"),
                Requisicao.warning,
                Requisicao.origem,
                Requisicao.motivo,
                Requisicao.status,
            )
            .select_from(Requisicao)
            .join(request_days, and_(
                request_days.c.ausente_id == Requisicao.ausente_id,
                request_days.c.cc == Requisicao.cc,
                request_days.c.motivo == Requisicao.motivo,
            ))
            .join(Ausente, Ausente.id == Requisicao.ausente_id)
            .outerjoin(Reserva, Reserva.id == Requisicao.reserva_id)
            .outerjoin(Floaters, Floaters.employee_id == Reserva.id)
            .join(CostCenters, CostCenters.id == Requisicao.cc)
            .outerjoin(SupervisorUsuario, SupervisorUsuario.id == Requisicao.supervisor_usuario_id)
            .outerjoin(Supervisors, Supervisors.id == Requisicao.supervisor_id)
            .order_by(Requisicao.created_at.desc())
        )
        
        reqs = apply_cost_center_scope(reqs, Requisicao.cc, token_data)
        if id: reqs = reqs.filter(Requisicao.id == id)
        if status:
            statuses = [s.strip() for s in status.split(",") if s.strip()]
            reqs = reqs.filter(Requisicao.status.in_(statuses))
        else:
            reqs = reqs.filter(Requisicao.status.in_(["pending", "updated"]))
        if limit: reqs = reqs.limit(limit=limit)
        reqs = reqs.all()
        return jsonify([r._asdict() for r in reqs]), 200

    @safe_route
    def kds(self, token_data):
        """Return every open requisition and only decisions finalized today."""
        Ausente = aliased(Employees)
        Reserva = aliased(Employees)
        SupervisorUsuario = aliased(Users)
        latest_history = (
            db.session.query(
                History.requisicao_id,
                func.max(History.id).label("history_id"),
            )
            .group_by(History.requisicao_id)
            .subquery()
        )
        sao_paulo = ZoneInfo("America/Sao_Paulo")
        today = dt.now(sao_paulo).date()
        day_start = dt.combine(today, dt.min.time())
        day_end = dt.combine(today, dt.max.time())

        kds_query = (
            db.session.query(
                Requisicao.id,
                Requisicao.created_at.label("abertura"),
                Requisicao.opened_at.label("aberta_em"),
                Requisicao.status,
                Requisicao.motivo,
                Requisicao.obs,
                Requisicao.warning,
                Ausente.nome.label("ausente"),
                Ausente.matricula.label("ausente_matricula"),
                case(
                    (Requisicao.reserva_id == 0, "SEM COBERTURA"),
                    else_=Reserva.nome,
                ).label("reserva"),
                Reserva.matricula.label("reserva_matricula"),
                CostCenters.local.label("contrato"),
                CostCenters.departamento.label("departamento"),
                func.coalesce(
                    SupervisorUsuario.nome,
                    Supervisors.nome,
                    "SEM SUPERVISOR",
                ).label("supervisor"),
                History.ended_at.label("decidida_em"),
            )
            .select_from(Requisicao)
            .join(Ausente, Ausente.id == Requisicao.ausente_id)
            .outerjoin(Reserva, Reserva.id == Requisicao.reserva_id)
            .join(CostCenters, CostCenters.id == Requisicao.cc)
            .outerjoin(SupervisorUsuario, SupervisorUsuario.id == Requisicao.supervisor_usuario_id)
            .outerjoin(Supervisors, Supervisors.id == Requisicao.supervisor_id)
            .outerjoin(latest_history, latest_history.c.requisicao_id == Requisicao.id)
            .outerjoin(History, History.id == latest_history.c.history_id)
            .filter(or_(
                Requisicao.status.in_(["pending", "updated"]),
                and_(
                    Requisicao.status.in_(["approved", "reproved"]),
                    History.ended_at.between(day_start, day_end),
                ),
            ))
        )
        rows = apply_cost_center_scope(kds_query, Requisicao.cc, token_data).all()

        def local_iso(value):
            if not value:
                return None
            if value.tzinfo:
                return value.astimezone(sao_paulo).isoformat()
            return value.replace(tzinfo=sao_paulo).isoformat()

        return jsonify({
            "servidor_em": dt.now(sao_paulo).isoformat(),
            "requisicoes": [{
                **row._asdict(),
                "abertura": local_iso(row.abertura),
                "aberta_em": local_iso(row.aberta_em or row.abertura),
                "decidida_em": local_iso(row.decidida_em),
            } for row in rows],
        }), 200

    @safe_route
    def create(self, token_data):
        if not has_permission(token_data, "reposicoes", "create"):
            return jsonify("Você não possui permissão para criar requisições."), 403
        bd = request.get_json(silent=True) or {}

        supervisor_usuario_id = bd.get("supervisor_usuario_id")
        reserva_id = bd.get("reserva_id")
        manual_coverage_id = bd.get("cobertura_colaborador_id")
        no_coverage_requested = str(bd.get("sem_cobertura", "")).strip().lower() in {
            "1", "true", "sim", "yes", "on",
        }
        ausente_id = bd.get("ausente_id")
        advertencia = str(bd.get("advertencia"))
        motivo = bd.get("motivo")
        data = bd.get("data")
        obs = bd.get("obs")
        status = "pending"

        ok, error = check_field(
            Supervisor=supervisor_usuario_id,
            Ausente=ausente_id,
            Motivo=motivo,
        )

        if not ok:
            return jsonify(error), 400
        try:
            supervisor_usuario_id = int(supervisor_usuario_id)
        except (TypeError, ValueError):
            return jsonify("Supervisor inválido."), 400
        supervisor_user = db.session.get(Users, supervisor_usuario_id)
        if not supervisor_user or str(supervisor_user.role or "").upper() != "SUPERVISOR":
            return jsonify("Supervisor não encontrado."), 404
        adv = True if advertencia and advertencia.lower() == "aplicado" else False
        created_at = self._parse_datetime(data)
        absent_employee = db.session.get(Employees, ausente_id)
        if not absent_employee:
            return jsonify("Colaborador ausente não encontrado."), 404
        duplicate_message = AbsenceControlService.duplicate_request_message(
            absent_employee.id,
            created_at,
        )
        if duplicate_message:
            return jsonify(duplicate_message), 409
        requested_center_id = bd.get("centro_id")
        if requested_center_id is not None:
            try:
                centro_id = int(requested_center_id)
            except (TypeError, ValueError):
                return jsonify("Local selecionado inválido."), 400
            if not db.session.get(CostCenters, centro_id):
                return jsonify("Local selecionado não encontrado."), 404
        else:
            centro_id = absent_employee.centro_id
        if not centro_id:
            return jsonify("O colaborador ausente não possui um local cadastrado."), 400
        if not can_access_cost_center(token_data, centro_id):
            return jsonify("Você não possui acesso à filial deste colaborador."), 403
        if not can_access_supervisor_user(token_data, supervisor_usuario_id, centro_id):
            return jsonify("Você não possui acesso à filial deste supervisor."), 403
        if not is_admin(token_data):
            if token_data.get("id") != supervisor_usuario_id:
                return jsonify("A requisição deve ser registrada pelo supervisor autenticado."), 403

        try:
            reserva_id = int(reserva_id or 0)
        except (TypeError, ValueError):
            return jsonify("Reserva inválida."), 400

        if reserva_id:
            reservation = Floaters.query.filter_by(employee_id=reserva_id).first()
            if not reservation:
                return jsonify("A pessoa selecionada não pertence às reservas técnicas."), 400
            if not reservation.disponivel:
                reason = (reservation.indisponibilidade_motivo or "indisponível").lower()
                return jsonify(f"Esta reserva está indisponível por {reason}."), 409
            if not _can_access_employee(token_data, reserva_id):
                return jsonify("Você não possui acesso à filial desta reserva."), 403
            day_start = created_at.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = created_at.replace(hour=23, minute=59, second=59, microsecond=999999)
            conflicting_request = Requisicao.query.filter(
                Requisicao.reserva_id == reserva_id,
                Requisicao.created_at.between(day_start, day_end),
                Requisicao.status.in_(["pending", "updated", "approved"]),
            ).first()
            if conflicting_request:
                return jsonify("Esta reserva está indisponível na data informada."), 409

        coverage_employee = db.session.get(Employees, reserva_id) if reserva_id else None
        if manual_coverage_id not in (None, "", 0, "0"):
            if reserva_id:
                return jsonify("Informe a cobertura manual somente quando não houver reserva técnica."), 400
            try:
                manual_coverage_id = int(manual_coverage_id)
            except (TypeError, ValueError):
                return jsonify("Cobertura manual inválida."), 400
            coverage_employee = db.session.get(Employees, manual_coverage_id)
            if not coverage_employee or coverage_employee.id == absent_employee.id:
                return jsonify("Selecione outro colaborador para realizar a cobertura."), 400
            if (
                coverage_employee.centro_id != absent_employee.centro_id
                or coverage_employee.situacao != 1
            ):
                return jsonify(
                    "A cobertura manual deve ser um colaborador ativo do mesmo centro de custo."
                ), 400
            if not _can_access_employee(token_data, coverage_employee.id):
                return jsonify("Você não possui acesso à filial da cobertura informada."), 403
        else:
            manual_coverage_id = None

        additional = self._additional_context(absent_employee, coverage_employee)
        if manual_coverage_id and not additional["has_additional"]:
            return jsonify("A cobertura manual do item 3 só é permitida para cargo com adicional cadastrado."), 400
        if (
            not reserva_id
            and additional["has_additional"]
            and not manual_coverage_id
            and not no_coverage_requested
            and self._coverage_candidates(absent_employee)
        ):
            return jsonify("Selecione quem realizará a cobertura no item 3."), 400

        new_rq = Requisicao(
            reserva_id=reserva_id,
            cobertura_colaborador_id=coverage_employee.id if coverage_employee else None,
            ausente_id=ausente_id,
            cc=centro_id,
            supervisor_id=None,
            supervisor_usuario_id=supervisor_usuario_id,
            warning=adv,
            adicional_tipo=additional["adicional_tipo"],
            adicional_valor_diaria=additional["adicional_valor_diaria"] or None,
            origem="requisicao",
            motivo=motivo,
            created_at=created_at,
            opened_at=dt.now(ZoneInfo("America/Sao_Paulo")),
            status=status
        )

        if obs: new_rq.obs = str(obs).strip().upper()
        db.session.add(new_rq)
        db.session.flush()
        if motivo not in self.ISNOTFAULT: AbsenceControlService.ensure_for_request(new_rq)
        db.session.commit()

        disciplinary_context = self._disciplinary_context(
            absent_employee.id,
        )

        TimelineService().create_event(
            req=new_rq,
            status=status,
            tipo="Criação da requisição",
            obs=obs,
            criado_por_usuario_id=token_data.get("id"),
        )
        
        socketio.emit("new_request")
        _emit_kds_update("created", new_rq.id, new_rq.status)
        return jsonify({
            "message": "Requisição criada",
            "resumo_disciplinar": disciplinary_context["contagens"],
            "avisos": disciplinary_context["avisos"],
        }), 201

    @safe_route
    def update(self, token_data):
        bd = request.get_json()
        id = bd.get("id")

        req = Requisicao.query.filter(Requisicao.id == id).first()
        if not req: return jsonify("Requisição não encontrada"), 404

        target_center = bd.get("centro_id", req.cc)
        if not can_access_cost_center(token_data, req.cc) or not can_access_cost_center(token_data, target_center):
            return jsonify("Você não possui acesso à filial desta requisição."), 403
        if "ausente_id" in bd and not _can_access_employee(token_data, bd.get("ausente_id")):
            return jsonify("Você não possui acesso à filial do colaborador ausente."), 403
        if "reserva_id" in bd and not _can_access_employee(
            token_data, bd.get("reserva_id"), allow_uncovered=True
        ):
            return jsonify("Você não possui acesso à filial desta reserva."), 403

        if "reserva_id" in bd and bd.get("reserva_id") not in (None, 0):
            reservation = Floaters.query.filter_by(employee_id=bd.get("reserva_id")).first()
            if not reservation:
                return jsonify("A pessoa selecionada não pertence às reservas técnicas."), 400
            if not reservation.disponivel:
                reason = (reservation.indisponibilidade_motivo or "indisponível").lower()
                return jsonify(f"Esta reserva está indisponível por {reason}."), 409

        next_absent_id = bd.get("ausente_id", req.ausente_id)
        next_date = (
            self._parse_datetime(bd.get("data"))
            if "data" in bd
            else req.created_at
        )
        duplicate_message = AbsenceControlService.duplicate_request_message(
            next_absent_id,
            next_date,
            exclude_request_id=req.id,
        )
        if duplicate_message:
            return jsonify(duplicate_message), 409

        if "reserva_id" in bd: req.reserva_id = bd.get("reserva_id")
        if "centro_id" in bd: req.cc = bd.get("centro_id")
        if "ausente_id" in bd: req.ausente_id = bd.get("ausente_id")
        if "motivo" in bd: req.motivo = bd.get("motivo")
        if "data" in bd: req.created_at = self._parse_datetime(bd.get("data"))
        req.status = "updated"
        AbsenceControlService.ensure_for_request(req)
        db.session.commit()

        TimelineService().create_event(
            req=req,
            status="updated",
            tipo="Alteração de Dados",
            obs=bd.get("obs", req.obs),
            alterado_por_usuario_id=token_data.get("id")
        )

        socketio.emit("new_request")
        _emit_kds_update("updated", req.id, req.status)
        return jsonify("Requisição alterada"), 200

    @safe_route
    def export(self, token_data):
        """Export only the operational queue, keeping approved/reproved items in history."""
        Ausente = aliased(Employees)
        Reserva = aliased(Employees)
        SupervisorUsuario = aliased(Users)
        request_days = (
            db.session.query(
                Requisicao.ausente_id.label("ausente_id"),
                Requisicao.cc.label("cc"),
                Requisicao.motivo.label("motivo"),
                func.count(func.distinct(func.date(Requisicao.created_at))).label("dias"),
            )
            .filter(Requisicao.status.in_(["pending", "updated"]))
            .group_by(Requisicao.ausente_id, Requisicao.cc, Requisicao.motivo)
            .subquery()
        )
        export_query = (
            db.session.query(
                Requisicao.id,
                Requisicao.created_at.label("data"),
                request_days.c.dias,
                Requisicao.status,
                Requisicao.motivo,
                Ausente.nome.label("ausente"),
                case((Requisicao.reserva_id == 0, "SEM COBERTURA"), else_=Reserva.nome).label("reserva"),
                CostCenters.local,
                CostCenters.departamento,
                func.coalesce(
                    SupervisorUsuario.nome,
                    Supervisors.nome,
                    "SEM SUPERVISOR",
                ).label("supervisor"),
            )
            .select_from(Requisicao)
            .join(request_days, and_(
                request_days.c.ausente_id == Requisicao.ausente_id,
                request_days.c.cc == Requisicao.cc,
                request_days.c.motivo == Requisicao.motivo,
            ))
            .join(Ausente, Ausente.id == Requisicao.ausente_id)
            .outerjoin(Reserva, Reserva.id == Requisicao.reserva_id)
            .join(CostCenters, CostCenters.id == Requisicao.cc)
            .outerjoin(SupervisorUsuario, SupervisorUsuario.id == Requisicao.supervisor_usuario_id)
            .outerjoin(Supervisors, Supervisors.id == Requisicao.supervisor_id)
            .filter(Requisicao.status.in_(["pending", "updated"]))
            .order_by(Requisicao.created_at.desc())
        )
        rows = apply_cost_center_scope(export_query, Requisicao.cc, token_data).all()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Requisicoes"
        headers = ["id", "data", "dias", "status", "motivo", "ausente", "reserva", "local", "departamento", "supervisor"]
        worksheet.append(headers)
        for row in rows:
            values = row._asdict()
            worksheet.append([values.get(header) for header in headers])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:J{max(len(rows) + 1, 1)}"
        for column, width in {"A": 8, "B": 20, "C": 8, "D": 14, "E": 22, "F": 32, "G": 32, "H": 40, "I": 16, "J": 28}.items():
            worksheet.column_dimensions[column].width = width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="requisicoes_abertas.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @safe_route
    def download_import_template(self, token_data):
        """Generate the canonical import sheet plus read-only ID reference tabs."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Requisicoes"
        headers = ["supervisor_usuario_id", "reserva_id", "centro_id", "ausente_id", "motivo", "data", "advertencia", "obs"]
        worksheet.append(headers)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = "A1:H1"
        for column, width in {"A": 16, "B": 14, "C": 14, "D": 14, "E": 22, "F": 16, "G": 18, "H": 36}.items():
            worksheet.column_dimensions[column].width = width

        instructions = workbook.create_sheet("Instrucoes")
        instructions.append(["Campo", "Regra"])
        instructions.append(["supervisor_usuario_id", "Obrigatório; usuário com role SUPERVISOR; consulte a aba Supervisores"])
        instructions.append(["reserva_id", "Obrigatório; consulte a aba Reservas ou use 0 para SEM COBERTURA"])
        instructions.append(["centro_id", "Obrigatório; consulte a aba Centros"])
        instructions.append(["ausente_id", "Obrigatório; consulte a aba Colaboradores"])
        instructions.append(["motivo", "AFASTAMENTO, ATESTADO, DECLARAÇÃO, POSTO VAGO, REMANEJAMENTO, INJUSTIFICADA ou OUTROS"])
        instructions.append(["data", "Obrigatório; somente hoje ou amanhã, no formato dd/mm/aaaa"])
        instructions.append(["advertencia", "Opcional; use APLICADO ou NÃO APLICADO"])
        instructions.append(["obs", "Opcional"])

        allowed_centers = apply_cost_center_scope(CostCenters.query, CostCenters.id, token_data).all()
        allowed_center_ids = {center.id for center in allowed_centers}
        allowed_supervisor_ids = {
            supervisor.id
            for supervisor in supervisor_users_query(token_data).all()
        }
        reference_sheets = [
            ("Supervisores", ["id", "nome"], db.session.query(Users.id, Users.nome).filter(Users.id.in_(allowed_supervisor_ids), func.upper(func.trim(Users.role)) == "SUPERVISOR").order_by(Users.nome).all()),
            ("Reservas", ["id", "matricula", "nome"], db.session.query(Employees.id, Employees.matricula, Employees.nome).select_from(Floaters).join(Employees, Employees.id == Floaters.employee_id).filter(Employees.centro_id.in_(allowed_center_ids)).order_by(Employees.nome).all()),
            ("Colaboradores", ["id", "matricula", "nome"], db.session.query(Employees.id, Employees.matricula, Employees.nome).filter(Employees.centro_id.in_(allowed_center_ids)).order_by(Employees.nome).all()),
            ("Centros", ["id", "local", "departamento"], [(center.id, center.local, center.departamento) for center in allowed_centers]),
        ]
        for title, sheet_headers, rows in reference_sheets:
            sheet = workbook.create_sheet(title)
            sheet.append(sheet_headers)
            for row in rows:
                sheet.append(list(row))
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="modelo_importacao_requisicoes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @safe_route
    def import_requests(self, token_data):
        """Validate every spreadsheet row before committing requests and timelines atomically."""
        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify("Envie uma planilha no formato .xlsx."), 400

        try:
            workbook = load_workbook(uploaded.stream, read_only=True, data_only=True)
            worksheet = workbook["Requisicoes"] if "Requisicoes" in workbook.sheetnames else workbook.active
            rows = worksheet.iter_rows(values_only=True)
            headers = [str(value or "").strip().lower() for value in next(rows)]
        except (StopIteration, ValueError, OSError):
            return jsonify("Não foi possível ler a planilha."), 400

        required = ["supervisor_usuario_id", "reserva_id", "centro_id", "ausente_id", "motivo", "data"]
        if any(field not in headers for field in required):
            return jsonify({"message": "Planilha fora do padrão.", "errors": [f"Colunas obrigatórias: {', '.join(required)}."]}), 400

        indexes = {header: position for position, header in enumerate(headers)}
        # Armazena chaves estrangeiras válidas para evitar uma consulta por linha da planilha.
        center_ids = {row[0] for row in apply_cost_center_scope(db.session.query(CostCenters.id), CostCenters.id, token_data).all()}
        supervisor_ids = {
            row[0] for row in supervisor_users_query(token_data).with_entities(Users.id).all()
        }
        employee_ids = {
            row[0] for row in db.session.query(Employees.id)
            .filter(Employees.centro_id.in_(center_ids)).all()
        }
        reservation_ids = {
            row[0] for row in db.session.query(Floaters.employee_id)
            .join(Employees, Employees.id == Floaters.employee_id)
            .filter(Employees.centro_id.in_(center_ids)).all()
        }
        today = dt.now().date()
        allowed_dates = {today, today + timedelta(days=1)}
        created = []
        errors = []
        batch_request_keys = set()

        for row_number, row in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            if row_number > 1001:
                errors.append("A planilha pode conter no máximo 1000 requisições.")
                break

            def value(field, default=None):
                position = indexes.get(field)
                return row[position] if position is not None and position < len(row) else default

            try:
                supervisor_usuario_id = int(value("supervisor_usuario_id"))
                reserva_id = int(value("reserva_id"))
                centro_id = int(value("centro_id"))
                ausente_id = int(value("ausente_id"))
                motivo = str(value("motivo") or "").strip().upper()
                created_at = self._spreadsheet_datetime(value("data"))
            except (TypeError, ValueError) as error:
                errors.append(f"Linha {row_number}: {error}.")
                continue

            row_errors = []
            if supervisor_usuario_id not in supervisor_ids: row_errors.append("supervisor_usuario_id não encontrado")
            elif not can_access_supervisor_user(token_data, supervisor_usuario_id, centro_id): row_errors.append("supervisor_usuario_id não está disponível para este centro")
            if reserva_id != 0 and reserva_id not in reservation_ids: row_errors.append("reserva_id não pertence às reservas técnicas")
            if centro_id not in center_ids: row_errors.append("centro_id não encontrado")
            if ausente_id not in employee_ids: row_errors.append("ausente_id não encontrado")
            if motivo not in self.REASONS: row_errors.append("motivo inválido")
            if created_at.date() not in allowed_dates: row_errors.append("a data deve ser hoje ou amanhã")
            request_key = (ausente_id, created_at.date())
            if request_key in batch_request_keys:
                row_errors.append("já existe outra requisição desta planilha para o colaborador nesta data")
            elif AbsenceControlService.duplicate_request_for_day(ausente_id, created_at):
                row_errors.append("já existe uma requisição para o colaborador nesta data")
            if row_errors:
                errors.append(f"Linha {row_number}: {', '.join(row_errors)}.")
                continue

            batch_request_keys.add(request_key)

            warning_value = str(value("advertencia") or "").strip().upper()
            obs = str(value("obs") or "").strip().upper() or None
            requisition = Requisicao(
                reserva_id=reserva_id,
                ausente_id=ausente_id,
                cc=centro_id,
                supervisor_id=None,
                supervisor_usuario_id=supervisor_usuario_id,
                warning=warning_value == "APLICADO",
                origem="requisicao",
                motivo=motivo,
                obs=obs,
                created_at=created_at,
                opened_at=dt.now(ZoneInfo("America/Sao_Paulo")),
                status="pending",
            )
            db.session.add(requisition)
            created.append(requisition)

        # Qualquer linha inválida cancela o lote completo; filas operacionais parciais são inseguras.
        if errors:
            db.session.rollback()
            return jsonify({"message": "A importação foi cancelada; nenhuma requisição foi criada.", "errors": errors}), 400
        if not created:
            return jsonify("A planilha não contém requisições para importar."), 400

        # Gera os IDs antes para que a requisição e o evento inicial da timeline usem a mesma transação.
        db.session.flush()
        for requisition in created:
            AbsenceControlService.ensure_for_request(requisition)
            db.session.add(Timeline(
                requisicao_id=requisition.id,
                reserva_id=requisition.reserva_id,
                ausente_id=requisition.ausente_id,
                cc=requisition.cc,
                supervisor_id=None,
                supervisor_usuario_id=requisition.supervisor_usuario_id,
                criado_por_usuario_id=token_data.get("id"),
                status="pending",
                tipo="Criação da requisição por planilha",
                motivo=requisition.motivo,
                obs=requisition.obs,
            ))
        db.session.commit()
        socketio.emit("new_request")
        _emit_kds_update("imported")
        return jsonify({"message": f"{len(created)} requisições importadas com sucesso.", "total": len(created)}), 201

    @safe_route
    def daily_reservations(self, token_data):
        """Split technical reserves by availability for one requested calendar day."""
        value = request.args.get("data")
        try:
            day = dt.strptime(value, "%Y-%m-%d") if value else dt.now()
        except ValueError:
            return jsonify("Data inválida."), 400
        init = day.replace(hour=0, minute=0, second=0, microsecond=0)
        end = day.replace(hour=23, minute=59, second=59, microsecond=999999)

        # Durações antigas não bloqueiam datas posteriores; somente o dia de abertura conta.
        used_ids = {
            row[0] for row in db.session.query(Requisicao.reserva_id)
            .filter(
                Requisicao.created_at.between(init, end),
                Requisicao.reserva_id > 0,
                Requisicao.status.in_(["pending", "updated", "approved"]),
            ).distinct().all()
        }
        # Último contrato considera todo uso registrado até o fim do dia consultado.
        # A janela mantém somente a requisição mais recente de cada reserva.
        last_usage = (
            db.session.query(
                Requisicao.reserva_id.label("reserva_id"),
                CostCenters.local.label("ultimo_contrato"),
                func.row_number().over(
                    partition_by=Requisicao.reserva_id,
                    order_by=Requisicao.created_at.desc(),
                ).label("ordem"),
            )
            .join(CostCenters, CostCenters.id == Requisicao.cc)
            .filter(
                Requisicao.created_at <= end,
                Requisicao.reserva_id > 0,
                Requisicao.status.in_(["pending", "updated", "approved"]),
            )
            .subquery()
        )

        reservation_query = (
            db.session.query(
                Employees.id,
                Employees.nome,
                Employees.matricula,
                Cargos.nome.label("cargo"),
                Situations.tipo.label("situacao"),
                last_usage.c.ultimo_contrato,
                Floaters.disponivel,
                Floaters.indisponibilidade_motivo,
            )
            .select_from(Floaters)
            .join(Employees, Employees.id == Floaters.employee_id)
            .join(CostCenters, CostCenters.id == Employees.centro_id)
            .join(Cargos, Cargos.id == Employees.cargo)
            .join(Situations, Situations.id == Employees.situacao)
            .outerjoin(last_usage, and_(
                last_usage.c.reserva_id == Employees.id,
                last_usage.c.ordem == 1,
            ))
            .order_by(Employees.nome)
        )
        reservation_query = apply_active_department_scope(
            reservation_query, Employees.centro_id
        )
        if not has_permission(token_data, "reposicoes", "create"):
            return jsonify("Você não possui permissão para consultar reservas."), 403
        reservation_query = apply_cost_center_scope(
            reservation_query, Employees.centro_id, token_data
        )
        reservations = reservation_query.all()
        response = [{**row._asdict(), "usada": row.id in used_ids} for row in reservations]
        return jsonify({
            "data": init.strftime("%Y-%m-%d"),
            "usadas": [row for row in response if row["usada"] and row["disponivel"]],
            "disponiveis": [row for row in response if not row["usada"] and row["disponivel"]],
            "indisponiveis": [row for row in response if not row["disponivel"]],
        }), 200
        
    @safe_route
    def delete(self, token_data):
        bd = request.get_json(silent=True) or request.args
        id = bd.get("id")

        req = Requisicao.query.filter(Requisicao.id == id).first()
        if not req: return jsonify("RequisiÃ§Ã£o nÃ£o encontrada"), 404

        if not can_access_cost_center(token_data, req.cc):
            return jsonify("Você não possui acesso à filial desta requisição."), 403
        requisicao_id = req.id
        self._remove_operational_coverage(req)
        History.query.filter(History.requisicao_id == requisicao_id).delete(synchronize_session=False)
        Timeline.query.filter(Timeline.requisicao_id == requisicao_id).delete(synchronize_session=False)
        db.session.delete(req)
        db.session.commit()

        socketio.emit("new_history")
        socketio.emit("new_request")
        socketio.emit("disallowance_update", {"action": "request_deleted"})
        _emit_kds_update("deleted", requisicao_id)
        return jsonify({
            "message": "RequisiÃ§Ã£o excluÃ­da",
            "requisicao_id": requisicao_id
        }), 200

class HistoryService:
    @safe_route
    def read(self, token_data):
        bd = request.get_json()

        init = bd.get("init", None)
        end = bd.get("end", None)
        
        if init and end: # Se passar os dois
            init = dt.now().strptime(init, "%d/%m/%Y").replace(hour=0, minute=0, second=0); 
            end = dt.now().strptime(end, "%d/%m/%Y").replace(hour=23, minute=59, second=59)
        elif init: # Se for passado somente o init
            init = dt.now().strptime(init, "%d/%m/%Y").replace(hour=0, minute=0, second=0); 
            end = init.replace(hour=23, minute=59, second=59)
        else: # Se nao for passado nenhum
            init = dt.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0); 
            dias_no_mes = monthrange(init.year, init.month)[1]
            end = init + relativedelta(day=dias_no_mes , hour=23, minute=59, second=59)
        
        Ausente = aliased(Employees)
        Reserva = aliased(Employees)
        latest_history = (
            db.session.query(
                History.requisicao_id,
                func.max(History.id).label("id")
            )
            .group_by(History.requisicao_id)
            .subquery()
        )
        absence_days = (
            db.session.query(
                History.ausente_id.label("ausente_id"),
                History.cc.label("cc"),
                History.motivo.label("motivo"),
                func.count(func.distinct(func.date(History.created_at))).label("dias"),
            )
            .join(latest_history, History.id == latest_history.c.id)
            .filter(History.created_at.between(init, end))
            .group_by(History.ausente_id, History.cc, History.motivo)
            .subquery()
        )

        SupervisorUsuario = aliased(Users)
        history_query = (
            db.session.query(
                History.id,
                History.requisicao_id,
                History.created_at.label("abertura"),
                History.ended_at.label("fechamento"),
                Ausente.nome.label("ausente"),
                case(
                    (History.reserva_id == 0, "SEM COBERTURA"), else_=Reserva.nome
                ).label("reserva"),
                History.motivo,
                History.obs,
                absence_days.c.dias,
                func.coalesce(
                    SupervisorUsuario.nome,
                    Supervisors.nome,
                    "SEM SUPERVISOR",
                ).label("supervisor"),
                CostCenters.local.label("local"),
                CostCenters.departamento.label("dpto"),
                History.status,
                Cargos.multa,
                
            )
            .select_from(History)
            .join(latest_history, History.id == latest_history.c.id)
            .join(absence_days, and_(
                absence_days.c.ausente_id == History.ausente_id,
                absence_days.c.cc == History.cc,
                absence_days.c.motivo == History.motivo,
            ))
            .join(Ausente, Ausente.id == History.ausente_id)
            .outerjoin(Reserva, Reserva.id == History.reserva_id)
            .outerjoin(Cargos, Cargos.id == Reserva.cargo)
            .join(CostCenters, CostCenters.id == History.cc)
            .outerjoin(SupervisorUsuario, SupervisorUsuario.id == History.supervisor_usuario_id)
            .outerjoin(Supervisors, Supervisors.id == History.supervisor_id)
            .filter(History.created_at.between(init, end))
            .order_by(History.created_at.desc())
        )
        hists = apply_cost_center_scope(history_query, History.cc, token_data).all()
        return jsonify([h._asdict() for h in hists]), 200
        
    @safe_route
    def create(self, token_data):
        bd = request.get_json()
        id = bd.get("id")
        status = bd.get("status", "reproved")
        req = Requisicao.query.filter(Requisicao.id == id).first()

        if not req:
            return jsonify("Requisição não encontrada."), 404
        if not can_access_cost_center(token_data, req.cc):
            return jsonify("Você não possui acesso à filial desta requisição."), 403
        requisicao_id = req.id
        reserva_id = req.reserva_id if status == "approved" else 0
        ausente_id = req.ausente_id
        cc_id = req.cc
        created_at = req.created_at
        supervisor_id = req.supervisor_id
        supervisor_usuario_id = req.supervisor_usuario_id
        motivo = req.motivo
        ended_at = dt.now()
        obs = req.obs

        hist = History.query.filter(History.requisicao_id == requisicao_id).order_by(History.id.desc()).first()
        if not hist:
            hist = History(
                requisicao_id=requisicao_id,
                created_at=created_at,
            )
            db.session.add(hist)

        hist.reserva_id = reserva_id
        hist.ausente_id = ausente_id
        hist.cc = cc_id
        hist.status = status
        hist.ended_at = ended_at
        hist.supervisor_id = supervisor_id
        hist.supervisor_usuario_id = supervisor_usuario_id
        hist.motivo = motivo
        hist.obs = obs
        
        req.status = status
        req.reserva_id = reserva_id
        RequestService.sync_operational_coverage(req)
        AbsenceControlService.ensure_for_request(req)
        db.session.commit()
        
        TimelineService().create_event(
            req= req,
            status= status,
            tipo = "Aprovado" if status == "approved" else "Reprovado, posto sem cobertura.",
            obs= obs,
            alterado_por_usuario_id=token_data.get("id")
        )

        socketio.emit("new_history")
        socketio.emit("disallowance_update", {"action": "request_decided"})
        _emit_kds_update("decided", requisicao_id, status)
        return jsonify("Sucesso"), 201

    @safe_route
    def update(self, token_data):
        bd = request.get_json()
        id = bd.get("id")

        hist = History.query.filter(History.id == id).first()
        if not hist: return jsonify("Histórico não encontrado"), 404

        target_center = bd.get("centro_id", hist.cc)
        if not can_access_cost_center(token_data, hist.cc) or not can_access_cost_center(token_data, target_center):
            return jsonify("Você não possui acesso à filial desta requisição."), 403
        if "ausente_id" in bd and not _can_access_employee(token_data, bd.get("ausente_id")):
            return jsonify("Você não possui acesso à filial do colaborador ausente."), 403
        if "reserva_id" in bd and not _can_access_employee(
            token_data, bd.get("reserva_id"), allow_uncovered=True
        ):
            return jsonify("Você não possui acesso à filial desta reserva."), 403
        if (
            "supervisor_usuario_id" in bd
            and not can_access_supervisor_user(
                token_data,
                bd.get("supervisor_usuario_id"),
                target_center,
            )
        ):
            return jsonify("Você não possui acesso à filial deste supervisor."), 403

        req = Requisicao.query.filter(Requisicao.id == hist.requisicao_id).first()
        if not req:
            req = Requisicao(
                id=hist.requisicao_id,
                reserva_id=hist.reserva_id,
                ausente_id=hist.ausente_id,
                cc=hist.cc,
                supervisor_id=hist.supervisor_id,
                supervisor_usuario_id=hist.supervisor_usuario_id,
                warning=False,
                origem="requisicao",
                motivo=hist.motivo,
                obs=hist.obs,
                created_at=hist.created_at,
                opened_at=dt.now(ZoneInfo("America/Sao_Paulo")),
                status="updated"
            )
            db.session.add(req)

        next_absent_id = bd.get("ausente_id", req.ausente_id)
        duplicate_message = AbsenceControlService.duplicate_request_message(
            next_absent_id,
            req.created_at,
            exclude_request_id=req.id,
        )
        if duplicate_message:
            return jsonify(duplicate_message), 409

        if "reserva_id" in bd:
            hist.reserva_id = bd.get("reserva_id")
            req.reserva_id = bd.get("reserva_id")
        if "centro_id" in bd:
            hist.cc = bd.get("centro_id")
            req.cc = bd.get("centro_id")
        if "ausente_id" in bd:
            hist.ausente_id = bd.get("ausente_id")
            req.ausente_id = bd.get("ausente_id")
        if "supervisor_usuario_id" in bd:
            try:
                supervisor_usuario_id = int(bd.get("supervisor_usuario_id"))
            except (TypeError, ValueError):
                return jsonify("Supervisor inválido."), 400
            supervisor = db.session.get(Users, supervisor_usuario_id)
            if not supervisor or str(supervisor.role or "").upper() != "SUPERVISOR":
                return jsonify("Supervisor não encontrado."), 404
            hist.supervisor_usuario_id = supervisor_usuario_id
            req.supervisor_usuario_id = supervisor_usuario_id
        if "motivo" in bd:
            hist.motivo = bd.get("motivo")
            req.motivo = bd.get("motivo")
        if "obs" in bd:
            hist.obs = str(bd.get("obs")).strip().upper()
            req.obs = hist.obs

        hist.status = "pending"
        req.status = "updated"

        RequestService.sync_operational_coverage(req)
        AbsenceControlService.ensure_for_request(req)

        db.session.commit()

        TimelineService().create_event(
            req=req,
            status="updated",
            tipo="Alteração do histórico",
            obs=bd.get("obs", req.obs),
            alterado_por_usuario_id=token_data.get("id")
        )

        socketio.emit("new_history")
        socketio.emit("new_request")
        socketio.emit("disallowance_update", {"action": "request_reopened"})
        _emit_kds_update("reopened", req.id, req.status)
        return jsonify("Histórico alterado"), 200

    @safe_route
    def delete(self, token_data):
        bd = request.get_json(silent=True) or request.args
        id = bd.get("id")

        hist = History.query.filter(History.id == id).first()
        if not hist: return jsonify("HistÃ³rico nÃ£o encontrado"), 404

        if not can_access_cost_center(token_data, hist.cc):
            return jsonify("Você não possui acesso à filial desta requisição."), 403
        requisicao_id = hist.requisicao_id
        req = Requisicao.query.filter(Requisicao.id == requisicao_id).first()

        if req:
            RequestService._remove_operational_coverage(req)
        History.query.filter(History.requisicao_id == requisicao_id).delete(synchronize_session=False)
        Timeline.query.filter(Timeline.requisicao_id == requisicao_id).delete(synchronize_session=False)
        if req: db.session.delete(req)
        db.session.commit()

        socketio.emit("new_history")
        socketio.emit("new_request")
        socketio.emit("disallowance_update", {"action": "history_deleted"})
        _emit_kds_update("deleted", requisicao_id)
        return jsonify({
            "message": "HistÃ³rico e requisiÃ§Ã£o excluÃ­dos",
            "history_id": id,
            "requisicao_id": requisicao_id
        }), 200

class TimelineService:
    def create_event(
        self,
        req,
        status,
        tipo,
        obs=None,
        criado_por_supervisor_id=None,
        criado_por_usuario_id=None,
        alterado_por_usuario_id=None,
    ):
        db.session.add(
            Timeline(
                requisicao_id=req.id,
                reserva_id=req.reserva_id,
                ausente_id=req.ausente_id,
                cc=req.cc,
                supervisor_id=req.supervisor_id,
                supervisor_usuario_id=req.supervisor_usuario_id,
                criado_por_supervisor_id=criado_por_supervisor_id,
                criado_por_usuario_id=criado_por_usuario_id,
                alterado_por_usuario_id=alterado_por_usuario_id,
                status=status,
                tipo=tipo,
                motivo=req.motivo,
                obs=obs or req.obs
            )
        )
        db.session.commit()

    @safe_route
    def read(self, token_data):
        requisicao_id = request.args.get("requisicao_id")

        Ausente = aliased(Employees)
        Reserva = aliased(Employees)
        SupervisorUsuario = aliased(Users)
        Criador = aliased(Supervisors)
        CriadorUsuario = aliased(Users)
        Alterador = aliased(Users)

        query = (
            db.session.query(
                Timeline.id,
                Timeline.requisicao_id,
                Timeline.created_at,
                Timeline.status,
                Timeline.tipo,
                Timeline.supervisor_id,
                Timeline.supervisor_usuario_id,
                Timeline.criado_por_supervisor_id,
                Timeline.criado_por_usuario_id,
                Timeline.alterado_por_usuario_id,
                Ausente.nome.label("ausente"),
                case(
                    (Timeline.reserva_id == 0, "SEM COBERTURA"),
                    else_=Reserva.nome
                ).label("reserva"),
                CostCenters.local,
                func.coalesce(
                    SupervisorUsuario.nome,
                    Supervisors.nome,
                    "SEM SUPERVISOR",
                ).label("supervisor"),
                Criador.nome.label("criado_por"),
                CriadorUsuario.nome.label("criado_por_usuario"),
                CriadorUsuario.foto_perfil.label("criado_por_usuario_foto"),
                Alterador.nome.label("alterado_por"),
                Alterador.foto_perfil.label("alterado_por_foto"),
                Timeline.motivo,
                Timeline.obs,
            )
            .select_from(Timeline)
            .join(Ausente, Ausente.id == Timeline.ausente_id)
            .outerjoin(Reserva, Reserva.id == Timeline.reserva_id)
            .join(CostCenters, CostCenters.id == Timeline.cc)
            .outerjoin(SupervisorUsuario, SupervisorUsuario.id == Timeline.supervisor_usuario_id)
            .outerjoin(Supervisors, Supervisors.id == Timeline.supervisor_id)
            .outerjoin(Criador, Criador.id == Timeline.criado_por_supervisor_id)
            .outerjoin(CriadorUsuario, CriadorUsuario.id == Timeline.criado_por_usuario_id)
            .outerjoin(Alterador, Alterador.id == Timeline.alterado_por_usuario_id)
            .order_by(Timeline.created_at.desc())
        )

        query = apply_cost_center_scope(query, Timeline.cc, token_data)
        if requisicao_id: query = query.filter(Timeline.requisicao_id == requisicao_id)
        timelines = query.all()
        return jsonify([t._asdict() for t in timelines]), 200

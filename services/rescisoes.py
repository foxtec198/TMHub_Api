from calendar import monthrange
from datetime import date, datetime as dt
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from unicodedata import normalize

from flask import jsonify, request
from openpyxl import load_workbook
from sqlalchemy import String, cast, or_

from models.centros_de_custo import CostCenters
from models.colaboradores import Employees
from models.filiais import Branch, filial_centros_custo, filial_departamentos
from models.rescisoes import Termination
from models.supervisores import Supervisors
from models.usuarios import Users
from utils.db import db
from utils.filial_scope import (
    allowed_cost_center_ids,
    apply_cost_center_scope,
    can_access_cost_center,
    is_admin,
)
from utils.permissions import has_permission
from utils.safe_route import safe_route
from utils.socket import socketio


MONEY = Decimal("0.01")
VALID_REASONS = {
    "sem_justa_causa": "Dispensa sem justa causa",
    "pedido_demissao": "Pedido de demissao",
    "acordo": "Extincao por acordo",
    "justa_causa": "Dispensa por justa causa",
    "termino_contrato": "Termino de contrato por prazo determinado",
    "morte": "Falecimento do colaborador",
}
VALID_NOTICE = {"indenizado", "trabalhado", "descontado", "dispensado", "nao_aplicavel"}


def _money(value, field="Valor"):
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(MONEY, rounding=ROUND_HALF_UP)
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)
    raw = str(value).strip().replace("R$", "").replace(" ", "")
    if not raw:
        return Decimal("0.00")
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        return Decimal(raw).quantize(MONEY, rounding=ROUND_HALF_UP)
    except InvalidOperation as error:
        raise ValueError(f"{field} invalido.") from error


def _parse_date(value, field):
    if isinstance(value, dt):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for date_format in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return dt.strptime(raw, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{field} invalida; use dd/mm/aaaa.")


def _notice_value(value):
    if isinstance(value, (date, dt)):
        parsed = value.date() if isinstance(value, dt) else value
        return parsed.strftime("%d/%m/%Y")
    raw = str(value or "").strip()
    return raw[:50] or None


def _cell(row, index):
    return row[index] if index < len(row) else None


def _first_filled_cell(row, indexes):
    for index in indexes:
        value = _cell(row, index)
        if value not in (None, ""):
            return value
    return None


def _normalize_text(value):
    return (
        normalize("NFKD", str(value or ""))
        .encode("ascii", "ignore")
        .decode("ascii")
        .casefold()
        .strip()
    )


def _employee_code(value):
    try:
        return int(str(value).strip().replace(".0", ""))
    except (TypeError, ValueError):
        return None


def _termination_reason(rows, employee_index, columns):
    """Localiza o motivo mesmo quando o relatório repete o cabeçalho em uma quebra de página."""
    ignored = {"empregado", "motivo da demissao"}
    limit = min(len(rows), employee_index + 26)

    for candidate_index in range(employee_index + 1, limit):
        candidate = rows[candidate_index]
        candidate_code = _employee_code(_cell(candidate, columns["matricula"]))
        candidate_dismissal = _cell(candidate, columns["demissao"])

        # Não usa o nome do próximo colaborador como motivo caso o relatório
        # realmente esteja sem a linha de motivo.
        if candidate_code is not None and candidate_dismissal not in (None, ""):
            break

        value = str(_cell(candidate, columns["nome"]) or "").strip()
        normalized = _normalize_text(value)
        if not value or normalized in ignored:
            continue
        return value

    return ""


def _complete_years(start, end):
    years = end.year - start.year
    if (end.month, end.day) < (start.month, start.day):
        years -= 1
    return max(0, years)


def _add_months(value, months):
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    return value.replace(year=year, month=month, day=min(value.day, monthrange(year, month)[1]))


def _thirteenth_months(admission, dismissal):
    total = 0
    for month in range(1, dismissal.month + 1):
        first = date(dismissal.year, month, 1)
        last = date(dismissal.year, month, monthrange(dismissal.year, month)[1])
        worked_from = max(first, admission)
        worked_to = min(last, dismissal)
        if worked_from <= worked_to and (worked_to - worked_from).days + 1 >= 15:
            total += 1
    return min(12, total)


def _vacation_months(admission, dismissal):
    anniversary_year = dismissal.year
    try:
        period_start = admission.replace(year=anniversary_year)
    except ValueError:
        period_start = date(anniversary_year, 2, 28)
    if period_start > dismissal:
        try:
            period_start = admission.replace(year=anniversary_year - 1)
        except ValueError:
            period_start = date(anniversary_year - 1, 2, 28)

    months = max(0, (dismissal.year - period_start.year) * 12 + dismissal.month - period_start.month)
    while months and _add_months(period_start, months) > dismissal:
        months -= 1
    remainder_start = _add_months(period_start, months)
    if (dismissal - remainder_start).days + 1 >= 15:
        months += 1
    return min(12, months)


def _branch_names(center_ids):
    center_ids = {value for value in center_ids if value is not None}
    if not center_ids:
        return {}
    result = {center_id: set() for center_id in center_ids}
    direct = (
        db.session.query(filial_centros_custo.c.centro_custo_id, Branch.nome)
        .join(Branch, Branch.id == filial_centros_custo.c.filial_id)
        .filter(filial_centros_custo.c.centro_custo_id.in_(center_ids))
        .all()
    )
    for center_id, name in direct:
        result.setdefault(center_id, set()).add(name)
    department = (
        db.session.query(CostCenters.id, Branch.nome)
        .join(
            filial_departamentos,
            filial_departamentos.c.departamento == CostCenters.departamento,
        )
        .join(Branch, Branch.id == filial_departamentos.c.filial_id)
        .filter(CostCenters.id.in_(center_ids))
        .all()
    )
    for center_id, name in department:
        result.setdefault(center_id, set()).add(name)
    return {center_id: sorted(names, key=str.casefold) for center_id, names in result.items()}


class TerminationService:
    @staticmethod
    def _query():
        return (
            db.session.query(
                Termination,
                Employees.id.label("colaborador_id"),
                Employees.matricula.label("matricula"),
                Employees.nome.label("colaborador_nome"),
                Employees.centro_id.label("centro_custo_id"),
                CostCenters.local.label("centro_custo"),
                CostCenters.departamento.label("departamento"),
                Supervisors.id.label("supervisor_id"),
                Supervisors.nome.label("supervisor_nome"),
                Users.nome.label("importado_por"),
            )
            .join(Employees, Employees.matricula == Termination.matricula)
            .outerjoin(CostCenters, CostCenters.id == Employees.centro_id)
            .outerjoin(Supervisors, Supervisors.id == CostCenters.supervisor_id)
            .outerjoin(Users, Users.id == Termination.importado_por_usuario_id)
        )

    @staticmethod
    def _serialize(row, branches):
        item = row.Termination
        return {
            "id": item.id,
            "colaborador_id": row.colaborador_id,
            "matricula": row.matricula,
            "nome": row.colaborador_nome,
            "centro_custo_id": row.centro_custo_id,
            "centro_custo": row.centro_custo or "Sem centro de custo",
            "departamento": row.departamento,
            "filiais": branches.get(row.centro_custo_id, []),
            "supervisor_id": row.supervisor_id,
            "supervisor": row.supervisor_nome or "Sem supervisor",
            "motivo_rescisao": item.motivo_rescisao,
            "data_admissao": item.data_admissao.isoformat(),
            "aviso": item.aviso,
            "data_demissao": item.data_demissao.isoformat(),
            "saldo_fgts": float(item.saldo_fgts or 0),
            "proventos": float(item.proventos or 0),
            "descontos": float(item.descontos or 0),
            "liquido": float(item.liquido or 0),
            "fgts_rescisorio": float(item.fgts_rescisorio or 0),
            "custo_bruto": float((item.proventos or 0) + (item.fgts_rescisorio or 0)),
            "arquivo_origem": item.arquivo_origem,
            "importado_por": row.importado_por,
            "importado_em": item.updated_at.isoformat() if item.updated_at else None,
        }

    @safe_route
    def read(self, token_data):
        if not has_permission(token_data, "controle_rescisoes", "view"):
            return jsonify("Voce nao possui acesso ao Controle de Rescisoes."), 403

        query = apply_cost_center_scope(
            self._query(),
            Employees.centro_id,
            token_data,
        )
        try:
            if request.args.get("inicio"):
                query = query.filter(Termination.data_demissao >= _parse_date(request.args["inicio"], "Data inicial"))
            if request.args.get("fim"):
                query = query.filter(Termination.data_demissao <= _parse_date(request.args["fim"], "Data final"))
            if request.args.get("centro_custo_id"):
                query = query.filter(Employees.centro_id == int(request.args["centro_custo_id"]))
            if request.args.get("supervisor_id"):
                query = query.filter(CostCenters.supervisor_id == int(request.args["supervisor_id"]))
            if request.args.get("departamento"):
                query = query.filter(CostCenters.departamento == int(request.args["departamento"]))
        except (TypeError, ValueError) as error:
            return jsonify(str(error) or "Filtro invalido."), 400

        reason = str(request.args.get("motivo") or "").strip()
        if reason:
            query = query.filter(Termination.motivo_rescisao == reason)
        search = str(request.args.get("busca") or "").strip()
        if search:
            pattern = f"%{search}%"
            query = query.filter(or_(
                Employees.nome.ilike(pattern),
                Termination.motivo_rescisao.ilike(pattern),
                cast(Employees.matricula, String).ilike(pattern),
                CostCenters.local.ilike(pattern),
                Supervisors.nome.ilike(pattern),
            ))

        rows = query.order_by(Termination.data_demissao.desc(), Termination.id.desc()).all()
        branches = _branch_names({row.centro_custo_id for row in rows})
        records = [self._serialize(row, branches) for row in rows]
        summary = {
            "total": len(records),
            "saldo_fgts": round(sum(item["saldo_fgts"] for item in records), 2),
            "proventos": round(sum(item["proventos"] for item in records), 2),
            "descontos": round(sum(item["descontos"] for item in records), 2),
            "liquido": round(sum(item["liquido"] for item in records), 2),
            "fgts_rescisorio": round(sum(item["fgts_rescisorio"] for item in records), 2),
            "custo_bruto": round(sum(item["custo_bruto"] for item in records), 2),
        }
        filters = {
            "motivos": sorted({item["motivo_rescisao"] for item in records}, key=str.casefold),
            "departamentos": sorted(
                {item["departamento"] for item in records if item["departamento"] is not None}
            ),
            "centros": sorted(
                [
                    {"label": item["centro_custo"], "value": item["centro_custo_id"]}
                    for item in {record["centro_custo_id"]: record for record in records}.values()
                ],
                key=lambda item: item["label"].casefold(),
            ),
            "supervisores": sorted(
                [
                    {"label": item["supervisor"], "value": item["supervisor_id"]}
                    for item in {
                        record["supervisor_id"]: record
                        for record in records
                        if record["supervisor_id"]
                    }.values()
                ],
                key=lambda item: item["label"].casefold(),
            ),
        }
        return jsonify({"registros": records, "resumo": summary, "filtros": filters}), 200

    @safe_route
    def import_xlsx(self, token_data):
        if not has_permission(token_data, "controle_rescisoes", "create"):
            return jsonify("Voce nao possui permissao para importar rescisoes."), 403
        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify("Envie a planilha padrao no formato .xlsx."), 400
        try:
            workbook = load_workbook(uploaded.stream, read_only=True, data_only=True)
            worksheet = next(
                (
                    workbook[name]
                    for name in workbook.sheetnames
                    if _normalize_text(name) == "relacao de rescisoes calculadas"
                ),
                workbook.active,
            )
            rows = list(worksheet.iter_rows(values_only=True))
        except (ValueError, OSError):
            return jsonify("Nao foi possivel ler a planilha."), 400

        header_index = next(
            (
                index for index, row in enumerate(rows)
                if _normalize_text(_cell(row, 0)) == "codigo"
                and "empregado" in _normalize_text(_cell(row, 2))
            ),
            None,
        )
        if header_index is None:
            return jsonify("Planilha fora do padrao: cabecalho Codigo/Empregado nao encontrado."), 400

        # Colunas do relatorio Relação de Rescisões Calculadas.
        columns = {
            "matricula": 0,
            "nome": 2,
            "admissao": 10,
            "aviso": 13,
            "demissao": 16,
        }
        money_columns = {
            "saldo_fgts": (17,),
            "proventos": (23,),
            "descontos": (26,),
            "liquido": (28,),
            # O relatório exibe o título em AH, porém algumas exportações
            # gravam o valor na célula anterior (AG) por causa da mesclagem.
            "fgts_rescisorio": (32, 33),
        }
        parsed_rows = []
        errors = []
        source_name = Path(uploaded.filename).name[:255]
        for index in range(header_index + 1, len(rows)):
            row = rows[index]
            raw_code = _cell(row, columns["matricula"])
            if raw_code in (None, ""):
                continue
            matricula = _employee_code(raw_code)
            if matricula is None:
                # Totais e rodapes nao sao registros de colaborador.
                continue
            raw_dismissal = _cell(row, columns["demissao"])
            if raw_dismissal in (None, ""):
                continue
            try:
                admission = _parse_date(_cell(row, columns["admissao"]), "Admissao")
                dismissal = _parse_date(raw_dismissal, "Demissao")
                reason = _termination_reason(rows, index, columns)
                name = str(_cell(row, columns["nome"]) or "").strip()
                if not name or not reason:
                    raise ValueError("nome ou motivo da rescisao nao informado")
                values = {
                    key: _money(_first_filled_cell(row, positions), key.replace("_", " ").title())
                    for key, positions in money_columns.items()
                }
            except ValueError as error:
                errors.append(f"Linha {index + 1}: {error}.")
                continue
            parsed_rows.append({
                "matricula": matricula,
                "nome": name,
                "admissao": admission,
                "aviso": _notice_value(_cell(row, columns["aviso"])),
                "demissao": dismissal,
                "motivo": reason[:500],
                **values,
            })

        if errors:
            return jsonify({"message": "A importacao foi cancelada; corrija a planilha.", "errors": errors[:50]}), 400
        if not parsed_rows:
            return jsonify("Nenhuma rescisao foi encontrada na planilha."), 400

        duplicate_keys = set()
        seen_keys = set()
        for item in parsed_rows:
            key = (item["matricula"], item["demissao"])
            if key in seen_keys:
                duplicate_keys.add(key)
            seen_keys.add(key)
        if duplicate_keys:
            formatted = ", ".join(f"{code} em {day.strftime('%d/%m/%Y')}" for code, day in sorted(duplicate_keys))
            return jsonify(f"A planilha possui rescisoes duplicadas: {formatted}."), 400

        matriculas = {item["matricula"] for item in parsed_rows}
        employees = {
            employee.matricula: employee
            for employee in Employees.query.filter(Employees.matricula.in_(matriculas)).all()
        }
        # Calcula o escopo uma única vez. Para usuários com seletor de filial,
        # recalcular dentro do laço executaria várias consultas por colaborador.
        allowed_centers = allowed_cost_center_ids(token_data)
        validation_errors = []
        for item in parsed_rows:
            employee = employees.get(item["matricula"])
            if not employee:
                validation_errors.append(f"Matricula {item['matricula']}: colaborador nao encontrado.")
            elif not employee.centro_id:
                validation_errors.append(f"Matricula {item['matricula']}: colaborador sem centro de custo.")
            elif allowed_centers is not None and employee.centro_id not in allowed_centers:
                validation_errors.append(f"Matricula {item['matricula']}: filial fora do seu acesso.")
        if validation_errors:
            return jsonify({
                "message": "A importacao foi cancelada; nenhum registro foi gravado.",
                "errors": validation_errors[:50],
            }), 400

        keys = {(item["matricula"], item["demissao"]) for item in parsed_rows}
        existing = {
            (record.matricula, record.data_demissao): record
            for record in Termination.query.filter(
                Termination.matricula.in_(matriculas),
                Termination.data_demissao.in_({item["demissao"] for item in parsed_rows}),
            ).all()
            if (record.matricula, record.data_demissao) in keys
        }
        created = 0
        updated = 0
        new_records = []

        # Uma única instrução UPDATE substitui até milhares de atualizações
        # individuais da situação dos colaboradores.
        situations_updated = (
            Employees.query
            .filter(
                Employees.matricula.in_(matriculas),
                or_(Employees.situacao.is_(None), Employees.situacao != 8),
            )
            .update({Employees.situacao: 8}, synchronize_session=False)
        )

        for data in parsed_rows:
            employee = employees[data["matricula"]]
            key = (employee.matricula, data["demissao"])
            record = existing.get(key)
            if record:
                updated += 1
            else:
                record = Termination(
                    matricula=employee.matricula,
                    data_demissao=data["demissao"],
                )
                new_records.append(record)
                created += 1
            record.motivo_rescisao = data["motivo"]
            record.data_admissao = data["admissao"]
            record.aviso = data["aviso"]
            record.saldo_fgts = data["saldo_fgts"]
            record.proventos = data["proventos"]
            record.descontos = data["descontos"]
            record.liquido = data["liquido"]
            record.fgts_rescisorio = data["fgts_rescisorio"]
            record.arquivo_origem = source_name
            record.importado_por_usuario_id = token_data.get("id")
            record.updated_at = dt.now()

        if new_records:
            db.session.add_all(new_records)
        db.session.commit()
        socketio.emit("termination_update", {
            "action": "imported",
            "created": created,
            "updated": updated,
            "situations_updated": situations_updated,
        })
        return jsonify({
            "message": (
                f"{created + updated} rescisao(oes) processada(s); "
                f"{situations_updated} colaborador(es) alterado(s) para a situacao 8."
            ),
            "criadas": created,
            "atualizadas": updated,
            "colaboradores_marcados_como_demitidos": situations_updated,
        }), 201

    @safe_route
    def calculate(self, token_data):
        if not has_permission(token_data, "controle_rescisoes", "view"):
            return jsonify("Voce nao possui acesso ao calculo de rescisao."), 403
        body = request.get_json(silent=True) or {}
        try:
            employee_id = int(body.get("colaborador_id"))
            dismissal = _parse_date(body.get("data_demissao"), "Data de demissao")
        except (TypeError, ValueError) as error:
            return jsonify(str(error) or "Informe o colaborador."), 400
        employee = db.session.get(Employees, employee_id)
        if not employee:
            return jsonify("Colaborador nao encontrado."), 404
        if not employee.centro_id or not can_access_cost_center(token_data, employee.centro_id):
            return jsonify("Voce nao possui acesso a filial deste colaborador."), 403
        if not employee.data_admissao:
            return jsonify("O colaborador nao possui data de admissao cadastrada."), 400
        salary_value = getattr(employee, "salario", None)
        if salary_value in (None, "") or _money(salary_value) <= 0:
            return jsonify("O colaborador nao possui salario valido cadastrado."), 400

        admission = employee.data_admissao.date() if isinstance(employee.data_admissao, dt) else employee.data_admissao
        if dismissal < admission:
            return jsonify("A data de demissao nao pode ser anterior a admissao."), 400
        reason = str(body.get("motivo") or "sem_justa_causa").strip()
        notice_type = str(body.get("tipo_aviso") or "indenizado").strip()
        if reason not in VALID_REASONS:
            return jsonify("Motivo de rescisao invalido."), 400
        if notice_type not in VALID_NOTICE:
            return jsonify("Tipo de aviso previo invalido."), 400

        try:
            salary = _money(salary_value, "Salario")
            fgts_balance = _money(body.get("saldo_fgts"), "Saldo FGTS")
            other_earnings = _money(body.get("outras_verbas"), "Outras verbas")
            manual_discounts = _money(body.get("descontos"), "Descontos")
            full_vacations = max(0, int(body.get("ferias_integrais") or 0))
            double_vacations = max(0, int(body.get("ferias_em_dobro") or 0))
            days_worked = int(body.get("dias_saldo") if body.get("dias_saldo") not in (None, "") else dismissal.day)
            thirteenth_months = int(body.get("avos_decimo_terceiro") if body.get("avos_decimo_terceiro") not in (None, "") else _thirteenth_months(admission, dismissal))
            vacation_months = int(body.get("avos_ferias") if body.get("avos_ferias") not in (None, "") else _vacation_months(admission, dismissal))
        except (TypeError, ValueError) as error:
            return jsonify(str(error) or "Parametros de calculo invalidos."), 400
        if not 0 <= days_worked <= 30 or not 0 <= thirteenth_months <= 12 or not 0 <= vacation_months <= 12:
            return jsonify("Dias de saldo devem estar entre 0 e 30 e os avos entre 0 e 12."), 400

        salary_balance = (salary / Decimal("30") * days_worked).quantize(MONEY)
        has_proportional_rights = reason != "justa_causa"
        thirteenth = (salary * thirteenth_months / Decimal("12")).quantize(MONEY) if has_proportional_rights else Decimal("0")
        proportional_vacation_base = (salary * vacation_months / Decimal("12")).quantize(MONEY) if has_proportional_rights else Decimal("0")
        proportional_vacation_third = (proportional_vacation_base / Decimal("3")).quantize(MONEY)
        full_vacation_base = (salary * full_vacations).quantize(MONEY)
        full_vacation_third = (full_vacation_base / Decimal("3")).quantize(MONEY)
        double_vacation_base = (salary * Decimal("2") * double_vacations).quantize(MONEY)
        double_vacation_third = (double_vacation_base / Decimal("3")).quantize(MONEY)

        notice_days = min(90, 30 + 3 * _complete_years(admission, dismissal))
        notice_earnings = Decimal("0")
        notice_discount = Decimal("0")
        if notice_type == "indenizado" and reason in {"sem_justa_causa", "acordo"}:
            factor = Decimal("0.5") if reason == "acordo" else Decimal("1")
            notice_earnings = (salary / Decimal("30") * notice_days * factor).quantize(MONEY)
        elif notice_type == "descontado" and reason == "pedido_demissao":
            notice_discount = salary

        components = {
            "saldo_salario": salary_balance,
            "decimo_terceiro_proporcional": thirteenth,
            "ferias_proporcionais": proportional_vacation_base,
            "terco_ferias_proporcionais": proportional_vacation_third,
            "ferias_integrais": full_vacation_base,
            "terco_ferias_integrais": full_vacation_third,
            "ferias_em_dobro": double_vacation_base,
            "terco_ferias_em_dobro": double_vacation_third,
            "aviso_previo_indenizado": notice_earnings,
            "outras_verbas": other_earnings,
        }
        earnings = sum(components.values(), Decimal("0")).quantize(MONEY)
        discounts = (manual_discounts + notice_discount).quantize(MONEY)
        liquid = (earnings - discounts).quantize(MONEY)
        fgts_base = salary_balance + thirteenth + notice_earnings
        fgts_termination = (fgts_base * Decimal("0.08")).quantize(MONEY)
        fine_rate = Decimal("0.40") if reason == "sem_justa_causa" else Decimal("0.20") if reason == "acordo" else Decimal("0")
        fgts_fine = ((fgts_balance + fgts_termination) * fine_rate).quantize(MONEY)
        company_cost = (earnings + fgts_termination + fgts_fine).quantize(MONEY)

        return jsonify({
            "colaborador": {
                "id": employee.id,
                "matricula": employee.matricula,
                "nome": employee.nome,
                "salario": float(salary),
                "data_admissao": admission.isoformat(),
                "data_demissao": dismissal.isoformat(),
            },
            "parametros": {
                "motivo": reason,
                "motivo_label": VALID_REASONS[reason],
                "tipo_aviso": notice_type,
                "dias_aviso": notice_days if notice_earnings else 0,
                "dias_saldo": days_worked,
                "avos_decimo_terceiro": thirteenth_months if has_proportional_rights else 0,
                "avos_ferias": vacation_months if has_proportional_rights else 0,
                "percentual_multa_fgts": float(fine_rate * 100),
            },
            "verbas": {key: float(value) for key, value in components.items()},
            "desconto_aviso": float(notice_discount),
            "descontos": float(discounts),
            "proventos": float(earnings),
            "liquido_estimado": float(liquid),
            "fgts_rescisorio_estimado": float(fgts_termination),
            "multa_fgts_estimada": float(fgts_fine),
            "custo_empresa_estimado": float(company_cost),
            "observacoes": [
                "Provisao estimada; valide o calculo final no sistema de folha/eSocial.",
                "INSS, IRRF, medias variaveis, convencao coletiva e eventos ja pagos nao foram calculados.",
                "Os avos de ferias usam a ultima data de aniversario da admissao e devem ser ajustados se houver historico de ferias diferente.",
                "O FGTS estimado usa 8% sobre saldo de salario, 13o proporcional e aviso indenizado.",
            ],
        }), 200

    @safe_route
    def delete_all(self, token_data):
        if not is_admin(token_data):
            return jsonify("Apenas administradores podem excluir todas as rescisoes."), 403

        affected = Termination.query.delete(synchronize_session=False)
        db.session.commit()
        socketio.emit("termination_update", {
            "action": "deleted_all",
            "deleted": affected,
        })
        return jsonify({
            "message": f"{affected} rescisao(oes) excluida(s).",
            "excluidas": affected,
        }), 200

    @safe_route
    def delete(self, termination_id, token_data):
        if not has_permission(token_data, "controle_rescisoes", "edit"):
            return jsonify("Voce nao possui permissao para excluir rescisoes."), 403
        item = db.session.get(Termination, termination_id)
        if not item:
            return jsonify("Rescisao nao encontrada."), 404
        employee = Employees.query.filter_by(matricula=item.matricula).first()
        if not employee or not can_access_cost_center(token_data, employee.centro_id):
            return jsonify("Voce nao possui acesso a filial desta rescisao."), 403
        db.session.delete(item)
        db.session.commit()
        socketio.emit("termination_update", {"action": "deleted", "id": termination_id})
        return jsonify("Rescisao excluida."), 200

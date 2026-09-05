# Regras de negócio de usuários.
# Dependências externas.
from flask import request as rq, jsonify, send_file
# Módulos internos da aplicação.
from models.usuarios import Users, db
from utils.check_field import check_field
# Biblioteca padrão.
from hashlib import sha256
# Módulos internos da aplicação.
from utils.safe_route import safe_route
# Biblioteca padrão.
from datetime import datetime as dt, timedelta
from email.message import EmailMessage
from os import getenv
from pathlib import Path
from uuid import uuid4
import json
import re
import secrets
import smtplib
from unicodedata import normalize
from io import BytesIO
# Dependências externas.
from openpyxl import Workbook, load_workbook
import pypdfium2 as pdfium
from PIL import Image, ImageDraw, ImageOps, ImageStat, UnidentifiedImageError
# Módulos internos da aplicação.
from models.filiais import Branch, filial_usuarios
from models.colaboradores import Employees
from models.marketplace import MarketplaceProduct, MarketplacePurchase
from utils.filial_scope import is_admin
from utils.permissions import PERMISSION_CATALOG, replace_permissions, serialize_permissions
from utils.password_security import (
    hash_password,
    is_default_password,
    is_strong_password,
    verify_password,
)
from utils.token import create_token
from utils.maintenance import maintenance_mode_enabled, update_maintenance_mode
from utils.theme_access import (
    CUSTOM_THEMES,
    available_themes_for,
    can_use_theme,
    effective_theme_for,
)
from utils.user_requirements import (
    auth_requirements,
    is_valid_cpf,
    normalize_cpf,
    refresh_user_requirements,
    validate_profile_photo,
)

EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
REGISTERED_SIGNATURE_DIR = Path(
    getenv("REGISTERED_SIGNATURE_DIR")
    or Path(__file__).resolve().parents[1] / "storage" / "assinaturas_cadastradas"
)
MAX_REGISTERED_SIGNATURE_SIZE = 5 * 1024 * 1024
MAX_REGISTERED_SIGNATURE_PIXELS = 16_000_000
REGISTERED_SIGNATURE_EXTENSION = ".png"
ALLOWED_REGISTERED_SIGNATURE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".pdf"}
ALLOWED_REGISTERED_SIGNATURE_FORMATS = {"PNG", "JPEG", "WEBP"}

class UserServices:
    @staticmethod
    def _is_admin(token_data):
        return is_admin(token_data)

    @staticmethod
    def _normalize_cpf(value):
        return normalize_cpf(value)

    @staticmethod
    def _issue_token(user):
        persistent = bool(user.token_sem_expiracao)
        return create_token({
            "id": user.id,
            "perm": user.role,
            "ver": int(user.token_version or 0),
            "sessao_persistente": persistent,
        }, expires=not persistent)

    @safe_route
    def read(self, token_data):
        detailed = rq.args.get("detail") == "1"
        include_photo = rq.args.get("include_photo") == "1"
        admin = self._is_admin(token_data)
        if detailed and not admin:
            return jsonify("Apenas administradores podem consultar configurações de usuários."), 403

        users_query = Users.query
        if not admin:
            current_user_branches = filial_usuarios.alias("current_user_branches")
            branch_ids = db.session.query(current_user_branches.c.filial_id).filter(
                current_user_branches.c.usuario_id == token_data.get("id")
            )
            users_query = (
                users_query.join(
                    filial_usuarios,
                    filial_usuarios.c.usuario_id == Users.id,
                )
                .filter(filial_usuarios.c.filial_id.in_(branch_ids))
                .distinct()
            )
        users = users_query.order_by(Users.nome).all()

        if not detailed:
            return jsonify([{
                "id": user.id,
                "nome": user.nome,
                "adorno_foto": user.adorno_foto,
                **({"foto_perfil": user.foto_perfil} if include_photo else {}),
            } for user in users]), 200

        return jsonify([{
            "id": user.id,
            "nome": user.nome,
            "email": user.email,
            "cpf": user.cpf if admin else None,
            "role": user.role,
            "gerencia_faltas": bool(user.gerencia_faltas),
            "created_at": user.created_at,
            "last_login": user.last_login,
            "filial_ids": sorted(branch.id for branch in user.filiais),
            "permissions": serialize_permissions(user),
            "assinatura_cadastrada": bool(user.assinatura_cadastrada),
            **({"foto_perfil": user.foto_perfil} if include_photo else {}),
        } for user in users]), 200

    @safe_route
    def create(self, token_data):
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem criar usuários."), 403

        user, error = self._build_user(rq.get_json(silent=True) or {})
        if error:
            return jsonify(error), 400

        branch_error = self._apply_branches(user, rq.get_json(silent=True) or {})
        if branch_error:
            return jsonify(branch_error), 400
        db.session.add(user)
        db.session.flush()
        permission_error = replace_permissions(user, (rq.get_json(silent=True) or {}).get("permissions"))
        if permission_error:
            db.session.rollback()
            return jsonify(permission_error), 400
        db.session.commit()
        return jsonify(self._serialize_admin(user)), 201

    @safe_route
    def update(self, token_data):
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem modificar usuários."), 403

        body = rq.get_json(silent=True) or {}
        user = db.session.get(Users, body.get("id"))
        if not user:
            return jsonify("Usuário não encontrado."), 404

        error = self._apply_user_changes(user, body)
        if error:
            return jsonify(error), 400

        branch_error = self._apply_branches(user, body)
        if branch_error:
            return jsonify(branch_error), 400
        permission_error = replace_permissions(user, body.get("permissions") if "permissions" in body else None)
        if permission_error:
            db.session.rollback()
            return jsonify(permission_error), 400

        db.session.commit()
        return jsonify(self._serialize_admin(user)), 200

    @staticmethod
    def _registered_signature_path(filename):
        """Resolve um nome de arquivo sem aceitar caminhos enviados pelo cliente."""
        safe_name = Path(str(filename or "")).name
        if not safe_name or safe_name != filename or Path(safe_name).suffix.lower() != REGISTERED_SIGNATURE_EXTENSION:
            return None
        return REGISTERED_SIGNATURE_DIR / safe_name

    @staticmethod
    def _signature_image_from_upload(upload, extension):
        """Lê imagem ou primeira página de PDF e valida o conteúdo real do arquivo."""
        upload.stream.seek(0, 2)
        size = upload.stream.tell()
        upload.stream.seek(0)
        if not size or size > MAX_REGISTERED_SIGNATURE_SIZE:
            return None, "O arquivo de assinatura deve ter no máximo 5 MB."

        raw_file = upload.stream.read()
        try:
            if extension == ".pdf":
                document = pdfium.PdfDocument(BytesIO(raw_file))
                if len(document) < 1:
                    document.close()
                    return None, "O PDF enviado não possui páginas."
                page = document[0]
                image = page.render(scale=2).to_pil().copy()
                page.close()
                document.close()
            else:
                with Image.open(BytesIO(raw_file)) as source:
                    if source.format not in ALLOWED_REGISTERED_SIGNATURE_FORMATS:
                        return None, "O arquivo enviado não é uma imagem de assinatura válida."
                    width, height = source.size
                    if not width or not height or width * height > MAX_REGISTERED_SIGNATURE_PIXELS:
                        return None, "A imagem da assinatura possui dimensões inválidas ou muito grandes."
                    source.load()
                    image = ImageOps.exif_transpose(source).copy()
        except (Image.DecompressionBombError, UnidentifiedImageError, OSError, ValueError, pdfium.PdfiumError):
            return None, "Não foi possível ler o arquivo de assinatura."

        width, height = image.size
        if not width or not height or width * height > MAX_REGISTERED_SIGNATURE_PIXELS:
            return None, "A imagem da assinatura possui dimensões inválidas ou muito grandes."
        return image, None

    @staticmethod
    def _normalize_signature(image):
        """Isola o papel claro, remove o fundo e gera uma PNG transparente."""
        image = image.convert("RGBA")
        image.thumbnail((1800, 1200), Image.Resampling.LANCZOS)

        # Fotos feitas sobre uma mesa ou parede escura precisam ser recortadas
        # primeiro na região clara do papel, antes de identificar o traço.
        visible_image = Image.alpha_composite(
            Image.new("RGBA", image.size, "white"),
            image,
        )
        visible_grayscale = ImageOps.grayscale(visible_image.convert("RGB"))
        paper_mask = visible_grayscale.point(lambda value: 255 if value >= 205 else 0)
        paper_bounds = paper_mask.getbbox()
        if paper_bounds:
            left, top, right, bottom = paper_bounds
            paper_area = (right - left) * (bottom - top)
            image_area = image.width * image.height
            if image_area and paper_area >= image_area * .08:
                paper_padding = max(12, round(max(image.size) * .015))
                image = image.crop((
                    max(0, left - paper_padding),
                    max(0, top - paper_padding),
                    min(image.width, right + paper_padding),
                    min(image.height, bottom + paper_padding),
                ))

        alpha = image.getchannel("A")
        grayscale = ImageOps.grayscale(image.convert("RGB"))
        # Um limite mais baixo evita transformar o papel sombreado em tinta.
        ink_mask = grayscale.point(lambda value: 255 if value < 150 else 0)
        if alpha.getextrema()[0] < 255:
            ink_mask = Image.composite(ink_mask, alpha, alpha)

        # Manchas ligadas à borda normalmente são a mesa, a sombra ou a margem
        # do papel. O recorte da interface deve manter uma pequena margem ao
        # redor da assinatura para que o traço real não seja removido aqui.
        border_points = [
            *((x, 0) for x in range(ink_mask.width)),
            *((x, ink_mask.height - 1) for x in range(ink_mask.width)),
            *((0, y) for y in range(ink_mask.height)),
            *((ink_mask.width - 1, y) for y in range(ink_mask.height)),
        ]
        for point in border_points:
            if ink_mask.getpixel(point):
                ImageDraw.floodfill(ink_mask, point, 0)

        bounds = ink_mask.getbbox()
        if not bounds:
            return None, "Não foi possível identificar o traço da assinatura."

        left, top, right, bottom = bounds
        padding = max(12, round(max(image.size) * .02))
        crop_box = (
            max(0, left - padding),
            max(0, top - padding),
            min(image.width, right + padding),
            min(image.height, bottom + padding),
        )
        signature = Image.new("RGBA", (crop_box[2] - crop_box[0], crop_box[3] - crop_box[1]), (20, 35, 40, 0))
        signature.putalpha(ink_mask.crop(crop_box))
        signature.thumbnail((1200, 420), Image.Resampling.LANCZOS)
        ink_coverage = ImageStat.Stat(signature.getchannel("A")).mean[0] / 255
        if ink_coverage > .55:
            return None, "Não foi possível separar a assinatura do fundo. Envie um recorte da assinatura sobre fundo claro."
        return signature, None

    @staticmethod
    def _apply_signature_crop(image, crop_data):
        """Aplica o recorte proporcional selecionado pelo administrador na prévia."""
        if not crop_data:
            return image, None
        try:
            crop = json.loads(crop_data)
            x = float(crop.get("x"))
            y = float(crop.get("y"))
            width = float(crop.get("width"))
            height = float(crop.get("height"))
        except (AttributeError, TypeError, ValueError, json.JSONDecodeError):
            return None, "O recorte da assinatura é inválido."

        if not 0 <= x < 1 or not 0 <= y < 1 or not .03 <= width <= 1 or not .03 <= height <= 1:
            return None, "Selecione uma área válida para o recorte da assinatura."
        if x + width > 1 or y + height > 1:
            return None, "O recorte da assinatura ultrapassa os limites da imagem."

        left = round(image.width * x)
        top = round(image.height * y)
        right = round(image.width * (x + width))
        bottom = round(image.height * (y + height))
        if right <= left or bottom <= top:
            return None, "Selecione uma área válida para o recorte da assinatura."
        return image.crop((left, top, right, bottom)), None

    @classmethod
    def _registered_signature_filename(cls, user):
        """Monta um nome legível com primeiro e último nome do titular."""
        normalized_name = normalize("NFKD", str(user.nome or "")).encode("ascii", "ignore").decode().lower()
        parts = [part for part in re.sub(r"[^a-z0-9]+", " ", normalized_name).split() if part]
        first_name = parts[0] if parts else f"usuario_{user.id}"
        last_name = parts[-1] if len(parts) > 1 else first_name
        filename = f"{first_name}_{last_name}{REGISTERED_SIGNATURE_EXTENSION}"

        # Evita que nomes iguais substituam a assinatura pertencente a outro usuário.
        existing_owner = Users.query.filter(
            Users.assinatura_cadastrada == filename,
            Users.id != user.id,
        ).first()
        return filename if not existing_owner else f"{first_name}_{last_name}_{user.id}{REGISTERED_SIGNATURE_EXTENSION}"

    @classmethod
    def _store_registered_signature(cls, upload, user):
        """Normaliza a assinatura e a grava temporariamente antes de nomeá-la."""
        if not upload or not upload.filename:
            return None, "Selecione um arquivo de assinatura."
        extension = Path(upload.filename).suffix.lower()
        if extension not in ALLOWED_REGISTERED_SIGNATURE_EXTENSIONS:
            return None, "Envie um arquivo PNG, JPG, JPEG, WEBP ou PDF."

        image, error = cls._signature_image_from_upload(upload, extension)
        if error:
            return None, error
        image, error = cls._apply_signature_crop(image, rq.form.get("recorte"))
        if error:
            return None, error
        signature, error = cls._normalize_signature(image)
        if error:
            return None, error

        output = BytesIO()
        signature.save(output, format="PNG", optimize=True)
        if output.tell() > MAX_REGISTERED_SIGNATURE_SIZE:
            return None, "A assinatura processada excedeu o limite de 5 MB."

        REGISTERED_SIGNATURE_DIR.mkdir(parents=True, exist_ok=True)
        filename = cls._registered_signature_filename(user)
        temporary_path = REGISTERED_SIGNATURE_DIR / f".{uuid4().hex}.tmp"
        try:
            with temporary_path.open("xb") as stored_file:
                stored_file.write(output.getvalue())
        except OSError:
            return None, "Não foi possível salvar a assinatura no servidor."
        return (filename, temporary_path), None

    @safe_route
    def register_signature(self, user_id, token_data):
        """Cadastra ou substitui a assinatura de um usuário somente por administrador."""
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem cadastrar assinaturas."), 403

        user = db.session.get(Users, user_id)
        if not user:
            return jsonify("Usuário não encontrado."), 404

        stored_signature, error = self._store_registered_signature(rq.files.get("arquivo"), user)
        if error:
            return jsonify(error), 400
        filename, temporary_path = stored_signature

        previous_filename = user.assinatura_cadastrada
        user.assinatura_cadastrada = filename
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            if temporary_path.is_file():
                temporary_path.unlink()
            raise

        try:
            temporary_path.replace(self._registered_signature_path(filename))
        except OSError:
            user.assinatura_cadastrada = previous_filename
            db.session.commit()
            if temporary_path.is_file():
                temporary_path.unlink()
            return jsonify("Não foi possível finalizar o salvamento da assinatura."), 500

        previous_path = self._registered_signature_path(previous_filename)
        if previous_path and previous_path.is_file() and previous_path.name != filename:
            try:
                previous_path.unlink()
            except OSError:
                pass
        return jsonify({
            "message": f"Assinatura de {user.nome} cadastrada com sucesso.",
            "usuario_id": user.id,
            "assinatura_cadastrada": True,
        }), 200

    @safe_route
    def delete(self): ...

    @safe_route
    def import_users(self, token_data):
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem importar usuários."), 403

        uploaded = rq.files.get("file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify("Envie uma planilha no formato .xlsx."), 400

        try:
            workbook = load_workbook(uploaded.stream, read_only=True, data_only=True)
            rows = workbook.active.iter_rows(values_only=True)
            headers = [str(value or "").strip().lower() for value in next(rows)]
        except (StopIteration, ValueError, OSError):
            return jsonify("Não foi possível ler a planilha."), 400

        required = ["nome", "cpf", "email", "role", "password"]
        if any(header not in headers for header in required):
            return jsonify(f"A planilha deve conter as colunas: {', '.join(required)}."), 400

        indexes = {header: headers.index(header) for header in required}
        created = []
        errors = []

        for row_number, row in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            if row_number > 1001:
                errors.append("A planilha pode conter no máximo 1000 usuários.")
                break

            data = {key: row[indexes[key]] for key in required}
            user, error = self._build_user(data)
            if error:
                errors.append(f"Linha {row_number}: {error}")
                continue
            db.session.add(user)
            db.session.flush()
            created.append(user)

        if errors:
            db.session.rollback()
            return jsonify({"message": "A importação foi cancelada.", "errors": errors}), 400
        if not created:
            return jsonify("A planilha não contém usuários para importar."), 400

        db.session.commit()
        return jsonify({"message": f"{len(created)} usuários importados com sucesso.", "total": len(created)}), 201

    @safe_route
    def download_template(self, token_data):
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem baixar o modelo de importação."), 403

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Usuarios"
        worksheet.append(["nome", "cpf", "email", "role", "password"])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = "A1:E1"
        for column, width in {"A": 32, "B": 16, "C": 34, "D": 12, "E": 24}.items():
            worksheet.column_dimensions[column].width = width

        instructions = workbook.create_sheet("Instrucoes")
        instructions.append(["Campo", "Regra"])
        instructions.append(["nome", "Obrigatório"])
        instructions.append(["cpf", "Opcional; quando informado, use 11 dígitos sem pontuação"])
        instructions.append(["email", "Opcional, válido e único"])
        instructions.append(["role", "SUPERVISOR, GERENTE, USER ou ADMIN"])
        instructions.append(["password", "Mínimo 8 caracteres, com maiúscula, minúscula, número e símbolo"])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="modelo_importacao_usuarios.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_user(self, body):
        nome = str(body.get("nome") or "").strip()
        cpf = self._normalize_cpf(body.get("cpf")) or None
        email = str(body.get("email") or "").strip().lower() or None
        role = str(body.get("role") or "USER").strip().upper()
        password = str(body.get("password") or "")

        # A seleção de colaborador é opcional e serve como origem segura para
        # preencher os dados básicos da nova conta. Campos informados
        # explicitamente pelo administrador continuam podendo ser ajustados.
        collaborator_id = body.get("colaborador_id")
        if collaborator_id not in (None, ""):
            try:
                collaborator = db.session.get(Employees, int(collaborator_id))
            except (TypeError, ValueError):
                collaborator = None
            if not collaborator:
                return None, "Colaborador não encontrado."

            collaborator_name = " ".join(str(collaborator.nome or "").split())
            name_parts = collaborator_name.split()
            collaborator_short_name = (
                f"{name_parts[0]} {name_parts[-1]}" if len(name_parts) > 1 else collaborator_name
            )
            if not nome:
                nome = collaborator_short_name
            if not cpf:
                cpf = self._normalize_cpf(collaborator.cpf) or None

        ok, error = check_field(nome=nome, senha=password)
        if not ok:
            return None, error
        if cpf and not is_valid_cpf(cpf):
            return None, "Informe um CPF válido."
        if email and not EMAIL_PATTERN.fullmatch(email):
            return None, "Informe um e-mail válido."
        if role not in {"SUPERVISOR", "GERENTE", "USER", "ADMIN"}:
            return None, "A role deve ser SUPERVISOR, GERENTE, USER ou ADMIN."
        if not is_strong_password(password):
            return None, "A senha deve ter ao menos 8 caracteres, com maiúscula, minúscula, número e caractere especial."
        if cpf and Users.query.filter_by(cpf=cpf).first():
            return None, "CPF já cadastrado."
        if email and Users.query.filter_by(email=email).first():
            return None, "E-mail já cadastrado."

        return Users(
            nome=nome,
            cpf=cpf,
            email=email,
            role=role,
            gerencia_faltas=bool(body.get("gerencia_faltas", False)),
            hash=hash_password(password),
            primeiro_acesso=True,
            cpf_pendente=not bool(cpf),
            foto_pendente=False,
            troca_senha_obrigatoria=False,
            senha_padrao=is_default_password(password),
            token_version=0,
        ), None

    def _apply_user_changes(self, user, body):
        if not any(key in body for key in ("nome", "cpf", "email", "role", "password", "filial_ids", "gerencia_faltas", "permissions")):
            return "Nenhuma alteração informada."

        if "nome" in body:
            nome = str(body.get("nome") or "").strip()
            if len(nome) < 2:
                return "Informe um nome válido."
            user.nome = nome

        if "cpf" in body:
            cpf = self._normalize_cpf(body.get("cpf")) or None
            if cpf and not is_valid_cpf(cpf):
                return "Informe um CPF válido."
            if cpf and Users.query.filter(Users.cpf == cpf, Users.id != user.id).first():
                return "CPF já cadastrado."
            user.cpf = cpf

        if "email" in body:
            email = str(body.get("email") or "").strip().lower() or None
            if email and not EMAIL_PATTERN.fullmatch(email):
                return "Informe um e-mail válido."
            if email and Users.query.filter(Users.email == email, Users.id != user.id).first():
                return "E-mail já cadastrado."
            user.email = email

        if "role" in body:
            role = str(body.get("role") or "").strip().upper()
            if role not in {"SUPERVISOR", "GERENTE", "USER", "ADMIN"}:
                return "A role deve ser SUPERVISOR, GERENTE, USER ou ADMIN."
            user.role = role

        if body.get("password"):
            password = str(body["password"])
            if not is_strong_password(password):
                return "A senha deve ter ao menos 8 caracteres, com maiúscula, minúscula, número e caractere especial."
            user.hash = hash_password(password)
            user.senha_padrao = is_default_password(password)
            user.troca_senha_obrigatoria = False
            user.token_version = int(user.token_version or 0) + 1
            user.senha_alterada_em = dt.now()

        if "gerencia_faltas" in body:
            user.gerencia_faltas = bool(body.get("gerencia_faltas"))

        refresh_user_requirements(user)
        return None

    @staticmethod
    def _apply_branches(user, body):
        if "filial_ids" not in body:
            return None
        try:
            ids = {int(value) for value in (body.get("filial_ids") or [])}
        except (TypeError, ValueError):
            return "Informe filiais válidas."
        branches = Branch.query.filter(Branch.id.in_(ids), Branch.ativa.is_(True)).all() if ids else []
        if len(branches) != len(ids):
            return "Uma ou mais filiais não foram encontradas ou estão inativas."
        user.filiais = branches
        return None

    @staticmethod
    def _serialize_admin(user):
        return {
            "id": user.id,
            "nome": user.nome,
            "email": user.email,
            "cpf": user.cpf,
            "role": user.role,
            "gerencia_faltas": bool(user.gerencia_faltas),
            "created_at": user.created_at,
            "last_login": user.last_login,
            "filial_ids": sorted(branch.id for branch in user.filiais),
            "permissions": serialize_permissions(user),
        }

    @safe_route
    def profile(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        if not user:
            return jsonify("Usuário não encontrado."), 404
        return jsonify(self._serialize(user))

    @safe_route
    def maintenance(self, token_data):
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem consultar a manutenção."), 403
        return jsonify({"manutencao_ativa": maintenance_mode_enabled()}), 200

    @safe_route
    def update_maintenance(self, token_data):
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem alterar a manutenção."), 403
        active = (rq.get_json(silent=True) or {}).get("manutencao_ativa")
        if not isinstance(active, bool):
            return jsonify("Informe um valor válido para manutenção."), 400
        update_maintenance_mode(active, token_data.get("id"))
        return jsonify({"manutencao_ativa": maintenance_mode_enabled()}), 200

    @safe_route
    def support_admins(self, token_data):
        admins = (
            Users.query
            .filter(db.func.upper(Users.role) == "ADMIN")
            .order_by(Users.nome)
            .all()
        )
        return jsonify([
            {
                "id": admin.id,
                "nome": admin.nome,
                "foto_perfil": admin.foto_perfil,
            }
            for admin in admins
        ]), 200

    @safe_route
    def update_profile(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        if not user:
            return jsonify("Usuário não encontrado."), 404

        body = rq.get_json(silent=True) or {}
        nome = body.get("nome")
        foto = body.get("foto_perfil")
        tema = body.get("tema")
        modo_tema = body.get("modo_tema")
        has_particles_update = "particulas_ativas" in body
        particulas_ativas = body.get("particulas_ativas")
        senha_atual = body.get("senha_atual")
        nova_senha = body.get("nova_senha")
        has_timo_update = "timo_ativo" in body
        timo_ativo = body.get("timo_ativo")
        has_timo_home_update = "timo_tela_inicial" in body
        timo_tela_inicial = body.get("timo_tela_inicial")
        timo_cenario = body.get("timo_cenario")

        if nome is not None:
            nome = nome.strip()
            if len(nome) < 2 or len(nome) > 120:
                return jsonify("O nome deve ter entre 2 e 120 caracteres."), 400
            user.nome = nome

        if foto is not None:
            if foto and not validate_profile_photo(foto):
                return jsonify("A foto deve ser PNG, JPG ou WEBP e ter até 1,5 MB."), 400
            user.foto_perfil = foto or None

        if tema is not None:
            tema = str(tema).lower()
            if tema not in {"tmhub", "light", "dark", *CUSTOM_THEMES}:
                return jsonify("Tema visual inválido."), 400
            if tema in {"light", "dark"}:
                if modo_tema is None:
                    user.modo_tema = tema
                user.tema = "tmhub"
            else:
                if tema != "tmhub" and not can_use_theme(user, tema):
                    return jsonify("Este tema ainda não está liberado para sua conta."), 403
                user.tema = tema

        if modo_tema is not None:
            modo_tema = str(modo_tema).lower()
            if modo_tema not in {"light", "dark"}:
                return jsonify("O modo deve ser light ou dark."), 400
            user.modo_tema = modo_tema

        if has_particles_update:
            if not isinstance(particulas_ativas, bool):
                return jsonify("Informe um valor válido para as partículas visuais."), 400
            user.particulas_ativas = particulas_ativas

        if has_timo_update:
            if str(user.role or "").upper() != "ADMIN":
                return jsonify("A ativação do Timo está disponível apenas para administradores."), 403
            if not isinstance(timo_ativo, bool):
                return jsonify("Informe um valor válido para ativar ou desativar o Timo."), 400
            user.timo_ativo = timo_ativo

        if has_timo_home_update:
            if not isinstance(timo_tela_inicial, bool):
                return jsonify("Informe um valor válido para a tela inicial do Timo."), 400
            user.timo_tela_inicial = timo_tela_inicial

        if timo_cenario is not None:
            timo_cenario = str(timo_cenario).strip().lower()
            # Compatibilidade com a primeira versão do front, que usava
            # "cyberpunk" enquanto o produto do Marketplace sempre foi
            # catalogado como timo_cenario_cyber.
            if timo_cenario == "cyberpunk":
                timo_cenario = "cyber"
            base_scenarios = {"workshop", "orbit", "garden"}
            premium_scenarios = {"christmas", "halloween", "muertos", "cyber"}
            if timo_cenario not in base_scenarios | premium_scenarios:
                return jsonify("Cenário do Timo inválido."), 400
            if timo_cenario in premium_scenarios:
                owned = (
                    MarketplacePurchase.query
                    .join(MarketplaceProduct, MarketplaceProduct.id == MarketplacePurchase.produto_id)
                    .filter(
                        MarketplacePurchase.usuario_id == user.id,
                        MarketplacePurchase.status == "concluida",
                        MarketplaceProduct.codigo == f"timo_cenario_{timo_cenario}",
                        MarketplaceProduct.ativo.is_(True),
                    )
                    .first()
                )
                if not owned:
                    return jsonify("Adquira esse cenário no Marketplace antes de usá-lo."), 403
            user.timo_cenario = timo_cenario

        if nova_senha is not None:
            valid_password, _, _ = verify_password(str(senha_atual or ""), user.hash)
            if not valid_password:
                return jsonify("Senha atual incorreta."), 400
            if not is_strong_password(nova_senha):
                return jsonify("A senha deve ter ao menos 8 caracteres, com maiúscula, minúscula, número e caractere especial."), 400
            user.hash = hash_password(nova_senha)
            user.senha_padrao = is_default_password(nova_senha)
            user.troca_senha_obrigatoria = False
            user.token_version = int(user.token_version or 0) + 1
            user.senha_alterada_em = dt.now()

        has_profile_update = any(value is not None for value in (
            nome,
            foto,
            tema,
            modo_tema,
            nova_senha,
            timo_cenario,
        ))
        if not has_profile_update and not has_timo_update and not has_timo_home_update and not has_particles_update:
            return jsonify("Nenhuma alteração informada."), 400

        refresh_user_requirements(user)
        db.session.commit()
        if tema is not None or modo_tema is not None:
            # O mascote desktop não depende do browser, então recebe a paleta
            # logo após o usuário trocar o tema na própria conta.
            from models.timo_voice_agents import TimoUserPreference
            from utils.timo_voice_socket import emit_agent_control

            preference = db.session.get(TimoUserPreference, user.id)
            if preference and preference.agente_preferido_id:
                emit_agent_control(
                    preference.agente_preferido_id,
                    bool(preference.habilitado),
                )
        response = self._serialize(user)
        if nova_senha is not None:
            response["access_token"] = self._issue_token(user)
        return jsonify(response)

    @safe_route
    def request_email_code(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        body = rq.get_json(silent=True) or {}
        email = str(body.get("email", "")).strip().lower()
        if not user:
            return jsonify("Usuário não encontrado."), 404
        if not EMAIL_PATTERN.fullmatch(email):
            return jsonify("Informe um e-mail válido."), 400
        if Users.query.filter(Users.email == email, Users.id != user.id).first():
            return jsonify("Este e-mail já está em uso."), 409

        code = f"{secrets.randbelow(1_000_000):06d}"
        try:
            self._send_email_code(email, code)
        except (OSError, RuntimeError, smtplib.SMTPException):
            return jsonify("O serviço de e-mail não está disponível no momento."), 503
        user.email_pendente = email
        user.email_codigo_hash = sha256(code.encode()).hexdigest()
        user.email_codigo_expira_em = dt.now() + timedelta(minutes=10)
        db.session.commit()
        return jsonify("Código enviado. Ele expira em 10 minutos."), 200

    @safe_route
    def confirm_email(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        code = str((rq.get_json(silent=True) or {}).get("codigo", "")).strip()
        if not user:
            return jsonify("Usuário não encontrado."), 404
        if not user.email_pendente or not user.email_codigo_hash:
            return jsonify("Solicite um novo código de verificação."), 400
        if not user.email_codigo_expira_em or user.email_codigo_expira_em < dt.now():
            return jsonify("O código expirou. Solicite um novo."), 400
        if sha256(code.encode()).hexdigest() != user.email_codigo_hash:
            return jsonify("Código de verificação inválido."), 400

        user.email = user.email_pendente
        user.email_pendente = None
        user.email_codigo_hash = None
        user.email_codigo_expira_em = None
        db.session.commit()
        return jsonify(self._serialize(user))

    @safe_route
    def pending_requirements(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        if not user:
            return jsonify("Usuário não encontrado."), 404
        return jsonify({
            "requirements": auth_requirements(user),
            "cpf": user.cpf,
            "foto_perfil": user.foto_perfil,
        })

    @safe_route
    def complete_required_profile(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        if not user:
            return jsonify("Usuário não encontrado."), 404
        body = rq.get_json(silent=True) or {}
        cpf = normalize_cpf(body.get("cpf") if "cpf" in body else user.cpf)
        photo = body.get("foto_perfil") if "foto_perfil" in body else None
        if not is_valid_cpf(cpf):
            return jsonify("Informe um CPF válido."), 400
        if Users.query.filter(Users.cpf == cpf, Users.id != user.id).first():
            return jsonify("CPF já cadastrado para outro usuário."), 409
        if photo and not validate_profile_photo(photo):
            return jsonify("A foto deve ser PNG, JPG ou WEBP e ter até 1,5 MB."), 400

        user.cpf = cpf
        if photo:
            user.foto_perfil = photo
        refresh_user_requirements(user)
        db.session.commit()
        return jsonify({
            "requirements": auth_requirements(user),
            "cpf": user.cpf,
            "foto_perfil": user.foto_perfil,
        })

    @safe_route
    def change_required_password(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        if not user:
            return jsonify("Usuário não encontrado."), 404
        requirements = auth_requirements(user)
        if requirements["cpf_pendente"]:
            return jsonify("Conclua o cadastro do CPF primeiro."), 409

        new_password = str((rq.get_json(silent=True) or {}).get("nova_senha") or "")
        if not is_strong_password(new_password):
            return jsonify("A senha deve ter ao menos 8 caracteres, com maiúscula, minúscula, número e caractere especial."), 400
        if is_default_password(new_password):
            return jsonify("Escolha uma senha diferente da senha padrão."), 400
        same_password, _, _ = verify_password(new_password, user.hash)
        if same_password:
            return jsonify("A nova senha deve ser diferente da senha atual."), 400

        user.hash = hash_password(new_password)
        user.troca_senha_obrigatoria = False
        user.senha_padrao = False
        user.token_version = int(user.token_version or 0) + 1
        user.senha_alterada_em = dt.now()
        db.session.commit()
        token = self._issue_token(user)
        return jsonify({
            "access_token": token,
            "requirements": auth_requirements(user),
        })

    @safe_route
    def ignore_default_password(self, token_data):
        user = db.session.get(Users, token_data.get("id"))
        if not user:
            return jsonify("Usuário não encontrado."), 404
        requirements = auth_requirements(user)
        if requirements["cpf_pendente"]:
            return jsonify("Conclua o cadastro do CPF primeiro."), 409
        if not user.senha_padrao:
            return jsonify({"requirements": requirements})
        user.senha_padrao = False
        db.session.commit()
        return jsonify({"requirements": auth_requirements(user)})

    @staticmethod
    def _serialize(user):
        return {
            "id": user.id,
            "nome": user.nome,
            "email": user.email,
            "foto_perfil": user.foto_perfil,
            "tema": effective_theme_for(user),
            "modo_tema": user.modo_tema or "light",
            "particulas_ativas": bool(user.particulas_ativas),
            "temas_disponiveis": available_themes_for(user),
            "adorno_foto": user.adorno_foto,
            "timo_skin": user.timo_skin or "default",
            "timo_cenario": user.timo_cenario or "workshop",
            "timo_tela_inicial": bool(user.timo_tela_inicial),
            "role": user.role,
            "timo_ativo": bool(user.timo_ativo) if str(user.role or "").upper() == "ADMIN" else False,
            "gerencia_faltas": bool(user.gerencia_faltas),
            "filiais": [{"id": branch.id, "nome": branch.nome} for branch in sorted(user.filiais, key=lambda item: item.nome) if branch.ativa],
            "permissions": serialize_permissions(user),
            "requirements": auth_requirements(user),
        }

    @safe_route
    def permission_catalog(self, token_data):
        if not self._is_admin(token_data):
            return jsonify("Apenas administradores podem consultar o catálogo de permissões."), 403
        return jsonify(PERMISSION_CATALOG), 200

    @staticmethod
    def _send_email_code(recipient, code):
        host = getenv("SMTP_HOST")
        sender = getenv("SMTP_FROM") or getenv("SMTP_USER")
        if not host or not sender:
            raise RuntimeError("SMTP_HOST e SMTP_FROM (ou SMTP_USER) devem estar configurados.")

        message = EmailMessage()
        message["Subject"] = "Confirme a troca de e-mail - TM Hub"
        message["From"] = sender
        message["To"] = recipient
        message.set_content(f"Seu código de verificação é {code}. Ele expira em 10 minutos.")

        port = int(getenv("SMTP_PORT", "587"))
        with smtplib.SMTP(host, port, timeout=15) as smtp:
            if getenv("SMTP_STARTTLS", "true").lower() == "true":
                smtp.starttls()
            username = getenv("SMTP_USER") or sender
            password = getenv("SMTP_PASSWORD")
            if username and password:
                smtp.login(username, password)
            smtp.send_message(message)
